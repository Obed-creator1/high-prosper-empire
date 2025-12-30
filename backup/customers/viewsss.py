from rest_framework import viewsets, filters
from rest_framework import viewsets, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, NumberFilter, CharFilter

from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import csv
from rest_framework import serializers
from .models import Customer
from rest_framework.permissions import AllowAny, IsAuthenticated

class CustomerSerializer(serializers.ModelSerializer):
    class Meta:
        model = Customer
        fields = [
            "id", "type", "names", "contact_no", "account_number",
            "monthly_fee", "sector", "cell", "village", "status", "outstanding", "created_at"
        ]


# -------------------- CUSTOM FILTER --------------------
class CustomerFilter(FilterSet):
    sector = CharFilter(field_name="sector", lookup_expr="icontains")
    cell = CharFilter(field_name="cell", lookup_expr="icontains")
    village = CharFilter(field_name="village", lookup_expr="icontains")
    status = CharFilter(field_name="status", lookup_expr="iexact")
    min_outstanding = NumberFilter(field_name="outstanding", lookup_expr="gte")
    max_outstanding = NumberFilter(field_name="outstanding", lookup_expr="lte")

    class Meta:
        model = Customer
        fields = ["sector", "cell", "village", "status"]

# -------------------- FILTER OPTIONS --------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def customer_filter_options(request):
    # Get unique values dynamically from your Customer model
    statuses = Customer.objects.values_list('status', flat=True).distinct()
    sectors = Customer.objects.values_list('sector', flat=True).distinct()
    cells = Customer.objects.values_list('cell', flat=True).distinct()
    villages = Customer.objects.values_list('village', flat=True).distinct()  # ✅ Added

    # Clean out empty/null values
    data = {
        "statuses": [s for s in statuses if s],
        "sectors": [s for s in sectors if s],
        "cells": [c for c in cells if c],
        "villages": [v for v in villages if v],  # ✅ Added
    }

    return Response(data)

# -------------------- PAGINATION --------------------
class CustomerPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'

# -------------------- VIEWSET --------------------
class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all().order_by("-created_at")
    serializer_class = CustomerSerializer
    pagination_class = CustomerPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["names", "sector", "cell", "village", "contact_no", "account_number"]
    filterset_fields = ['status', 'sector', 'cell', 'village']
    ordering_fields = ["id", "names", "monthly_fee", "outstanding", "status", "sector", "cell", "village"]
    ordering = ["-created_at"]
    filterset_class = CustomerFilter
    permission_classes = [permissions.IsAuthenticated]

    # ---------------- CSV EXPORT ----------------
    @action(detail=False, methods=["get"], url_path="export_csv")
    def export_csv(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID','Type','Name','Contact','Account','Monthly Fee','Outstanding','Status','Sector','Cell','Village'])
        for c in queryset:
            writer.writerow([c.id,c.type,c.names,c.contact_no,c.account_number,c.monthly_fee,c.outstanding,c.status,c.sector,c.cell,c.village])
        return response

    # ---------------- PDF EXPORT ----------------
    @action(detail=False, methods=["get"], url_path="export_pdf")
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = [Paragraph("Customer Report", styles["Title"]), Spacer(1,12)]
        data = [['ID','Type','Name','Contact','Account','Fee','Outstanding','Status','Sector','Cell','Village']]
        for c in queryset:
            data.append([c.id,c.type,c.names,c.contact_no,c.account_number,str(c.monthly_fee),str(c.outstanding),c.status,c.sector,c.cell,c.village])
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.grey),
            ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('GRID',(0,0),(-1,-1),0.5,colors.black),
            ('FONTSIZE',(0,0),(-1,-1),8)
        ]))
        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="customers_report.pdf"'
        response.write(pdf)
        return response

    # ---------------- EXCEL EXPORT ----------------
    @action(detail=False, methods=["get"], url_path="export_excel")
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        headers = ['ID','Type','Name','Contact','Account','Monthly Fee','Outstanding','Status','Sector','Cell','Village']
        sheet.append(headers)
        header_font = Font(bold=True,color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
        for c in queryset:
            sheet.append([c.id,c.type,c.names,c.contact_no,c.account_number,c.monthly_fee,c.outstanding,c.status,c.sector,c.cell,c.village])
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            sheet.column_dimensions[column[0].column_letter].width = max_length+2
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="customers.xlsx"'
        workbook.save(response)
        return response

