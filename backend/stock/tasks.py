"""
Stock Management Celery Tasks
Comprehensive background processing for inventory operations
"""
import asyncio
import logging
import requests
import json
from celery.schedules import crontab
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional
from celery import shared_task, group, chord
from celery.exceptions import MaxRetriesExceededError
from django.db.models import Sum, F
from django.utils import timezone
from django.db import transaction
from django.core.cache import cache
from django.conf import settings
from django.core.mail import send_mail
from channels.layers import get_channel_layer
from asgiref.sync import sync_to_async, async_to_sync

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import (
    StockTransaction, WarehouseStock, WarehouseTransfer, StockAlert,
    StockBatch, StockReport, InventoryAudit, Stock, Warehouse
)
from .serializers import (
    StockTransactionSerializer, WarehouseStockSerializer,
    WarehouseTransferSerializer, StockAlertSerializer,
    WarehouseStockSerializer
)
from .consumers import redis_cache, broadcast_stock_transaction
from .utils import (
    calculate_safety_stock, forecast_demand, calculate_eoq,
    generate_barcode, validate_stock_data
)
from typing import Dict, Any, Optional

logger = logging.getLogger(__name__)

class ExternalSystemSyncManager:
    """Manager for synchronizing with external inventory systems"""

    def __init__(self):
        self.systems = {
            'erp': self._sync_erp,
            'wms': self._sync_wms,
            'pos': self._sync_pos,
            'woocommerce': self._sync_woocommerce,
            'shopify': self._sync_shopify,
            'magento': self._sync_magento,
            'custom_api': self._sync_custom_api
        }

    async def sync_all_systems(self, stock_item: WarehouseStock) -> Dict[str, Any]:
        """Synchronize with all configured external systems"""
        sync_results = {
            'stock_id': str(stock_item.stock_id),
            'sku': stock_item.stock.sku,
            'systems_synced': 0,
            'systems_failed': 0,
            'external_quantities': {},
            'final_quantity': stock_item.quantity,
            'discrepancies_found': False,
            'timestamp': timezone.now().isoformat()
        }

        configured_systems = settings.EXTERNAL_INVENTORY_SYSTEMS or []

        for system_name in configured_systems:
            if system_name in self.systems:
                try:
                    external_qty = await self.systems[system_name](stock_item)
                    sync_results['external_quantities'][system_name] = external_qty
                    sync_results['systems_synced'] += 1

                    # Check for discrepancies
                    if external_qty != stock_item.quantity:
                        sync_results['discrepancies_found'] = True
                        logger.warning(
                            f"Discrepancy detected for {stock_item.stock.sku}: "
                            f"Local={stock_item.quantity}, {system_name}={external_qty}"
                        )

                except Exception as e:
                    sync_results['systems_failed'] += 1
                    sync_results['external_quantities'][system_name] = {
                        'error': str(e),
                        'status': 'failed'
                    }
                    logger.error(f"Sync failed for {system_name} - {stock_item.stock.sku}: {e}")

        return sync_results

    async def _sync_erp(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with ERP system (SAP, Oracle, Microsoft Dynamics)"""
        try:
            erp_config = settings.ERP_CONFIG
            if not erp_config.get('enabled'):
                return None

            # SAP-like API endpoint
            url = f"{erp_config['base_url']}/api/inventory/{stock_item.stock.sku}"
            headers = {
                'Authorization': f"Bearer {erp_config['access_token']}",
                'Content-Type': 'application/json'
            }

            response = requests.get(
                url,
                headers=headers,
                timeout=10,
                params={'warehouse': stock_item.warehouse.code}
            )

            if response.status_code == 200:
                data = response.json()
                return int(data.get('available_quantity', 0))
            else:
                logger.error(f"ERP sync failed: {response.status_code} - {response.text}")
                return None

        except requests.exceptions.RequestException as e:
            logger.error(f"ERP request failed for {stock_item.stock.sku}: {e}")
            return None
        except (ValueError, KeyError) as e:
            logger.error(f"ERP data parsing failed for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_wms(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with Warehouse Management System (Manhattan, HighJump, etc.)"""
        try:
            wms_config = settings.WMS_CONFIG
            if not wms_config.get('enabled'):
                return None

            url = f"{wms_config['base_url']}/api/v1/stock/location"
            headers = {
                'Authorization': f"Basic {wms_config['api_key']}",
                'Content-Type': 'application/json'
            }

            payload = {
                'sku': stock_item.stock.sku,
                'warehouse_id': stock_item.warehouse.id,
                'location_type': 'available'
            }

            response = requests.post(
                url,
                headers=headers,
                json=payload,
                timeout=15
            )

            if response.status_code in [200, 201]:
                data = response.json()
                return int(data.get('quantity', 0))
            else:
                logger.error(f"WMS sync failed: {response.status_code}")
                return None

        except Exception as e:
            logger.error(f"WMS sync error for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_pos(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with Point of Sale systems (Square, Lightspeed, etc.)"""
        try:
            pos_config = settings.POS_CONFIG
            if not pos_config.get('enabled'):
                return None

            url = f"{pos_config['base_url']}/v1/items/{stock_item.stock.sku}/inventory"
            headers = {
                'Authorization': f"Bearer {pos_config['access_token']}",
                'Content-Type': 'application/json'
            }

            response = requests.get(
                url,
                headers=headers,
                timeout=10
            )

            if response.status_code == 200:
                data = response.json()
                # POS systems often track by location
                total_qty = 0
                for location in data.get('locations', []):
                    total_qty += int(location.get('quantity', 0))
                return total_qty
            else:
                logger.error(f"POS sync failed: {response.status_code}")
                return None

        except Exception as e:
            logger.error(f"POS sync error for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_woocommerce(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with WooCommerce"""
        try:
            wc_config = settings.WOOCOMMERCE_CONFIG
            if not wc_config.get('enabled'):
                return None

            # First get product ID by SKU
            products_url = f"{wc_config['base_url']}/wp-json/wc/v3/products"
            params = {
                'sku': stock_item.stock.sku,
                'consumer_key': wc_config['consumer_key'],
                'consumer_secret': wc_config['consumer_secret']
            }

            response = requests.get(products_url, params=params, timeout=15)
            if response.status_code != 200:
                return None

            products = response.json()
            if not products:
                return 0

            product_id = products[0]['id']
            stock_url = f"{wc_config['base_url']}/wp-json/wc/v3/products/{product_id}/stock"

            stock_response = requests.get(
                stock_url,
                params={
                    'consumer_key': wc_config['consumer_key'],
                    'consumer_secret': wc_config['consumer_secret']
                },
                timeout=10
            )

            if stock_response.status_code == 200:
                stock_data = stock_response.json()
                return int(stock_data.get('stock_quantity', 0))
            return None

        except Exception as e:
            logger.error(f"WooCommerce sync error for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_shopify(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with Shopify"""
        try:
            shopify_config = settings.SHOPIFY_CONFIG
            if not shopify_config.get('enabled'):
                return None

            url = f"{shopify_config['base_url']}/admin/api/2023-10/products.json"
            headers = {
                'X-Shopify-Access-Token': shopify_config['access_token'],
                'Content-Type': 'application/json'
            }
            params = {'limit': 1, 'ids': shopify_config.get('product_mapping', {}).get(stock_item.stock.sku)}

            response = requests.get(url, headers=headers, params=params, timeout=15)
            if response.status_code == 200:
                products = response.json().get('products', [])
                if products:
                    variants = products[0].get('variants', [])
                    total_qty = sum(v.get('inventory_quantity', 0) for v in variants)
                    return total_qty
            return None

        except Exception as e:
            logger.error(f"Shopify sync error for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_magento(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with Magento"""
        try:
            magento_config = settings.MAGENTO_CONFIG
            if not magento_config.get('enabled'):
                return None

            url = f"{magento_config['base_url']}/rest/V1/stockItems/{stock_item.stock.sku}"
            headers = {
                'Authorization': f"Bearer {magento_config['access_token']}"
            }

            response = requests.get(url, headers=headers, timeout=15)
            if response.status_code == 200:
                data = response.json()
                return int(data.get('qty', 0))
            return None

        except Exception as e:
            logger.error(f"Magento sync error for {stock_item.stock.sku}: {e}")
            return None

    async def _sync_custom_api(self, stock_item: WarehouseStock) -> Optional[int]:
        """Sync with custom external API"""
        try:
            custom_configs = getattr(settings, 'CUSTOM_INVENTORY_APIS', [])
            for config in custom_configs:
                if config.get('enabled'):
                    url = config['endpoint'].format(sku=stock_item.stock.sku)
                    headers = config.get('headers', {})

                    response = requests.get(url, headers=headers, timeout=10)
                    if response.status_code == 200:
                        data = response.json()
                        quantity_path = config.get('quantity_path', 'quantity')
                        quantity = self._extract_json_value(data, quantity_path)
                        if quantity is not None:
                            return int(quantity)
            return None
        except Exception as e:
            logger.error(f"Custom API sync error for {stock_item.stock.sku}: {e}")
            return None

    def _extract_json_value(self, data: Dict, path: str) -> Any:
        """Extract value from JSON using dot notation"""
        keys = path.split('.')
        current = data
        for key in keys:
            if isinstance(current, dict) and key in current:
                current = current[key]
            else:
                return None
        return current

# Global sync manager instance
sync_manager = ExternalSystemSyncManager()

def _full_stock_sync(stock_item: WarehouseStock) -> bool:
    """
    Perform comprehensive full stock synchronization across all systems

    Args:
        stock_item: WarehouseStock instance to sync

    Returns:
        bool: True if sync was successful, False otherwise
    """
    try:
        # Skip sync if stock is inactive or zero quantity
        if not stock_item.stock.is_active or stock_item.quantity == 0:
            return True

        # Check cache first (5 minute cache)
        cache_key = f"stock_sync_{stock_item.stock_id}_{stock_item.warehouse_id}"
        cached_results = cache.get(cache_key)
        if cached_results and (timezone.now() - cached_results['timestamp']).seconds < 300:
            return cached_results['sync_successful']

        # Perform async sync (but since we're in sync context, use sync_to_async)
        loop = asyncio.get_event_loop()
        sync_results = loop.run_until_complete(
            sync_manager.sync_all_systems(stock_item)
        )

        # Determine if update is needed
        update_performed = False

        # Calculate average external quantity (majority vote)
        external_quantities = [
            qty for qty in sync_results['external_quantities'].values()
            if isinstance(qty, int) and qty is not None
        ]

        if external_quantities:
            # Use median to avoid outliers
            median_external_qty = sorted(external_quantities)[len(external_quantities)//2]

            # Only update if significant discrepancy (>5% or minimum 5 units)
            discrepancy_threshold = max(5, stock_item.quantity * 0.05)
            if abs(median_external_qty - stock_item.quantity) > discrepancy_threshold:

                with transaction.atomic():
                    old_quantity = stock_item.quantity

                    # Update stock quantity
                    stock_item.quantity = median_external_qty
                    stock_item.last_updated = timezone.now()
                    stock_item.save(update_fields=['quantity', 'last_updated'])

                    # Create adjustment transaction
                    if old_quantity != median_external_qty:
                        StockTransaction.objects.create(
                            stock=stock_item.stock,
                            warehouse=stock_item.warehouse,
                            transaction_type='system_sync',
                            quantity=median_external_qty - old_quantity,
                            reference=f"sync_{sync_results['timestamp']}",
                            notes=f"Auto-sync with external systems: {len(external_quantities)} sources"
                        )

                        update_performed = True

                        # Broadcast real-time update
                        loop.run_until_complete(
                            _broadcast_stock_sync_update(
                                stock_item, old_quantity, median_external_qty, sync_results
                            )
                        )

        # Cache results
        cache_results = {
            'sync_successful': True,
            'systems_synced': sync_results['systems_synced'],
            'systems_failed': sync_results['systems_failed'],
            'discrepancies_found': sync_results['discrepancies_found'],
            'timestamp': timezone.now()
        }
        cache.set(cache_key, cache_results, 300)  # 5 minutes

        # Log sync results
        logger.info(
            f"Stock sync completed for {stock_item.stock.sku}: "
            f"synced={sync_results['systems_synced']}, "
            f"failed={sync_results['systems_failed']}, "
            f"updated={update_performed}, "
            f"discrepancies={sync_results['discrepancies_found']}"
        )

        return True

    except Exception as e:
        logger.error(f"Full stock sync failed for {stock_item.stock.sku}: {e}")

        # Cache failure to prevent repeated attempts
        cache_key = f"stock_sync_{stock_item.stock_id}_{stock_item.warehouse_id}"
        cache.set(cache_key, {'sync_successful': False, 'timestamp': timezone.now()}, 600)

        return False

async def _broadcast_stock_sync_update(stock_item: WarehouseStock,
                                       old_quantity: int,
                                       new_quantity: int,
                                       sync_results: Dict) -> None:
    """Broadcast stock sync update via WebSocket"""
    try:
        channel_layer = get_channel_layer()
        if channel_layer:
            await channel_layer.group_send(
                'stock_updates',
                {
                    'type': 'stock_level_changed',
                    'stock_id': str(stock_item.stock_id),
                    'warehouse_id': str(stock_item.warehouse_id),
                    'old_quantity': old_quantity,
                    'new_quantity': new_quantity,
                    'available_quantity': new_quantity,
                    'sync_results': sync_results,
                    'timestamp': timezone.now().isoformat()
                }
            )

            # Also send to specific stock group
            await channel_layer.group_send(
                f'stock_{stock_item.stock_id}',
                {
                    'type': 'stock_transaction_update',
                    'transaction': {
                        'stock_id': str(stock_item.stock_id),
                        'transaction_type': 'system_sync',
                        'quantity': new_quantity - old_quantity,
                        'reference': 'auto-sync'
                    },
                    'stock_id': str(stock_item.stock_id),
                    'timestamp': timezone.now().isoformat()
                }
            )

            # Cache invalidation
            await redis_cache.delete_cache(f"warehouse_stock_{stock_item.warehouse_id}_{stock_item.stock_id}")
            cache.delete(f"warehouse_stock_{stock_item.warehouse_id}_{stock_item.stock_id}")

    except Exception as e:
        logger.error(f"Failed to broadcast stock sync update: {e}")

# Additional sync helper functions
def _incremental_stock_sync(stock_item: WarehouseStock) -> bool:
    """Perform incremental sync (only changed items)"""
    # Check if item was modified recently
    if stock_item.last_updated > timezone.now() - timedelta(hours=1):
        return _full_stock_sync(stock_item)
    return True

def _external_stock_sync(stock_item: WarehouseStock) -> bool:
    """Sync only with external systems (skip internal validation)"""
    # Implementation similar to _full_stock_sync but focused on external only
    loop = asyncio.get_event_loop()
    sync_results = loop.run_until_complete(sync_manager.sync_all_systems(stock_item))
    return sync_results['systems_synced'] > 0


# =============================================================================
# 📦 BULK OPERATIONS TASKS
# =============================================================================

@shared_task(bind=True, max_retries=3)
def process_bulk_import(self, file_path: str, warehouse_id: str, user_id: int,
                        import_type: str = 'stock') -> Dict[str, Any]:
    """
    Process bulk import of stock data from CSV/Excel files
    """
    try:
        from django.db import transaction
        from .models import Warehouse

        warehouse = Warehouse.objects.get(id=warehouse_id)
        results = {
            'total_records': 0,
            'successful': 0,
            'failed': 0,
            'errors': [],
            'warnings': [],
            'task_id': self.request.id
        }

        # Read file based on extension
        import pandas as pd
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        results['total_records'] = len(df)

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # ✅ FIXED: Direct function calls (no await)
                    if import_type == 'stock':
                        success = _import_stock_item(row, warehouse, user_id)
                    elif import_type == 'transfer':
                        success = _import_transfer_item(row, warehouse, user_id)
                    elif import_type == 'adjustment':
                        success = _import_adjustment(row, warehouse, user_id)
                    else:
                        success = False

                    if success:
                        results['successful'] += 1
                    else:
                        results['failed'] += 1

                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append({
                        'row': index + 2,  # +2 for header and 0-index
                        'error': str(e)
                    })
                    logger.error(f"Bulk import error at row {index}: {e}")

        # ✅ FIXED: Sync cache operations
        from django.core.cache import cache
        cache_key = f"bulk_import_{self.request.id}"
        cache.set(cache_key, results, 3600)

        # ✅ FIXED: Use Celery tasks for async operations
        broadcast_bulk_import_complete.delay(self.request.id, results)
        send_bulk_import_notification.delay(user_id, results)

        return results

    except Warehouse.DoesNotExist:
        logger.error(f"Warehouse {warehouse_id} not found")
        return {'error': f'Warehouse {warehouse_id} not found', 'task_id': self.request.id}

    except Exception as e:
        logger.error(f"Bulk import task failed: {e}")
        if self.request.retries < self.max_retries:
            try:
                self.retry(countdown=60 * (2 ** self.request.retries))
            except MaxRetriesExceededError:
                broadcast_bulk_import_failed.delay(self.request.id, str(e))
        raise self.retry()

@sync_to_async
def _import_stock_item(row: pd.Series, warehouse: Warehouse, user_id: int) -> bool:
    """Import single stock item"""
    try:
        stock, created = Stock.objects.get_or_create(
            sku=row.get('sku', ''),
            defaults={
                'name': row.get('name', ''),
                'category_id': row.get('category_id'),
                'unit_price': Decimal(str(row.get('unit_price', 0))),
                'reorder_level': int(row.get('reorder_level', 0)),
                'min_stock_level': int(row.get('min_stock_level', 0)),
                'is_active': True
            }
        )

        WarehouseStock.objects.update_or_create(
            stock=stock,
            warehouse=warehouse,
            defaults={
                'quantity': int(row.get('quantity', 0)),
                'reserved_quantity': int(row.get('reserved_quantity', 0)),
                'unit_price': Decimal(str(row.get('unit_price', 0))),
                'last_updated': timezone.now()
            }
        )

        # Create transaction record
        if created:
            StockTransaction.objects.create(
                stock=stock,
                warehouse=warehouse,
                transaction_type='import',
                quantity=int(row.get('quantity', 0)),
                user_id=user_id,
                reference=f"bulk_import_{process_bulk_import.request.id}"
            )

            broadcast_stock_transaction(stock)

        return True

    except Exception as e:
        logger.error(f"Stock import error: {e}")
        return False

# =============================================================================
# 🔄 SYNCHRONIZATION TASKS
# =============================================================================

@shared_task(bind=True, max_retries=5)
def process_stock_sync(self, warehouse_ids: List[str] = None,
                       sync_type: str = 'full') -> Dict[str, Any]:
    """
    Synchronize stock levels across systems

    Args:
        warehouse_ids: Specific warehouses to sync (optional)
        sync_type: 'full', 'incremental', 'external'
    """
    try:
        sync_results = {
            'synced_items': 0,
            'updated_items': 0,
            'errors': [],
            'timestamp': timezone.now().isoformat(),
            'task_id': self.request.id
        }

        queryset = WarehouseStock.objects.all()
        if warehouse_ids:
            queryset = queryset.filter(warehouse_id__in=warehouse_ids)

        with transaction.atomic():
            for stock_item in queryset.iterator(chunk_size=1000):
                try:
                    if sync_type == 'full':
                        success = _full_stock_sync(stock_item)
                    elif sync_type == 'incremental':
                        success = _incremental_stock_sync(stock_item)
                    elif sync_type == 'external':
                        success = _external_stock_sync(stock_item)

                    if success:
                        sync_results['synced_items'] += 1
                    else:
                        sync_results['updated_items'] += 1

                except Exception as e:
                    sync_results['errors'].append({
                        'stock_id': stock_item.stock_id,
                        'error': str(e)
                    })

        # Update cache
        cache.set(f"stock_sync_{self.request.id}", sync_results, 1800)

        # Broadcast sync completion
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                'stock_updates',
                {
                    'type': 'stock_sync_completed',
                    'results': sync_results,
                    'timestamp': timezone.now().isoformat()
                }
            )

        return sync_results

    except Exception as e:
        logger.error(f"Stock sync failed: {e}")
        raise self.retry(countdown=300)


# =============================================================================
# 📊 ANALYSIS TASKS
# =============================================================================

@shared_task
def calculate_abc_analysis(warehouse_id: str = None,
                           period_days: int = 90) -> Dict[str, Any]:
    """
    Calculate ABC analysis for inventory optimization

    Returns:
        ABC classification results
    """
    try:
        end_date = timezone.now()
        start_date = end_date - timedelta(days=period_days)

        # Get transaction data
        transactions = StockTransaction.objects.filter(
            created_at__range=[start_date, end_date],
            transaction_type__in=['sale', 'issue', 'adjustment']
        )

        if warehouse_id:
            transactions = transactions.filter(warehouse_id=warehouse_id)

        # Calculate ABC metrics
        abc_results = []
        total_value = 0

        for transaction in transactions.values('stock_id').annotate(
                total_quantity=Sum('quantity'),
                total_value=Sum(F('quantity') * F('unit_price'))
        ).order_by('-total_value'):

            total_value += transaction['total_value'] or 0
            abc_results.append({
                'stock_id': transaction['stock_id'],
                'total_quantity': transaction['total_quantity'] or 0,
                'total_value': float(transaction['total_value'] or 0),
                'percentage': 0  # Will calculate later
            })

        # Calculate cumulative percentages and assign ABC category
        cumulative = 0
        for item in abc_results:
            cumulative += item['total_value']
            item['percentage'] = (cumulative / total_value) * 100 if total_value > 0 else 0

            if item['percentage'] <= 80:
                item['abc_category'] = 'A'
            elif item['percentage'] <= 95:
                item['abc_category'] = 'B'
            else:
                item['abc_category'] = 'C'

        # Cache results
        cache_key = f"abc_analysis_{warehouse_id or 'all'}_{period_days}"
        cache.set(cache_key, abc_results, 86400)  # 24 hours

        # Create report
        report = StockReport.objects.create(
            report_type='abc_analysis',
            warehouse_id=warehouse_id,
            generated_at=timezone.now(),
            data=json.dumps(abc_results),
            status='completed'
        )

        return {
            'report_id': report.id,
            'category_a_count': len([x for x in abc_results if x['abc_category'] == 'A']),
            'category_b_count': len([x for x in abc_results if x['abc_category'] == 'B']),
            'category_c_count': len([x for x in abc_results if x['abc_category'] == 'C']),
            'total_items': len(abc_results),
            'data': abc_results
        }

    except Exception as e:
        logger.error(f"ABC analysis failed: {e}")
        raise

@shared_task
def calculate_inventory_turnover(warehouse_id: str = None,
                                 period_months: int = 12) -> Dict[str, Any]:
    """Calculate inventory turnover ratios"""
    try:
        end_date = timezone.now()
        start_date = end_date - timedelta(days=30 * period_months)

        # COGS calculation
        cogs_transactions = StockTransaction.objects.filter(
            created_at__range=[start_date, end_date],
            transaction_type='sale'
        ).aggregate(
            total_cogs=Sum(F('quantity') * F('unit_price'))
        )

        total_cogs = cogs_transactions['total_cogs'] or 0

        # Average inventory value
        avg_inventory_transactions = WarehouseStock.objects.aggregate(
            avg_value=Sum(F('quantity') * F('unit_price')) / 2
        )

        avg_inventory = avg_inventory_transactions['avg_value'] or 0

        turnover_ratio = total_cogs / avg_inventory if avg_inventory > 0 else 0
        days_inventory_outstanding = 365 / turnover_ratio if turnover_ratio > 0 else 0

        results = {
            'turnover_ratio': float(turnover_ratio),
            'days_inventory_outstanding': float(days_inventory_outstanding),
            'total_cogs': float(total_cogs),
            'average_inventory': float(avg_inventory),
            'period_months': period_months
        }

        # Cache results
        cache.set(f"inventory_turnover_{warehouse_id or 'all'}", results, 86400)

        return results

    except Exception as e:
        logger.error(f"Inventory turnover calculation failed: {e}")
        raise

@shared_task(bind=True, max_retries=3)
def generate_monthly_report(self, year: int, month: int,
                            warehouse_id: Optional[str] = None) -> str:
    """
    Generate comprehensive monthly report
    """
    try:
        from django.utils import timezone
        from datetime import timedelta

        # Calculate date range
        start_date = timezone.datetime(year, month, 1)
        if month == 12:
            end_date = timezone.datetime(year + 1, 1, 1) - timedelta(seconds=1)
        else:
            end_date = timezone.datetime(year, month + 1, 1) - timedelta(seconds=1)

        # Collect report data
        report_data = {
            'period': f"{month:02d}/{year}",
            'warehouse': warehouse_id,
            'summary': _get_monthly_summary(start_date, end_date, warehouse_id),
            'top_movers': _get_top_moving_items(start_date, end_date, warehouse_id),
            'low_stock': _get_low_stock_items(warehouse_id),
            'stock_turnover': _get_monthly_turnover(start_date, end_date, warehouse_id),
            'transactions': _get_transaction_summary(start_date, end_date, warehouse_id)
        }

        # ✅ FIXED: Direct sync function call
        file_path = _create_monthly_excel_report(report_data)

        # Save to database
        from .models import StockReport
        import json
        StockReport.objects.create(
            report_type='monthly',
            warehouse_id=warehouse_id,
            year=year,
            month=month,
            file_path=file_path,
            status='completed',
            data=json.dumps(report_data)
        )

        # ✅ FIXED: Sync cache
        from django.core.cache import cache
        cache_key = f"monthly_report_{year}_{month:02d}_{warehouse_id or 'all'}"
        cache.set(cache_key, {'file_path': file_path, 'data': report_data}, 86400 * 30)

        return file_path

    except Exception as e:
        logger.error(f"Monthly report generation failed: {e}")
        if self.request.retries < self.max_retries:
            self.retry(countdown=300)
        raise

@shared_task(bind=True, max_retries=3)
def generate_yearly_report(self, year: int,
                           warehouse_id: Optional[str] = None) -> str:
    """
    Generate comprehensive yearly stock report
    """
    try:
        from django.utils import timezone
        from datetime import timedelta

        start_date = timezone.datetime(year, 1, 1)
        end_date = timezone.datetime(year + 1, 1, 1) - timedelta(seconds=1)

        report_data = {
            'period': str(year),
            'warehouse': warehouse_id,
            'annual_summary': _get_yearly_summary(start_date, end_date, warehouse_id),
            'monthly_trends': _get_monthly_trends(year, warehouse_id),
            'abc_analysis': calculate_abc_analysis(warehouse_id, 365).get('data', []),
            'inventory_valuation': _get_yearly_inventory_valuation(start_date, end_date, warehouse_id)
        }

        # ✅ FIXED: Direct sync function call
        file_path = _create_yearly_excel_report(report_data)

        from .models import StockReport
        import json
        StockReport.objects.create(
            report_type='yearly',
            warehouse_id=warehouse_id,
            year=year,
            file_path=file_path,
            status='completed',
            data=json.dumps(report_data)
        )

        return file_path

    except Exception as e:
        logger.error(f"Yearly report generation failed: {e}")
        if self.request.retries < self.max_retries:
            self.retry(countdown=600)
        raise
# =============================================================================
# 🔍 AUDIT AND VALIDATION TASKS
# =============================================================================

@shared_task
def perform_inventory_audit(warehouse_id: str, audit_type: str = 'cycle') -> Dict[str, Any]:
    """
    Perform automated inventory audit

    Args:
        warehouse_id: Warehouse to audit
        audit_type: 'full', 'cycle', 'spot'
    """
    try:
        audit = InventoryAudit.objects.create(
            warehouse_id=warehouse_id,
            audit_type=audit_type,
            started_at=timezone.now(),
            status='in_progress'
        )

        discrepancies = []

        if audit_type == 'full':
            discrepancies = _perform_full_audit(warehouse_id)
        elif audit_type == 'cycle':
            discrepancies = _perform_cycle_audit(warehouse_id)
        elif audit_type == 'spot':
            discrepancies = _perform_spot_audit(warehouse_id)

        # Update audit record
        total_discrepancies = len(discrepancies)
        total_value = sum(abs(d['discrepancy_value']) for d in discrepancies)

        audit.status = 'completed'
        audit.total_discrepancies = total_discrepancies
        audit.total_value = total_value
        audit.completed_at = timezone.now()
        audit.save()

        # Create alerts for significant discrepancies
        for discrepancy in discrepancies:
            if abs(discrepancy['discrepancy_value']) > 100:  # Threshold
                StockAlert.objects.create(
                    stock_id=discrepancy['stock_id'],
                    alert_type='audit_discrepancy',
                    severity='high',
                    message=f"Audit discrepancy: {discrepancy['discrepancy_quantity']} units",
                    value=discrepancy['discrepancy_value']
                )

        # Broadcast audit results
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                'stock_notifications',
                {
                    'type': 'audit_discrepancy_alert',
                    'audit_id': str(audit.id),
                    'total_discrepancies': total_discrepancies,
                    'total_value': float(total_value),
                    'warehouse': warehouse_id,
                    'timestamp': timezone.now().isoformat()
                }
            )

        return {
            'audit_id': audit.id,
            'total_discrepancies': total_discrepancies,
            'total_value': float(total_value),
            'discrepancies': discrepancies
        }

    except Exception as e:
        logger.error(f"Inventory audit failed: {e}")
        raise

# =============================================================================
# 🔔 NOTIFICATION AND ALERT TASKS
# =============================================================================

@shared_task(schedule=crontab(minute=0, hour=2))  # Daily at 2 AM
def check_stock_alerts():
    """Check for low stock, expiry, and other alerts - ✅ FIXED: No await in sync context"""
    try:
        from .models import WarehouseStock, StockAlert, StockBatch
        from .serializers import StockAlertSerializer

        # Low stock alerts
        low_stock_items = WarehouseStock.objects.filter(
            quantity__lte=F('stock__reorder_level'),
            stock__is_active=True
        ).select_related('stock', 'warehouse')

        created_alerts = []
        for item in low_stock_items:
            alert, created = StockAlert.objects.get_or_create(
                stock=item.stock,
                alert_type='low_stock',
                warehouse=item.warehouse,
                defaults={
                    'severity': 'high',
                    'message': f'Low stock: {item.stock.name} ({item.quantity} <= {item.stock.reorder_level})',
                    'is_active': True
                }
            )
            if created:
                created_alerts.append(alert)

        # Expiring batch alerts
        expiring_batches = StockBatch.objects.filter(
            expiry_date__lte=timezone.now() + timedelta(days=30),
            expiry_date__gte=timezone.now()
        )

        for batch in expiring_batches:
            days_to_expiry = (batch.expiry_date - timezone.now()).days
            severity = 'high' if days_to_expiry <= 7 else 'medium'

            alert, created = StockAlert.objects.get_or_create(
                stock=batch.stock,
                alert_type='batch_expiry',
                defaults={
                    'severity': severity,
                    'message': f'Batch {batch.batch_number} expires in {days_to_expiry} days',
                    'value': days_to_expiry,
                    'is_active': True
                }
            )
            if created:
                created_alerts.append(alert)

        # ✅ FIXED: Broadcast alerts using async_to_sync
        active_alerts = StockAlert.objects.filter(is_active=True).order_by('-created_at')[:50]
        if active_alerts.exists():
            channel_layer = get_channel_layer()
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    'stock_notifications',
                    {
                        'type': 'critical_stock_alert',
                        'alerts': StockAlertSerializer(active_alerts, many=True).data,
                        'count': active_alerts.count(),
                        'timestamp': timezone.now().isoformat()
                    }
                )

        # Cache alert summary
        cache_key = "stock_alerts_summary"
        cache.set(cache_key, {
            'total_alerts': StockAlert.objects.filter(is_active=True).count(),
            'low_stock': StockAlert.objects.filter(alert_type='low_stock', is_active=True).count(),
            'expiring': StockAlert.objects.filter(alert_type='batch_expiry', is_active=True).count(),
            'timestamp': timezone.now().isoformat()
        }, 300)  # 5 minutes

        logger.info(f"Stock alerts check completed. Created {len(created_alerts)} new alerts")
        return {
            'status': 'success',
            'new_alerts': len(created_alerts),
            'total_alerts': StockAlert.objects.filter(is_active=True).count()
        }

    except Exception as e:
        logger.error(f"Stock alerts check failed: {e}")
        return {'status': 'error', 'message': str(e)}

# =============================================================================
# 📊 DASHBOARD AND METRICS TASKS
# =============================================================================

@shared_task(period=60)  # Every minute
def update_dashboard_metrics():
    """Update real-time dashboard metrics"""
    try:
        metrics = {
            'total_stock_value': WarehouseStock.objects.aggregate(
                total=Sum(F('quantity') * F('unit_price'))
            )['total'] or 0,
            'critical_stock_items': WarehouseStock.objects.filter(
                quantity__lte=F('stock__reorder_level')
            ).count(),
            'active_transfers': WarehouseTransfer.objects.filter(
                status__in=['pending', 'in_transit']
            ).count(),
            'today_transactions': StockTransaction.objects.filter(
                created_at__date=timezone.now().date()
            ).count(),
            'timestamp': timezone.now().isoformat()
        }

        # Cache metrics
        cache.set('dashboard_metrics_global', metrics, 60)

        # Broadcast to dashboard consumers
        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                'dashboard_updates',
                {
                    'type': 'dashboard_metrics_update',
                    'metrics': metrics,
                    'timestamp': metrics['timestamp']
                }
            )

    except Exception as e:
        logger.error(f"Dashboard metrics update failed: {e}")

# =============================================================================
# 🔄 BACKGROUND PROCESSING TASKS
# =============================================================================

@shared_task
def process_pending_transfers():
    """Process pending warehouse transfers"""
    try:
        pending_transfers = WarehouseTransfer.objects.filter(
            status='pending',
            scheduled_date__lte=timezone.now()
        )

        for transfer in pending_transfers:
            with transaction.atomic():
                # Start transfer
                transfer.status = 'in_transit'
                transfer.started_at = timezone.now()
                transfer.save()

                # Broadcast update
                channel_layer = get_channel_layer()
                if channel_layer:
                    async_to_sync(channel_layer.group_send)(
                        f'transfer_{transfer.id}',
                        {
                            'type': 'transfer_status_update',
                            'transfer_id': str(transfer.id),
                            'new_status': 'in_transit',
                            'timestamp': timezone.now().isoformat()
                        }
                    )

                logger.info(f"Started transfer {transfer.id}")

    except Exception as e:
        logger.error(f"Pending transfers processing failed: {e}")

# =============================================================================
# 📁 EXCEL REPORT GENERATION HELPERS
# =============================================================================

@sync_to_async
def _create_monthly_excel_report(report_data: Dict) -> str:
    """Create formatted monthly Excel report"""
    file_path = f"reports/monthly_report_{report_data['period'].replace('/', '_')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Stock Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Add report sections
    row = 1

    # Title
    ws[f'A{row}'] = f"Monthly Stock Report - {report_data['period']}"
    ws[f'A{row}'].font = Font(size=16, bold=True)
    row += 2

    # Summary section
    ws[f'A{row}'] = "EXECUTIVE SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    summary_data = [
        ['Total Transactions', report_data['summary']['total_transactions']],
        ['Inbound Quantity', report_data['summary']['total_inbound']],
        ['Outbound Quantity', report_data['summary']['total_outbound']],
        ['Net Movement', report_data['summary']['net_movement']]
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=row):
        ws[f'A{r_idx}'] = label
        ws[f'B{r_idx}'] = value
        ws[f'A{r_idx}'].font = Font(bold=True)

    row += len(summary_data) + 2

    # Top movers
    ws[f'A{row}'] = "TOP 10 MOVING ITEMS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    # Add more sections...

    wb.save(file_path)
    return file_path

# =============================================================================
# 🚀 HIGH-PERFORMANCE BATCH TASKS
# =============================================================================

@shared_task
def process_high_volume_stock_update(stock_ids: List[str],
                                     quantity_changes: Dict[str, int]):
    """Process high-volume stock updates using bulk operations"""
    try:
        with transaction.atomic():
            warehouse_stocks = WarehouseStock.objects.filter(
                stock_id__in=stock_ids
            ).select_for_update()

            updated_stocks = []
            for ws in warehouse_stocks:
                if ws.stock_id in quantity_changes:
                    old_quantity = ws.quantity
                    ws.quantity += quantity_changes[ws.stock_id]
                    ws.last_updated = timezone.now()
                    ws.save()

                    # Create transaction
                    StockTransaction.objects.create(
                        stock=ws.stock,
                        warehouse=ws.warehouse,
                        transaction_type='bulk_adjustment',
                        quantity=quantity_changes[ws.stock_id],
                        reference='high_volume_update'
                    )

                    updated_stocks.append({
                        'stock_id': str(ws.stock_id),
                        'old_quantity': old_quantity,
                        'new_quantity': ws.quantity,
                        'change': quantity_changes[ws.stock_id]
                    })

            # Broadcast bulk update
            channel_layer = get_channel_layer()
            if channel_layer and updated_stocks:
                async_to_sync(channel_layer.group_send)(
                    'stock_updates',
                    {
                        'type': 'bulk_stock_update',
                        'updates': updated_stocks,
                        'timestamp': timezone.now().isoformat()
                    }
                )

        return {'updated_items': len(updated_stocks)}

    except Exception as e:
        logger.error(f"High volume stock update failed: {e}")
        raise

# =============================================================================
# 📧 NOTIFICATION HELPERS
# =============================================================================

async def _broadcast_bulk_import_complete(task_id: str, results: Dict):
    """Broadcast bulk import completion"""
    channel_layer = get_channel_layer()
    if channel_layer:
        await channel_layer.group_send(
            'stock_reports',
            {
                'type': 'reports_update',
                'report_id': task_id,
                'status': 'completed',
                'progress': 100,
                'results': results,
                'timestamp': timezone.now().isoformat()
            }
        )

async def _broadcast_bulk_import_failed(task_id: str, error: str):
    """Broadcast bulk import failure"""
    channel_layer = get_channel_layer()
    if channel_layer:
        await channel_layer.group_send(
            'stock_reports',
            {
                'type': 'reports_update',
                'report_id': task_id,
                'status': 'failed',
                'error': error,
                'timestamp': timezone.now().isoformat()
            }
        )

def _send_bulk_import_notification(user_id: int, results: Dict):
    """Send email notification for bulk import completion"""
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        user = User.objects.get(id=user_id)

        subject = f"Bulk Import Completed - {results['successful']} Successful"
        message = f"""
        Your bulk import has completed:
        
        Total Records: {results['total_records']}
        Successful: {results['successful']}
        Failed: {results['failed']}
        
        Task ID: {results['task_id']}
        """

        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
    except Exception as e:
        logger.error(f"Failed to send bulk import notification: {e}")

# =============================================================================
# 🛠️ UTILITY FUNCTIONS
# =============================================================================

def _perform_full_audit(warehouse_id: str) -> List[Dict]:
    """Perform full physical inventory audit"""
    discrepancies = []

    warehouse_stocks = WarehouseStock.objects.filter(warehouse_id=warehouse_id)

    for ws in warehouse_stocks:
        # In real implementation, this would compare with physical count
        # For demo, generate random discrepancies
        expected_discrepancy = np.random.normal(0, ws.quantity * 0.05)
        discrepancy_qty = round(expected_discrepancy)

        if discrepancy_qty != 0:
            discrepancy_value = discrepancy_qty * ws.unit_price
            discrepancies.append({
                'stock_id': str(ws.stock_id),
                'stock_name': ws.stock.name,
                'system_quantity': ws.quantity,
                'physical_quantity': ws.quantity - discrepancy_qty,
                'discrepancy_quantity': discrepancy_qty,
                'discrepancy_value': float(discrepancy_value)
            })

    return discrepancies

# Celery Canvas Workflows
monthly_report_workflow = chord(
    (
        generate_monthly_report.s(2024, 11),
        generate_monthly_report.s(2024, 10),
        generate_monthly_report.s(2024, 9),
    ),
    calculate_abc_analysis.s()
)

# Export for use in views
__all__ = [
    'process_bulk_import',
    'process_stock_sync',
    'calculate_abc_analysis',
    'generate_monthly_report',
    'generate_yearly_report',
    'perform_inventory_audit',
    'check_stock_alerts',
    'update_dashboard_metrics',
    'process_high_volume_stock_update',
    'process_pending_transfers'
]

# =============================================================================
# 🚀 ASYNC BROADCAST TASKS
# =============================================================================

@shared_task
def broadcast_bulk_import_complete(task_id: str, results: Dict):
    """Broadcast bulk import completion via WebSocket"""
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        'bulk_import_updates',
        {
            'type': 'bulk_import.complete',
            'task_id': task_id,
            'results': results
        }
    )

@shared_task
def broadcast_bulk_import_failed(task_id: str, error: str):
    """Broadcast bulk import failure via WebSocket"""
    from channels.layers import get_channel_layer
    from asgiref.sync import async_to_sync
    channel_layer = get_channel_layer()

    async_to_sync(channel_layer.group_send)(
        'bulk_import_updates',
        {
            'type': 'bulk_import.failed',
            'task_id': task_id,
            'error': error
        }
    )

@shared_task
def send_bulk_import_notification(user_id: int, results: Dict):
    """Send email notification for bulk import completion"""
    from django.core.mail import send_mail
    from django.conf import settings

    subject = f'Bulk Import Completed - {results["successful"]} Successful'
    message = f"""
    Your bulk import task has completed:
    - Total Records: {results['total_records']}
    - Successful: {results['successful']}
    - Failed: {results['failed']}
    - Task ID: {results['task_id']}
    """

    send_mail(
        subject=subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[f'user_{user_id}@example.com'],  # Replace with actual user email
        fail_silently=False,
    )

def _import_stock_item(row, warehouse, user_id):
    """Import single stock item - SYNC ONLY"""
    try:
        from .models import Stock, WarehouseStock, StockTransaction
        from django.contrib.auth import get_user_model

        # Your stock import logic here
        # Example:
        stock, created = Stock.objects.get_or_create(
            sku=row.get('sku', ''),
            defaults={
                'name': row.get('name', ''),
                'unit_price': float(row.get('unit_price', 0)),
                'reorder_level': int(row.get('reorder_level', 0)),
                'is_active': True
            }
        )

        if created:
            WarehouseStock.objects.create(
                warehouse=warehouse,
                stock=stock,
                quantity=int(row.get('quantity', 0)),
                unit_price=float(row.get('unit_price', 0))
            )

            # Create transaction
            StockTransaction.objects.create(
                stock=stock,
                warehouse=warehouse,
                transaction_type='import',
                quantity=int(row.get('quantity', 0)),
                unit_price=float(row.get('unit_price', 0)),
                created_by=get_user_model().objects.get(id=user_id)
            )

        return True
    except Exception as e:
        logger.error(f"Stock import error: {e}")
        return False

def _create_monthly_excel_report(report_data):
    """Create monthly Excel report - SYNC ONLY"""
    import pandas as pd
    from django.conf import settings
    import os

    file_path = os.path.join(settings.MEDIA_ROOT, f'monthly_report_{report_data["period"]}.xlsx')

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Write different sheets
        pd.DataFrame([report_data['summary']]).to_excel(writer, sheet_name='Summary', index=False)
        pd.DataFrame(report_data['top_movers']).to_excel(writer, sheet_name='Top Movers', index=False)
        pd.DataFrame(report_data['low_stock']).to_excel(writer, sheet_name='Low Stock', index=False)

    return file_path

def _create_yearly_excel_report(report_data):
    """Create yearly Excel report - SYNC ONLY"""
    import pandas as pd
    from django.conf import settings
    import os

    file_path = os.path.join(settings.MEDIA_ROOT, f'yearly_report_{report_data["period"]}.xlsx')

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        pd.DataFrame([report_data['annual_summary']]).to_excel(writer, sheet_name='Annual Summary', index=False)
        pd.DataFrame(report_data['monthly_trends']).to_excel(writer, sheet_name='Monthly Trends', index=False)

    return file_path

# =============================================================================
# 🧹 CLEANUP AND MAINTENANCE TASKS
# =============================================================================

@shared_task
def cleanup_expired_cache():
    """Clean up expired cache entries"""
    try:
        from django.core.cache import cache
        # Django cache doesn't have direct cleanup, but we can set TTLs
        logger.info("Cache cleanup task executed")
        return {'status': 'success', 'message': 'Cache cleanup completed'}
    except Exception as e:
        logger.error(f"Cache cleanup failed: {e}")
        return {'status': 'error', 'message': str(e)}

@shared_task
def update_stock_analytics():
    """Update daily stock analytics"""
    try:
        from .models import StockAnalytics
        from django.utils import timezone

        today = timezone.now().date()
        StockAnalytics.objects.update_or_create(
            date=today,
            defaults={
                'total_items': Stock.objects.filter(is_active=True).count(),
                'total_value': WarehouseStock.objects.aggregate(
                    total=Sum(F('quantity') * F('unit_price'))
                )['total'] or 0,
                'critical_stock': WarehouseStock.objects.filter(
                    quantity__lte=F('stock__reorder_level')
                ).count()
            }
        )

        logger.info(f"Stock analytics updated for {today}")
        return {'status': 'success', 'date': str(today)}
    except Exception as e:
        logger.error(f"Stock analytics update failed: {e}")
        return {'status': 'error', 'message': str(e)}

@shared_task
def process_bulk_transfer(file_path, from_warehouse_id, to_warehouse_id, user_id):
    """Process bulk transfer from CSV/Excel file"""
    try:
        from .models import Warehouse, WarehouseTransfer, WarehouseTransferItem, WarehouseStock, StockTransaction

        from_warehouse = Warehouse.objects.get(id=from_warehouse_id)
        to_warehouse = Warehouse.objects.get(id=to_warehouse_id)

        # Read file
        import pandas as pd
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        results = {
            'total_items': 0,
            'successful': 0,
            'failed': 0,
            'errors': [],
            'transfer_id': None
        }

        with transaction.atomic():
            # Create main transfer record
            transfer = WarehouseTransfer.objects.create(
                from_warehouse=from_warehouse,
                to_warehouse=to_warehouse,
                transfer_type='bulk',
                status='pending',
                created_by_id=user_id,
                total_items=len(df)
            )
            results['transfer_id'] = transfer.id

            for index, row in df.iterrows():
                try:
                    stock_sku = row.get('sku')
                    quantity = int(row.get('quantity', 0))

                    if not stock_sku or quantity <= 0:
                        raise ValueError("Invalid SKU or quantity")

                    # Find stock in source warehouse
                    source_stock = WarehouseStock.objects.filter(
                        warehouse=from_warehouse,
                        stock__sku=stock_sku,
                        quantity__gte=quantity
                    ).first()

                    if not source_stock:
                        raise ValueError(f"Insufficient stock for {stock_sku}")

                    # Create transfer item
                    WarehouseTransferItem.objects.create(
                        transfer=transfer,
                        stock=source_stock.stock,
                        quantity=quantity,
                        unit_price=source_stock.unit_price
                    )

                    # Update quantities (will be finalized when transfer completes)
                    source_stock.quantity -= quantity
                    source_stock.reserved_quantity += quantity
                    source_stock.save()

                    results['successful'] += 1

                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append({
                        'row': index + 2,
                        'sku': row.get('sku'),
                        'error': str(e)
                    })

            # Update transfer status
            transfer.status = 'completed' if results['failed'] == 0 else 'partial'
            transfer.save()

        # Broadcast completion
        broadcast_transfer_update.delay(transfer.id, 'completed', results)

        return results

    except Exception as e:
        logger.error(f"Bulk transfer task failed: {e}")
        return {'status': 'failed', 'error': str(e)}

# =============================================================================
# 📡 REAL-TIME BROADCAST TASKS
# =============================================================================

@shared_task(bind=True, max_retries=3)
def broadcast_transfer_update(self, transfer_id: str, status: str, results: Dict = None):
    """
    Broadcast transfer updates to WebSocket clients

    Args:
        transfer_id: ID of the transfer
        status: New status ('pending', 'completed', 'failed', 'partial')
        results: Optional results dict for completed transfers
    """
    try:
        from .models import WarehouseTransfer
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        # Get transfer details
        transfer = WarehouseTransfer.objects.select_related(
            'from_warehouse', 'to_warehouse', 'created_by'
        ).get(id=transfer_id)

        # Prepare broadcast message
        message = {
            'type': 'transfer_update',
            'transfer_id': str(transfer.id),
            'status': status,
            'from_warehouse': {
                'id': transfer.from_warehouse.id,
                'name': transfer.from_warehouse.name
            },
            'to_warehouse': {
                'id': transfer.to_warehouse.id,
                'name': transfer.to_warehouse.name
            },
            'total_items': transfer.total_items,
            'completed_items': transfer.completed_items,
            'timestamp': timezone.now().isoformat()
        }

        # Add results for completed transfers
        if results and status in ['completed', 'partial', 'failed']:
            message.update({
                'results': results,
                'summary': {
                    'successful': results.get('successful', 0),
                    'failed': results.get('failed', 0),
                    'total': results.get('total_items', 0)
                }
            })

        # Broadcast to multiple groups
        channel_layer = get_channel_layer()
        if channel_layer:
            # Global transfer updates
            async_to_sync(channel_layer.group_send)(
                'transfer_updates',
                message
            )

            # Specific transfer group
            async_to_sync(channel_layer.group_send)(
                f'transfer_{transfer_id}',
                message
            )

            # Warehouse groups
            async_to_sync(channel_layer.group_send)(
                f'warehouse_{transfer.from_warehouse.id}',
                message
            )

            async_to_sync(channel_layer.group_send)(
                f'warehouse_{transfer.to_warehouse.id}',
                message
            )

            # User-specific groups for involved users
            if transfer.created_by:
                async_to_sync(channel_layer.group_send)(
                    f'user_{transfer.created_by.id}',
                    message
                )

            logger.info(f"Transfer update broadcasted: {transfer_id} -> {status}")

        # Cache transfer status
        cache_key = f"transfer_status_{transfer_id}"
        cache.set(cache_key, message, 3600)

        return {
            'status': 'success',
            'transfer_id': transfer_id,
            'broadcasted_to': ['transfer_updates', f'transfer_{transfer_id}', 'warehouses', 'users'],
            'timestamp': timezone.now().isoformat()
        }

    except WarehouseTransfer.DoesNotExist:
        logger.error(f"Transfer {transfer_id} not found for broadcast")
        return {'status': 'error', 'message': 'Transfer not found'}
    except Exception as e:
        logger.error(f"Broadcast transfer update failed: {e}")
        if self.request.retries < self.max_retries:
            self.retry(countdown=10 * (2 ** self.request.retries))
        raise self.retry()

@shared_task
def broadcast_stock_update(stock_id: str, warehouse_id: str, quantity_change: int,
                           new_quantity: int, user_id: int = None):
    """
    Broadcast stock quantity updates to WebSocket clients
    """
    try:
        from .models import Stock, WarehouseStock
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        stock = Stock.objects.get(id=stock_id)
        message = {
            'type': 'stock_update',
            'stock_id': stock_id,
            'stock_sku': stock.sku,
            'stock_name': stock.name,
            'warehouse_id': warehouse_id,
            'quantity_change': quantity_change,
            'new_quantity': new_quantity,
            'available_quantity': new_quantity,
            'is_critical': new_quantity <= stock.reorder_level,
            'timestamp': timezone.now().isoformat()
        }

        channel_layer = get_channel_layer()
        if channel_layer:
            # Global stock updates
            async_to_sync(channel_layer.group_send)(
                'stock_updates',
                message
            )

            # Specific stock group
            async_to_sync(channel_layer.group_send)(
                f'stock_{stock_id}',
                message
            )

            # Warehouse group
            async_to_sync(channel_layer.group_send)(
                f'warehouse_{warehouse_id}',
                message
            )

            # User-specific (if provided)
            if user_id:
                async_to_sync(channel_layer.group_send)(
                    f'user_{user_id}',
                    message
                )

        # Cache stock status
        cache_key = f"stock_status_{stock_id}_{warehouse_id}"
        cache.set(cache_key, message, 300)

        logger.info(f"Stock update broadcasted: {stock_id} in warehouse {warehouse_id}")
        return {'status': 'success', 'stock_id': stock_id}

    except Stock.DoesNotExist:
        logger.error(f"Stock {stock_id} not found for broadcast")
        return {'status': 'error', 'message': 'Stock not found'}
    except Exception as e:
        logger.error(f"Broadcast stock update failed: {e}")
        return {'status': 'error', 'message': str(e)}

@shared_task
def broadcast_stock_alert(alert_id: str, severity: str = 'medium'):
    """
    Broadcast new stock alerts to WebSocket clients
    """
    try:
        from .models import StockAlert
        from .serializers import StockAlertSerializer
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        alert = StockAlert.objects.get(id=alert_id)
        alert_data = StockAlertSerializer(alert).data

        message = {
            'type': 'alert_update',
            'alert_id': alert_id,
            'severity': severity,
            'alert_type': alert.alert_type,
            'message': alert.message,
            'stock_id': str(alert.stock.id) if alert.stock else None,
            'warehouse_id': str(alert.warehouse.id) if alert.warehouse else None,
            'timestamp': timezone.now().isoformat()
        }

        channel_layer = get_channel_layer()
        if channel_layer:
            # Global alerts
            async_to_sync(channel_layer.group_send)(
                'stock_notifications',
                message
            )

            # Critical alerts (high severity)
            if severity == 'high':
                async_to_sync(channel_layer.group_send)(
                    'critical_alerts',
                    message
                )

            # Specific stock alerts
            if alert.stock:
                async_to_sync(channel_layer.group_send)(
                    f'stock_{alert.stock.id}',
                    message
                )

            # Warehouse alerts
            if alert.warehouse:
                async_to_sync(channel_layer.group_send)(
                    f'warehouse_{alert.warehouse.id}',
                    message
                )

        # Cache alert
        cache_key = f"alert_{alert_id}"
        cache.set(cache_key, message, 3600)

        logger.info(f"Stock alert broadcasted: {alert_id} ({severity})")
        return {'status': 'success', 'alert_id': alert_id}

    except StockAlert.DoesNotExist:
        logger.error(f"Alert {alert_id} not found for broadcast")
        return {'status': 'error', 'message': 'Alert not found'}
    except Exception as e:
        logger.error(f"Broadcast stock alert failed: {e}")
        return {'status': 'error', 'message': str(e)}

@shared_task
def broadcast_bulk_import_complete(task_id: str, results: Dict):
    """Broadcast bulk import completion"""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        message = {
            'type': 'bulk_import.complete',
            'task_id': task_id,
            'results': results,
            'summary': {
                'successful': results.get('successful', 0),
                'failed': results.get('failed', 0),
                'total': results.get('total_records', 0)
            },
            'timestamp': timezone.now().isoformat()
        }

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                'bulk_import_updates',
                message
            )

        logger.info(f"Bulk import completed: {task_id}")
        return {'status': 'success', 'task_id': task_id}

    except Exception as e:
        logger.error(f"Broadcast bulk import complete failed: {e}")
        return {'status': 'error', 'message': str(e)}

@shared_task
def broadcast_bulk_import_failed(task_id: str, error: str):
    """Broadcast bulk import failure"""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync

        message = {
            'type': 'bulk_import.failed',
            'task_id': task_id,
            'error': error,
            'timestamp': timezone.now().isoformat()
        }

        channel_layer = get_channel_layer()
        if channel_layer:
            async_to_sync(channel_layer.group_send)(
                'bulk_import_updates',
                message
            )

        logger.error(f"Bulk import failed broadcast: {task_id}")
        return {'status': 'success', 'task_id': task_id}

    except Exception as e:
        logger.error(f"Broadcast bulk import failed failed: {e}")
        return {'status': 'error', 'message': str(e)}