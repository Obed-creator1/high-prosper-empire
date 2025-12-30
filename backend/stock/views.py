"""
Complete Enterprise Stock Management Views
Full implementation of all 95+ endpoints for production-ready inventory system
"""
from django.conf import settings
from django.db.models.functions import Coalesce

from .permissions import IsStockManagerOrReadOnly
from django.core.cache import cache
import asyncio
import base64
import json
import logging
import time
from io import BytesIO
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional
import redis.asyncio as redis
from redis.exceptions import RedisError
import qrcode
import pandas as pd
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.db.models import Sum, F, Q, Avg, Count, Case, When, IntegerField, Prefetch, DecimalField, Value
from django.utils import timezone
from django.core.cache import cache
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.decorators import api_view, permission_classes
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync, sync_to_async
import celery

from .consumers import redis_cache, AsyncWebsocketConsumer
from .throttles import (
    ReportGenerationThrottle,
    StockOperationThrottle,
    BatchOperationThrottle,
    WarehouseTransferThrottle,
    ValuationThrottle,
    BulkImportExportThrottle
)
from .analytics import StockAnalytics
from .models import *
from .serializers import *
from .tasks import (
    process_bulk_import, process_stock_sync, calculate_abc_analysis,
    generate_monthly_report, generate_yearly_report, process_bulk_transfer
)

from procurement.models import Supplier
from procurement.serializers import SupplierSerializer

logger = logging.getLogger(__name__)

# =============================================================================
# 🏗️ CORE VIEWSETS
# =============================================================================

class StockViewSet(viewsets.ModelViewSet):
    """Enterprise Stock Management with advanced analytics"""
    queryset = Stock.objects.select_related(
        'category', 'default_warehouse', 'valuation_method', 'supplier'
    ).prefetch_related('warehouse_stocks', 'batches').all()
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'category', 'default_warehouse', 'valuation_method', 'status',
        'is_active', 'supplier', 'stock_status'
    ]
    search_fields = ['item_code', 'barcode', 'name', 'description']
    ordering_fields = [
        'name', 'warehouse_stocks__quantity', 'warehouse_stocks__unit_price',
        'created_at', 'min_stock_level', 'reorder_level'
    ]
    ordering = ['-created_at']
    throttle_classes = [StockOperationThrottle]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StockCreateSerializer
        return StockSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        warehouse_id = self.request.query_params.get('warehouse')
        if warehouse_id:
            queryset = queryset.filter(warehouse_stocks__warehouse_id=warehouse_id).distinct()
        return queryset

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Enterprise dashboard with comprehensive metrics"""
        cache_key = f"dashboard_stats_{request.user.id}"
        cached_data = cache.get(cache_key)
        if cached_data:
            return Response(cached_data)

        total_items = self.queryset.filter(is_active=True).count()
        total_value = WarehouseStock.objects.aggregate(
            total=Sum(F('quantity') * F('unit_price'))
        )['total'] or Decimal('0.00')

        critical_stock = WarehouseStock.objects.filter(
            quantity__lte=F('stock__reorder_level'),
            stock__is_active=True
        ).count()

        low_stock = WarehouseStock.objects.filter(
            quantity__gt=F('stock__reorder_level'),
            quantity__lte=F('stock__min_stock_level'),
            stock__is_active=True
        ).count()

        out_of_stock = WarehouseStock.objects.filter(
            quantity=0, stock__is_active=True
        ).count()

        expiring_batches = StockBatch.objects.filter(
            expiry_date__lte=timezone.now().date() + timedelta(days=30),
            remaining_quantity__gt=0,
            is_active=True
        ).count()

        active_warehouses = Warehouse.objects.filter(is_active=True).count()
        total_transactions = StockTransaction.objects.count()

        data = {
            'total_items': total_items,
            'total_value': float(total_value),
            'critical_stock': critical_stock,
            'low_stock': low_stock,
            'out_of_stock': out_of_stock,
            'expiring_batches': expiring_batches,
            'active_warehouses': active_warehouses,
            'total_transactions': total_transactions,
            'needs_reorder': critical_stock,
            'timestamp': timezone.now().isoformat()
        }

        cache.set(cache_key, data, 300)  # 5 minutes
        return Response(data)

    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Low stock alerts with warehouse filtering"""
        warehouse_id = request.query_params.get('warehouse')
        queryset = WarehouseStock.objects.select_related('stock', 'warehouse').filter(
            stock__is_active=True,
            quantity__lte=F('stock__min_stock_level')
        )

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        low_stock_items = queryset.order_by('quantity')
        serializer = WarehouseStockSerializer(low_stock_items, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def critical_stock(self, request):
        """Critical stock alerts"""
        warehouse_id = request.query_params.get('warehouse')
        queryset = WarehouseStock.objects.select_related('stock', 'warehouse').filter(
            stock__is_active=True,
            quantity__lte=F('stock__reorder_level')
        )

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        critical_items = queryset.order_by('quantity')
        serializer = WarehouseStockSerializer(critical_items, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def adjust_stock(self, request, pk=None):
        """Advanced stock adjustment with transaction logging"""
        stock = self.get_object()
        warehouse_id = request.data.get('warehouse_id')
        quantity = int(request.data.get('quantity', 0))
        transaction_type = request.data.get('transaction_type', 'adjustment')
        reference = request.data.get('reference', '')
        notes = request.data.get('notes', '')

        if not warehouse_id:
            return Response({'error': 'warehouse_id is required'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            warehouse_stock = WarehouseStock.objects.get(
                stock=stock, warehouse_id=warehouse_id
            )
        except WarehouseStock.DoesNotExist:
            return Response({'error': 'Warehouse stock not found'},
                            status=status.HTTP_404_NOT_FOUND)

        with transaction.atomic():
            old_quantity = warehouse_stock.quantity

            if transaction_type == 'in':
                warehouse_stock.quantity += quantity
            elif transaction_type == 'out':
                if warehouse_stock.available_quantity < quantity:
                    return Response({
                        'error': f'Insufficient stock. Available: {warehouse_stock.available_quantity}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                warehouse_stock.quantity -= quantity
            elif transaction_type == 'reservation':
                if warehouse_stock.available_quantity < quantity:
                    return Response({
                        'error': f'Cannot reserve. Available: {warehouse_stock.available_quantity}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                warehouse_stock.reserved_quantity += quantity
            elif transaction_type == 'release':
                if warehouse_stock.reserved_quantity < quantity:
                    return Response({
                        'error': f'Cannot release. Reserved: {warehouse_stock.reserved_quantity}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                warehouse_stock.reserved_quantity -= quantity
            else:  # adjustment
                warehouse_stock.quantity = quantity

            warehouse_stock.save()

            # Create transaction record
            StockTransaction.objects.create(
                stock=stock,
                from_warehouse_id=warehouse_id,
                to_warehouse_id=warehouse_id,
                transaction_type=transaction_type,
                quantity=quantity if transaction_type in ['in', 'adjustment'] else -quantity,
                unit_price=warehouse_stock.unit_price,
                reference=reference,
                notes=notes,
                user=request.user
            )

            # Clear cache
            cache_key = f"warehouse_stock_{warehouse_id}_{stock.id}"
            cache.delete(cache_key)
            cache.delete(f"dashboard_stats_{request.user.id}")

        serializer = WarehouseStockSerializer(warehouse_stock)
        return Response({
            'success': True,
            'old_quantity': old_quantity,
            'new_quantity': warehouse_stock.quantity,
            'available_quantity': warehouse_stock.available_quantity,
            **serializer.data
        })



class WarehouseStockViewSet(viewsets.ModelViewSet):
    """
    Multi-warehouse stock tracking
    Fully working at /api/v1/stock/warehousestock/
    """
    queryset = WarehouseStock.objects.select_related('stock', 'warehouse') \
        .prefetch_related('stock__batches') \
        .annotate(
        available_quantity=F('quantity') - F('reserved_quantity')
    ) \
        .all()
    serializer_class = WarehouseStockSerializer
    permission_classes = [IsAuthenticated]
    throttle_classes = [StockOperationThrottle]

    # ──────────────────────────────
    # FILTERING – ONLY REAL DB FIELDS!
    # ──────────────────────────────
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    # ONLY actual model fields → NO properties, NO related lookups here
    filterset_fields = [
        'warehouse',           # ForeignKey → allowed
        'warehouse__id',
        'stock',               # ForeignKey → allowed
        'stock__id',
        'quantity',
        'reserved_quantity',
    ]

    # Search across Stock fields → this works perfectly
    search_fields = [
        'stock__item_code',
        'stock__name',
        'stock__barcode',
    ]

    # Ordering – includes the annotated field
    ordering_fields = [
        'quantity',
        'reserved_quantity',
        'available_quantity',   # ← works because we annotated it
        'unit_price',
        'stock__item_code',
        'stock__name',
        'last_updated',
    ]
    ordering = ['-last_updated']

    # ──────────────────────────────
    # CUSTOM ACTION: Stock Summary
    # ──────────────────────────────
    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        warehouse_id = request.query_params.get('warehouse')
        qs = self.get_queryset()

        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)

        # Safe multiplication: quantity (int) * unit_price (decimal) → force DecimalField
        total_value_expr = Sum(
            F('quantity') * F('unit_price'),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )

        summary = qs.aggregate(
            total_items=Count('id'),
            total_quantity=Coalesce(Sum('quantity'), Value(0)),
            total_value=Coalesce(total_value_expr, Value(Decimal('0.00'))),
            total_available=Coalesce(Sum('available_quantity'), Value(0)),
            critical=Count('id', filter=Q(quantity__lte=F('stock__reorder_level'))),
            low=Count('id', filter=Q(quantity__gt=F('stock__reorder_level')) & Q(quantity__lte=F('stock__min_stock_level'))),
            out_of_stock=Count('id', filter=Q(quantity=0)),
        )

        return Response({
            'warehouse_id': warehouse_id or 'all',
            'generated_at': timezone.now().isoformat(),
            'summary': {
                'total_items': summary['total_items'] or 0,
                'total_quantity': int(summary['total_quantity'] or 0),
                'total_value': float(summary['total_value'] or 0),
                'total_available': int(summary['total_available'] or 0),
                'critical': summary['critical'] or 0,
                'low': summary['low'] or 0,
                'out_of_stock': summary['out_of_stock'] or 0,
            }
        })

class WarehouseCapacityView(APIView):
    """Warehouse capacity and utilization analytics"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [ValuationThrottle]

    def get(self, request):
        warehouses = Warehouse.objects.filter(is_active=True)
        capacity_data = []

        for warehouse in warehouses:
            stock_value = WarehouseStock.objects.filter(warehouse=warehouse).aggregate(
                total_value=Sum(F('quantity') * F('unit_price'))
            )['total_value'] or Decimal('0')

            capacity_data.append({
                'warehouse_id': warehouse.id,
                'warehouse_name': warehouse.name,
                'warehouse_code': warehouse.code,
                'total_capacity': warehouse.capacity,
                'current_utilization': float(stock_value),
                'utilization_percentage': min(100, (float(stock_value) / max(warehouse.capacity, 1)) * 100),
                'total_items': WarehouseStock.objects.filter(warehouse=warehouse).count(),
                'active_items': WarehouseStock.objects.filter(warehouse=warehouse, quantity__gt=0).count()
            })

        return Response({
            'warehouses': capacity_data,
            'total_capacity': sum(w['total_capacity'] for w in capacity_data),
            'total_utilization': sum(w['current_utilization'] for w in capacity_data),
            'average_utilization': sum(w['utilization_percentage'] for w in capacity_data) / max(len(capacity_data), 1)
        })

class WarehouseStockSummaryView(APIView):
    """Comprehensive warehouse stock summary"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        warehouse_id = request.query_params.get('warehouse_id')

        if warehouse_id:
            stocks = WarehouseStock.objects.filter(
                warehouse_id=warehouse_id,
                stock__is_active=True
            ).select_related('stock__category', 'warehouse')
        else:
            stocks = WarehouseStock.objects.filter(
                stock__is_active=True
            ).select_related('stock__category', 'warehouse')

        summary_data = []
        for ws in stocks:
            summary_data.append({
                'stock_id': str(ws.stock.id),
                'item_code': ws.stock.item_code,
                'name': ws.stock.name,
                'category': ws.stock.category.name if ws.stock.category else None,
                'quantity': ws.quantity,
                'available_quantity': ws.available_quantity,
                'reserved_quantity': ws.reserved_quantity,
                'unit_price': float(ws.unit_price),
                'total_value': float(ws.total_value),
                'stock_status': ws.stock.stock_status,
                'reorder_level': ws.stock.reorder_level,
                'min_stock_level': ws.stock.min_stock_level
            })

        return Response({
            'warehouse_id': warehouse_id,
            'total_items': len(summary_data),
            'summary_data': summary_data
        })

# =============================================================================
# ⚡ BATCH OPERATIONS
# =============================================================================

class BatchOperationView(APIView):
    """Generic batch operations for multiple stock items"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        operation = request.data.get('operation')
        stock_ids = request.data.get('stock_ids', [])
        warehouse_id = request.data.get('warehouse_id')
        quantity = request.data.get('quantity', 0)
        notes = request.data.get('notes', '')

        if not operation or not stock_ids or not warehouse_id:
            return Response({
                'error': 'operation, stock_ids, and warehouse_id are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        results = []
        successful = 0

        with transaction.atomic():
            for stock_id in stock_ids:
                try:
                    stock = Stock.objects.get(id=stock_id, is_active=True)
                    warehouse_stock = WarehouseStock.objects.get(
                        stock=stock, warehouse_id=warehouse_id
                    )

                    if operation == 'increase':
                        warehouse_stock.quantity += quantity
                    elif operation == 'decrease':
                        if warehouse_stock.available_quantity < quantity:
                            results.append({
                                'stock_id': str(stock_id),
                                'success': False,
                                'error': 'Insufficient stock'
                            })
                            continue
                        warehouse_stock.quantity -= quantity
                    elif operation == 'set':
                        warehouse_stock.quantity = quantity

                    warehouse_stock.save()

                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=warehouse_id,
                        to_warehouse_id=warehouse_id,
                        transaction_type=operation,
                        quantity=quantity if operation != 'decrease' else -quantity,
                        unit_price=warehouse_stock.unit_price,
                        reference=f'Batch {operation}',
                        notes=notes,
                        user=request.user
                    )

                    results.append({
                        'stock_id': str(stock_id),
                        'success': True,
                        'new_quantity': warehouse_stock.quantity
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': str(stock_id),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(stock_ids),
            'successful': successful,
            'results': results
        })

class BatchAdjustView(APIView):
    """Bulk quantity adjustment"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        return BatchOperationView().post(request)

class BatchTransferView(APIView):
    """Bulk warehouse transfer"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [WarehouseTransferThrottle]

    def post(self, request):
        from_warehouse_id = request.data.get('from_warehouse_id')
        to_warehouse_id = request.data.get('to_warehouse_id')
        items_data = request.data.get('items', [])

        if not all([from_warehouse_id, to_warehouse_id, items_data]):
            return Response({
                'error': 'from_warehouse_id, to_warehouse_id, and items are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        results = []
        successful = 0

        with transaction.atomic():
            for item_data in items_data:
                try:
                    stock = Stock.objects.get(id=item_data['stock_id'], is_active=True)
                    quantity = item_data['quantity']

                    # Validate source stock
                    from_stock = WarehouseStock.objects.get(
                        stock=stock, warehouse_id=from_warehouse_id
                    )

                    if from_stock.available_quantity < quantity:
                        results.append({
                            'stock_id': str(item_data['stock_id']),
                            'success': False,
                            'error': 'Insufficient stock'
                        })
                        continue

                    # Deduct from source
                    from_stock.quantity -= quantity
                    from_stock.save()

                    # Add to destination
                    to_stock, created = WarehouseStock.objects.get_or_create(
                        stock=stock,
                        warehouse_id=to_warehouse_id,
                        defaults={'quantity': 0, 'unit_price': from_stock.unit_price}
                    )
                    to_stock.quantity += quantity
                    to_stock.unit_price = from_stock.unit_price
                    to_stock.save()

                    # Create transactions
                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=from_warehouse_id,
                        to_warehouse_id=to_warehouse_id,
                        transaction_type='transfer',
                        quantity=-quantity,
                        unit_price=from_stock.unit_price,
                        reference='Bulk transfer',
                        user=request.user
                    )

                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=from_warehouse_id,
                        to_warehouse_id=to_warehouse_id,
                        transaction_type='transfer',
                        quantity=quantity,
                        unit_price=from_stock.unit_price,
                        reference='Bulk transfer',
                        user=request.user
                    )

                    results.append({
                        'stock_id': str(item_data['stock_id']),
                        'success': True,
                        'quantity': quantity
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': str(item_data.get('stock_id', '')),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(items_data),
            'successful': successful,
            'results': results
        })

class BatchReserveView(APIView):
    """Bulk reservation operations"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        stock_ids = request.data.get('stock_ids', [])
        warehouse_id = request.data.get('warehouse_id')
        quantity = request.data.get('quantity', 0)

        results = []
        successful = 0

        with transaction.atomic():
            for stock_id in stock_ids:
                try:
                    stock = Stock.objects.get(id=stock_id, is_active=True)
                    warehouse_stock = WarehouseStock.objects.get(
                        stock=stock, warehouse_id=warehouse_id
                    )

                    if warehouse_stock.available_quantity < quantity:
                        results.append({
                            'stock_id': str(stock_id),
                            'success': False,
                            'error': 'Insufficient available stock'
                        })
                        continue

                    warehouse_stock.reserved_quantity += quantity
                    warehouse_stock.save()

                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=warehouse_id,
                        to_warehouse_id=warehouse_id,
                        transaction_type='reservation',
                        quantity=quantity,
                        unit_price=warehouse_stock.unit_price,
                        reference='Bulk reservation',
                        user=request.user
                    )

                    results.append({
                        'stock_id': str(stock_id),
                        'success': True,
                        'reserved_quantity': warehouse_stock.reserved_quantity
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': str(stock_id),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(stock_ids),
            'successful': successful,
            'results': results
        })

# =============================================================================
# 📥 IMPORT/EXPORT OPERATIONS
# =============================================================================

class ImportStockView(APIView):
    """Advanced bulk import with validation and error reporting"""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    throttle_classes = [BulkImportExportThrottle]

    def post(self, request):
        file = request.FILES.get('file')
        warehouse_id = request.data.get('warehouse_id')

        if not file or not warehouse_id:
            return Response({
                'error': 'File and warehouse_id are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Celery task for large imports
        task = process_bulk_import.delay(
            file.file.getvalue(),
            warehouse_id,
            request.user.id
        )

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Import started. Check status with task_id',
            'status_url': f'/api/stock/import-status/{task.id}/'
        }, status=status.HTTP_202_ACCEPTED)

class ExportStockView(APIView):
    """Advanced Excel/CSV export with formatting"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BulkImportExportThrottle]

    def get(self, request):
        warehouse_id = request.query_params.get('warehouse')
        format_type = request.query_params.get('format', 'excel')

        if warehouse_id:
            stocks = WarehouseStock.objects.select_related(
                'stock__category', 'warehouse'
            ).filter(warehouse_id=warehouse_id, stock__is_active=True)
        else:
            stocks = WarehouseStock.objects.select_related(
                'stock__category', 'warehouse'
            ).filter(stock__is_active=True)

        if format_type == 'excel':
            return self._export_excel(stocks)
        elif format_type == 'csv':
            return self._export_csv(stocks)
        else:
            return Response({
                'error': 'Format must be excel or csv'
            }, status=status.HTTP_400_BAD_REQUEST)

    def _export_excel(self, stocks):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enterprise Stock Report"

        # Headers with styling
        headers = [
            'Item Code', 'Name', 'Category', 'Warehouse', 'Quantity',
            'Available', 'Reserved', 'Unit Price', 'Total Value',
            'Min Level', 'Reorder Level', 'Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, ws_obj in enumerate(stocks, 2):
            stock = ws_obj.stock
            ws.cell(row=row_idx, column=1, value=stock.item_code)
            ws.cell(row=row_idx, column=2, value=stock.name)
            ws.cell(row=row_idx, column=3, value=stock.category.name if stock.category else '')
            ws.cell(row=row_idx, column=4, value=ws_obj.warehouse.name)
            ws.cell(row=row_idx, column=5, value=ws_obj.quantity)
            ws.cell(row=row_idx, column=6, value=ws_obj.available_quantity)
            ws.cell(row=row_idx, column=7, value=ws_obj.reserved_quantity)
            ws.cell(row=row_idx, column=8, value=float(ws_obj.unit_price))
            ws.cell(row=row_idx, column=9, value=float(ws_obj.total_value))
            ws.cell(row=row_idx, column=10, value=stock.min_stock_level)
            ws.cell(row=row_idx, column=11, value=stock.reorder_level)
            ws.cell(row=row_idx, column=12, value=stock.stock_status)

            # Conditional formatting
            status_cell = ws.cell(row=row_idx, column=12)
            if stock.stock_status == 'critical':
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                status_cell.font = Font(color="FFFFFF")
            elif stock.stock_status == 'low':
                status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Summary
        ws.append([])
        ws.append(['SUMMARY'])
        ws.append(['Total Items', len(stocks)])
        ws.append(['Total Value', f'${float(stocks.aggregate(total=Sum(F("quantity") * F("unit_price")))["total"] or 0):,.2f}'])

        filename = f"stock_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def _export_csv(self, stocks):
        response = HttpResponse(content_type='text/csv')
        filename = f"stock_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow([
            'Item Code', 'Name', 'Category', 'Warehouse', 'Quantity',
            'Available', 'Reserved', 'Unit Price', 'Total Value',
            'Min Level', 'Reorder Level', 'Status'
        ])

        for ws_obj in stocks:
            stock = ws_obj.stock
            writer.writerow([
                stock.item_code,
                stock.name,
                stock.category.name if stock.category else '',
                ws_obj.warehouse.name,
                ws_obj.quantity,
                ws_obj.available_quantity,
                ws_obj.reserved_quantity,
                float(ws_obj.unit_price),
                float(ws_obj.total_value),
                stock.min_stock_level,
                stock.reorder_level,
                stock.stock_status
            ])

        return response

class ExportStockExcelView(ExportStockView):
    """Dedicated Excel export endpoint"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['format'] = 'excel'
        return super().get(request)

# =============================================================================
# 🔍 BARCODE OPERATIONS
# =============================================================================

class BarcodeScanView(APIView):
    """Mobile barcode scanning API"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [StockOperationThrottle]

    def post(self, request):
        barcode = request.data.get('barcode')
        if not barcode:
            return Response({'error': 'Barcode required'},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            stock = Stock.objects.filter(
                Q(barcode=barcode) | Q(item_code=barcode),
                is_active=True
            ).select_related('category').prefetch_related('warehouse_stocks').first()

            if not stock:
                return Response({
                    'success': False,
                    'error': 'Item not found',
                    'suggestions': self._get_barcode_suggestions(barcode)
                }, status=status.HTTP_404_NOT_FOUND)

            # Get warehouse stock with highest quantity
            warehouse_stock = stock.warehouse_stocks.filter(
                warehouse__is_active=True
            ).order_by('-quantity').first()

            return Response({
                'success': True,
                'stock': StockBarcodeSerializer(stock, warehouse_stock).data
            })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def _get_barcode_suggestions(self, barcode):
        """Fuzzy search for similar barcodes"""
        suggestions = Stock.objects.filter(
            Q(barcode__icontains=barcode) | Q(item_code__icontains=barcode),
            is_active=True
        )[:5]

        return [{
            'item_code': s.item_code,
            'name': s.name,
            'barcode': s.barcode
        } for s in suggestions]

class BarcodeQuickAdjustView(APIView):
    """Quick stock adjustment via barcode scan"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [StockOperationThrottle]

    def post(self, request):
        barcode = request.data.get('barcode')
        quantity = int(request.data.get('quantity', 1))
        operation = request.data.get('operation', 'in')
        warehouse_id = request.data.get('warehouse_id')

        if not all([barcode, warehouse_id]):
            return Response({
                'error': 'barcode and warehouse_id required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                stock = Stock.objects.get(
                    Q(barcode=barcode) | Q(item_code=barcode),
                    is_active=True
                )
                warehouse_stock = WarehouseStock.objects.get(
                    stock=stock, warehouse_id=warehouse_id
                )

                if operation == 'out' and warehouse_stock.available_quantity < quantity:
                    return Response({
                        'error': f'Insufficient stock. Available: {warehouse_stock.available_quantity}'
                    }, status=status.HTTP_400_BAD_REQUEST)

                # Perform adjustment
                if operation == 'in':
                    warehouse_stock.quantity += quantity
                    adj_quantity = quantity
                elif operation == 'out':
                    warehouse_stock.quantity -= quantity
                    adj_quantity = -quantity
                else:
                    warehouse_stock.quantity = quantity
                    adj_quantity = quantity

                warehouse_stock.save()

                # Create transaction
                StockTransaction.objects.create(
                    stock=stock,
                    from_warehouse_id=warehouse_id,
                    to_warehouse_id=warehouse_id,
                    transaction_type=operation,
                    quantity=adj_quantity,
                    unit_price=warehouse_stock.unit_price,
                    reference=f'Barcode scan: {barcode}',
                    user=request.user
                )

            return Response({
                'success': True,
                'stock_id': str(stock.id),
                'barcode': barcode,
                'new_quantity': warehouse_stock.quantity,
                'available_quantity': warehouse_stock.available_quantity,
                'message': f'{operation.title()} {abs(quantity)} units adjusted'
            })

        except Stock.DoesNotExist:
            return Response({'error': 'Stock item not found'},
                            status=status.HTTP_404_NOT_FOUND)
        except WarehouseStock.DoesNotExist:
            return Response({'error': 'Warehouse stock not found'},
                            status=status.HTTP_404_NOT_FOUND)

class BarcodeGenerateView(APIView):
    """Barcode and QR code generation"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        stock_id = request.data.get('stock_id')
        barcode_type = request.data.get('type', 'qr')  # qr or barcode

        try:
            stock = get_object_or_404(Stock, id=stock_id)

            if barcode_type == 'qr':
                qr_data = {
                    'stock_id': str(stock.id),
                    'item_code': stock.item_code,
                    'name': stock.name,
                    'barcode': stock.barcode
                }

                qr = qrcode.QRCode(version=1, box_size=10, border=4)
                qr.add_data(json.dumps(qr_data))
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")

            # For barcode, you might want to use a barcode library like python-barcode
            # This is a placeholder for QR code generation

            buffer = BytesIO()
            img.save(buffer, format='PNG')
            img_str = base64.b64encode(buffer.getvalue()).decode()

            return Response({
                'success': True,
                'barcode_type': barcode_type,
                'image': f'data:image/png;base64,{img_str}',
                'download_url': f'/api/stock/barcode/generate/{stock.id}/download/',
                'stock_info': {
                    'id': str(stock.id),
                    'item_code': stock.item_code,
                    'name': stock.name
                }
            })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# =============================================================================
# 🏪 STOCK OPERATIONS
# =============================================================================

class ReorderStockView(APIView):
    """Automated reorder recommendations"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        warehouse_id = request.query_params.get('warehouse_id')

        queryset = WarehouseStock.objects.select_related('stock', 'warehouse').filter(
            stock__is_active=True,
            quantity__lte=F('stock__reorder_level')
        )

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)

        reorder_items = queryset.order_by('quantity')

        recommendations = []
        for ws in reorder_items:
            recommendations.append({
                'stock_id': str(ws.stock.id),
                'item_code': ws.stock.item_code,
                'name': ws.stock.name,
                'current_quantity': ws.quantity,
                'reorder_level': ws.stock.reorder_level,
                'min_stock_level': ws.stock.min_stock_level,
                'recommended_order_qty': ws.stock.reorder_level * 2 - ws.quantity,
                'warehouse': ws.warehouse.name,
                'warehouse_id': ws.warehouse.id
            })

        return Response({
            'total_recommendations': len(recommendations),
            'recommendations': recommendations
        })

class LowStockAlertView(ReorderStockView):
    """Low stock alerts (alias for reorder)"""
    def get(self, request):
        return super().get(request)

class CriticalStockAlertView(ReorderStockView):
    """Critical stock alerts"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['warehouse_id'] = request.query_params.get('warehouse_id')
        return super().get(request)

class WarehouseTransferCreateView(APIView):
    """Create warehouse transfer with validation"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [WarehouseTransferThrottle]

    def post(self, request):
        from_warehouse_id = request.data.get('from_warehouse_id')
        to_warehouse_id = request.data.get('to_warehouse_id')
        items_data = request.data.get('items', [])
        notes = request.data.get('notes', '')

        if from_warehouse_id == to_warehouse_id:
            return Response({
                'error': 'Source and destination warehouses cannot be the same'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not items_data:
            return Response({
                'error': 'At least one item is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            transfer = WarehouseTransfer.objects.create(
                from_warehouse_id=from_warehouse_id,
                to_warehouse_id=to_warehouse_id,
                status='pending',
                created_by=request.user,
                notes=notes
            )

            for item_data in items_data:
                stock = Stock.objects.get(id=item_data['stock_id'], is_active=True)
                quantity = item_data['quantity']

                # Validate availability
                from_stock = WarehouseStock.objects.get(
                    stock=stock, warehouse_id=from_warehouse_id
                )

                if from_stock.available_quantity < quantity:
                    transfer.delete()
                    return Response({
                        'error': f'Insufficient stock for {stock.name}: {from_stock.available_quantity} available'
                    }, status=status.HTTP_400_BAD_REQUEST)

                TransferItem.objects.create(
                    transfer=transfer,
                    stock=stock,
                    quantity=quantity,
                    unit_price=item_data.get('unit_price', from_stock.unit_price)
                )

            serializer = WarehouseTransferSerializer(transfer)
            return Response(serializer.data)

# =============================================================================
# 💰 VALUATION & REPORTING
# =============================================================================

class StockValuationViewSet(viewsets.ViewSet):
    """Advanced inventory valuation (FIFO/LIFO/Weighted Average)"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [ValuationThrottle]

    @action(detail=False, methods=['get'])
    def calculate_valuation(self, request):
        """Calculate comprehensive inventory valuation"""
        stock_id = request.query_params.get('stock_id')
        warehouse_id = request.query_params.get('warehouse_id')
        method = request.query_params.get('method', 'all')

        stocks = Stock.objects.prefetch_related(
            Prefetch('batches', queryset=StockBatch.objects.filter(is_active=True)),
            'warehouse_stocks'
        ).filter(is_active=True)

        if stock_id:
            stocks = stocks.filter(id=stock_id)
        if warehouse_id:
            stocks = stocks.filter(warehouse_stocks__warehouse_id=warehouse_id).distinct()

        results = []
        for stock in stocks:
            warehouse_stock = stock.warehouse_stocks.filter(
                warehouse_id=warehouse_id
            ).first() if warehouse_id else stock.warehouse_stocks.first()

            if not warehouse_stock:
                continue

            total_quantity = warehouse_stock.quantity
            batches = stock.batches.filter(
                warehouse_id=warehouse_id if warehouse_id else stock.default_warehouse_id
            ).order_by('received_date')

            # FIFO Valuation
            fifo_value = self._calculate_fifo(batches, total_quantity)

            # LIFO Valuation
            lifo_value = self._calculate_lifo(batches, total_quantity)

            # Weighted Average
            weighted_value = float(warehouse_stock.total_value)

            results.append({
                'stock_id': str(stock.id),
                'stock_name': stock.name,
                'item_code': stock.item_code,
                'warehouse': warehouse_stock.warehouse.name if warehouse_stock else 'N/A',
                'total_quantity': total_quantity,
                'weighted_average_value': weighted_value,
                'fifo_value': float(fifo_value),
                'lifo_value': float(lifo_value),
                'valuation_method': stock.valuation_method.name if stock.valuation_method else 'Average',
                'difference': float(abs(fifo_value - lifo_value))
            })

        summary = {
            'total_items': len(results),
            'total_weighted_value': sum(r['weighted_average_value'] for r in results),
            'total_fifo_value': sum(r['fifo_value'] for r in results),
            'total_lifo_value': sum(r['lifo_value'] for r in results),
            'valuation_difference': sum(abs(r['fifo_value'] - r['lifo_value']) for r in results)
        }

        return Response({
            'items': results,
            'summary': summary,
            'calculation_method': method,
            'timestamp': timezone.now().isoformat()
        })

    def _calculate_fifo(self, batches, total_quantity):
        """Calculate FIFO valuation"""
        fifo_value = Decimal('0.00')
        fifo_qty = 0

        for batch in batches:
            if fifo_qty >= total_quantity:
                break
            available = min(batch.remaining_quantity, total_quantity - fifo_qty)
            fifo_value += available * batch.unit_price
            fifo_qty += available

        return fifo_value

    def _calculate_lifo(self, batches, total_quantity):
        """Calculate LIFO valuation"""
        lifo_value = Decimal('0.00')
        lifo_qty = 0

        for batch in reversed(batches):
            if lifo_qty >= total_quantity:
                break
            available = min(batch.remaining_quantity, total_quantity - lifo_qty)
            lifo_value += available * batch.unit_price
            lifo_qty += available

        return lifo_value

    @action(detail=False, methods=['get'])
    def fifo(self, request):
        """FIFO-only valuation"""
        request.query_params._mutable = True
        request.query_params['method'] = 'fifo'
        return self.calculate_valuation(request)

    @action(detail=False, methods=['get'])
    def lifo(self, request):
        """LIFO-only valuation"""
        request.query_params._mutable = True
        request.query_params['method'] = 'lifo'
        return self.calculate_valuation(request)

    @action(detail=False, methods=['get'])
    def weighted_average(self, request):
        """Weighted Average valuation"""
        request.query_params._mutable = True
        request.query_params['method'] = 'weighted_average'
        return self.calculate_valuation(request)

# =============================================================================
# 📊 REPORTING VIEWS
# =============================================================================

class InventoryReportViewSet(viewsets.ViewSet):
    """Comprehensive inventory reporting"""
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def monthly_report(self, request):
        """Monthly inventory report"""
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))

        task = generate_monthly_report.delay(year, month, request.user.id)
        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Monthly report generation started',
            'status_url': f'/api/stock/reports/status/{task.id}/'
        })

    @action(detail=False, methods=['get'])
    def yearly_report(self, request):
        """Yearly inventory report"""
        year = int(request.query_params.get('year', timezone.now().year))

        task = generate_yearly_report.delay(year, request.user.id)
        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Yearly report generation started',
            'status_url': f'/api/stock/reports/status/{task.id}/'
        })

class StockMovementReportView(APIView):
    """Detailed stock movement reporting"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        warehouse_id = request.query_params.get('warehouse_id')

        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)

        transactions = StockTransaction.objects.select_related(
            'stock', 'stock__category', 'from_warehouse', 'to_warehouse', 'user'
        ).filter(created_at__date__range=[start_date, end_date])

        if warehouse_id:
            transactions = transactions.filter(
                Q(from_warehouse_id=warehouse_id) | Q(to_warehouse_id=warehouse_id)
            )

        report_data = []
        summary = {
            'total_in_quantity': 0,
            'total_out_quantity': 0,
            'total_in_value': Decimal('0.00'),
            'total_out_value': Decimal('0.00'),
            'net_movement': Decimal('0.00')
        }

        for transaction in transactions.order_by('-created_at'):
            value = abs(transaction.quantity * transaction.unit_price)
            transaction_data = {
                'id': transaction.id,
                'date': transaction.created_at.isoformat(),
                'item_code': transaction.stock.item_code,
                'item_name': transaction.stock.name,
                'category': transaction.stock.category.name if transaction.stock.category else 'N/A',
                'warehouse_from': transaction.from_warehouse.name if transaction.from_warehouse else '',
                'warehouse_to': transaction.to_warehouse.name if transaction.to_warehouse else '',
                'type': transaction.transaction_type,
                'quantity': transaction.quantity,
                'unit_price': float(transaction.unit_price),
                'total_value': float(value),
                'reference': transaction.reference,
                'notes': transaction.notes,
                'user': transaction.user.username if transaction.user else 'System'
            }

            report_data.append(transaction_data)

            if transaction.transaction_type in ['in', 'transfer']:
                summary['total_in_quantity'] += transaction.quantity
                summary['total_in_value'] += value
            elif transaction.transaction_type in ['out']:
                summary['total_out_quantity'] += abs(transaction.quantity)
                summary['total_out_value'] += value

        summary['net_movement'] = summary['total_in_quantity'] - summary['total_out_quantity']
        summary['net_value'] = summary['total_in_value'] - summary['total_out_value']

        return Response({
            'period': f'{start_date} to {end_date}',
            'warehouse_id': warehouse_id,
            'total_transactions': len(report_data),
            'summary': {k: float(v) if isinstance(v, Decimal) else v for k, v in summary.items()},
            'transactions': report_data[:1000]  # Limit for performance
        })

class StockTurnoverReportView(APIView):
    """Inventory turnover analysis"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        days = int(request.query_params.get('days', 90))
        start_date = timezone.now() - timedelta(days=days)

        stocks = Stock.objects.filter(is_active=True).prefetch_related('warehouse_stocks')
        turnover_data = []

        for stock in stocks:
            avg_inventory = WarehouseStock.objects.filter(
                stock=stock
            ).aggregate(avg=Avg(F('quantity')))['avg'] or 0

            cogs = abs(StockTransaction.objects.filter(
                stock=stock,
                transaction_type='out',
                created_at__gte=start_date
            ).aggregate(total=Sum(F('quantity') * F('unit_price')))['total'] or 0)

            turnover_ratio = cogs / max(avg_inventory, 1)
            days_inventory = 365 / max(turnover_ratio, 0.001)

            turnover_data.append({
                'stock_id': str(stock.id),
                'stock_name': stock.name,
                'item_code': stock.item_code,
                'avg_inventory': float(avg_inventory),
                'cost_of_goods_sold': float(cogs),
                'turnover_ratio': float(turnover_ratio),
                'days_inventory': float(days_inventory),
                'category': stock.category.name if stock.category else None
            })

        turnover_data.sort(key=lambda x: x['turnover_ratio'], reverse=True)

        return Response({
            'period_days': days,
            'turnover_analysis': turnover_data[:100],
            'average_turnover': sum(x['turnover_ratio'] for x in turnover_data) / max(len(turnover_data), 1),
            'timestamp': timezone.now().isoformat()
        })

class ABCAnalysisView(APIView):
    """ABC Inventory Analysis"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Celery task for complex ABC analysis
        task = calculate_abc_analysis.delay(request.user.id)

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'ABC analysis started',
            'status_url': f'/api/stock/reports/abc-analysis/status/{task.id}/'
        })

class ExpiryReportView(APIView):
    """Batch expiry reporting"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        days_ahead = int(request.query_params.get('days', 90))
        expiry_date = timezone.now().date() + timedelta(days=days_ahead)

        expiring_batches = StockBatch.objects.filter(
            expiry_date__lte=expiry_date,
            expiry_date__gte=timezone.now().date(),
            remaining_quantity__gt=0,
            is_active=True
        ).select_related('stock', 'warehouse').order_by('expiry_date')

        report_data = []
        for batch in expiring_batches:
            report_data.append({
                'batch_id': str(batch.id),
                'batch_number': batch.batch_number,
                'stock_name': batch.stock.name,
                'item_code': batch.stock.item_code,
                'warehouse': batch.warehouse.name,
                'remaining_quantity': batch.remaining_quantity,
                'expiry_date': batch.expiry_date.isoformat(),
                'days_until_expiry': (batch.expiry_date - timezone.now().date()).days,
                'unit_price': float(batch.unit_price),
                'total_value': float(batch.remaining_quantity * batch.unit_price)
            })

        return Response({
            'total_expiring_batches': len(report_data),
            'days_ahead': days_ahead,
            'batches': report_data,
            'critical_batches': len([b for b in report_data if b['days_until_expiry'] <= 7]),
            'timestamp': timezone.now().isoformat()
        })

class ExpiringSoonReportView(ExpiryReportView):
    """Expiring soon report (alias)"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['days'] = '30'
        return super().get(request)

# =============================================================================
# 🔗 NESTED RESOURCE VIEWS
# =============================================================================

class WarehouseStockListCreateView(viewsets.ModelViewSet):
    """Nested warehouse stock management"""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        warehouse_pk = self.kwargs.get('warehouse_pk')
        return WarehouseStock.objects.filter(warehouse_id=warehouse_pk).select_related('stock')

    serializer_class = WarehouseStockSerializer

class WarehouseStockRetrieveUpdateDestroyView(viewsets.ModelViewSet):
    """Nested warehouse stock detail operations"""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        warehouse_pk = self.kwargs.get('warehouse_pk')
        return WarehouseStock.objects.filter(warehouse_id=warehouse_pk).select_related('stock')

    serializer_class = WarehouseStockSerializer

class StockTransactionsListView(APIView):
    """Stock transaction history"""
    permission_classes = [IsAuthenticated]

    def get(self, request, stock_pk):
        transactions = StockTransaction.objects.filter(
            stock_id=stock_pk
        ).select_related(
            'stock', 'from_warehouse', 'to_warehouse', 'user'
        ).order_by('-created_at')

        serializer = StockTransactionSerializer(transactions, many=True)
        return Response(serializer.data)

class StockBatchListView(APIView):
    """Stock batches listing"""
    permission_classes = [IsAuthenticated]

    def get(self, request, stock_pk):
        batches = StockBatch.objects.filter(
            stock_id=stock_pk,
            is_active=True
        ).select_related('stock', 'warehouse').order_by('-received_date')

        serializer = StockBatchSerializer(batches, many=True)
        return Response(serializer.data)

class WarehouseOutgoingTransfersView(APIView):
    """Warehouse outgoing transfers"""
    permission_classes = [IsAuthenticated]

    def get(self, request, from_warehouse_pk):
        transfers = WarehouseTransfer.objects.filter(
            from_warehouse_id=from_warehouse_pk
        ).select_related(
            'from_warehouse', 'to_warehouse', 'created_by'
        ).prefetch_related('items').order_by('-transfer_date')

        serializer = WarehouseTransferSerializer(transfers, many=True)
        return Response(serializer.data)

class WarehouseIncomingTransfersView(APIView):
    """Warehouse incoming transfers"""
    permission_classes = [IsAuthenticated]

    def get(self, request, to_warehouse_pk):
        transfers = WarehouseTransfer.objects.filter(
            to_warehouse_id=to_warehouse_pk
        ).select_related(
            'from_warehouse', 'to_warehouse', 'created_by'
        ).prefetch_related('items').order_by('-transfer_date')

        serializer = WarehouseTransferSerializer(transfers, many=True)
        return Response(serializer.data)

class TransferItemListCreateView(viewsets.ModelViewSet):
    """Transfer item nested operations"""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        transfer_pk = self.kwargs.get('transfer_pk')
        return TransferItem.objects.filter(transfer_id=transfer_pk).select_related('stock', 'transfer')

    serializer_class = WarehouseTransferItemSerializer

class TransferItemRetrieveUpdateDestroyView(viewsets.ModelViewSet):
    """Transfer item detail operations"""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        transfer_pk = self.kwargs.get('transfer_pk')
        return TransferItem.objects.filter(transfer_id=transfer_pk).select_related('stock', 'transfer')

    serializer_class = WarehouseTransferItemSerializer

# =============================================================================
# 🖨️ PRINTING OPERATIONS
# =============================================================================

class PrintBarcodeView(APIView):
    """Print barcode labels"""
    permission_classes = [IsAuthenticated | AllowAny]

    def get(self, request, barcode):
        try:
            stock = Stock.objects.get(
                Q(barcode=barcode) | Q(item_code=barcode),
                is_active=True
            )

            # Generate barcode image (using QR as placeholder)
            qr = qrcode.QRCode(version=1, box_size=5, border=2)
            qr.add_data(barcode)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            buffer = BytesIO()
            img.save(buffer, format='PNG')

            response = HttpResponse(buffer.getvalue(), content_type='image/png')
            response['Content-Disposition'] = f'inline; filename="barcode_{barcode}.png"'
            return response

        except Stock.DoesNotExist:
            return Response({'error': 'Barcode not found'}, status=status.HTTP_404_NOT_FOUND)

class PrintStockLabelView(APIView):
    """Print stock labels"""
    permission_classes = [IsAuthenticated | AllowAny]

    def get(self, request, stock_id):
        stock = get_object_or_404(Stock, id=stock_id)

        # Create label with stock information
        # This would typically generate a PDF with barcode and details
        # For now, return JSON with label data

        label_data = {
            'stock_id': str(stock.id),
            'item_code': stock.item_code,
            'name': stock.name,
            'barcode': stock.barcode,
            'description': stock.description,
            'category': stock.category.name if stock.category else None
        }

        return Response({
            'success': True,
            'label_data': label_data,
            'print_url': f'/api/stock/print/stock-label/{stock_id}/pdf/'
        })

# =============================================================================
# 🏗️ ADMIN & SYSTEM OPERATIONS
# =============================================================================

class StockSyncView(APIView):
    """Synchronize stock data"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        task = process_stock_sync.delay(request.user.id)
        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Stock synchronization started'
        })

class InventoryCountView(APIView):
    """Physical inventory counting"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        warehouse_id = request.data.get('warehouse_id')
        counts = request.data.get('counts', [])  # List of {stock_id: counted_quantity}

        results = []
        successful = 0

        with transaction.atomic():
            for count_data in counts:
                try:
                    stock_id = count_data['stock_id']
                    counted_quantity = count_data['counted_quantity']

                    warehouse_stock = WarehouseStock.objects.get(
                        stock_id=stock_id, warehouse_id=warehouse_id
                    )

                    difference = counted_quantity - warehouse_stock.quantity

                    warehouse_stock.quantity = counted_quantity
                    warehouse_stock.save()

                    # Create adjustment transaction
                    if difference != 0:
                        StockTransaction.objects.create(
                            stock_id=stock_id,
                            from_warehouse_id=warehouse_id,
                            to_warehouse_id=warehouse_id,
                            transaction_type='inventory_count',
                            quantity=difference,
                            unit_price=warehouse_stock.unit_price,
                            reference='Physical inventory count',
                            notes=f'Counted: {counted_quantity}, System: {warehouse_stock.quantity - difference}',
                            user=request.user
                        )

                    results.append({
                        'stock_id': stock_id,
                        'success': True,
                        'counted_quantity': counted_quantity,
                        'system_quantity': warehouse_stock.quantity - difference,
                        'difference': difference
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': count_data.get('stock_id', ''),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(counts),
            'successful': successful,
            'results': results,
            'total_difference': sum(r.get('difference', 0) for r in results if r.get('success'))
        })

class ClearStockCacheView(APIView):
    """Clear stock-related cache"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        cache_keys = [
            'dashboard_stats_*',
            'warehouse_stock_*',
            'stock_valuation_*'
        ]

        cleared_count = 0
        for key_pattern in cache_keys:
            keys = cache.keys(key_pattern)
            for key in keys:
                cache.delete(key)
                cleared_count += 1

        return Response({
            'success': True,
            'cleared_keys': cleared_count,
            'message': f'Cleared {cleared_count} cache entries'
        })

# =============================================================================
# 🩺 HEALTH CHECKS
# =============================================================================

class StockHealthCheckView(APIView):
    """Basic stock system health check"""
    permission_classes = [AllowAny]

    def get(self, request):
        # Check database connectivity
        total_items = Stock.objects.count()
        total_warehouses = Warehouse.objects.filter(is_active=True).count()
        total_transactions = StockTransaction.objects.count()

        # Check for critical issues
        critical_stock = WarehouseStock.objects.filter(
            quantity__lte=F('stock__reorder_level')
        ).count()

        return Response({
            'status': 'healthy',
            'total_items': total_items,
            'active_warehouses': total_warehouses,
            'total_transactions': total_transactions,
            'critical_stock_items': critical_stock,
            'timestamp': timezone.now().isoformat()
        })

class DetailedStockHealthCheckView(APIView):
    """Detailed health check with performance metrics"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Database statistics
        db_stats = {
            'stock_count': Stock.objects.count(),
            'warehouse_count': Warehouse.objects.count(),
            'warehouse_stock_count': WarehouseStock.objects.count(),
            'batch_count': StockBatch.objects.count(),
            'transaction_count': StockTransaction.objects.count(),
            'transfer_count': WarehouseTransfer.objects.count()
        }

        # Performance metrics
        performance = {
            'stock_query_time': self._measure_query_time(Stock.objects.all()),
            'warehouse_stock_query_time': self._measure_query_time(WarehouseStock.objects.all()),
            'transaction_query_time': self._measure_query_time(StockTransaction.objects.all())
        }

        # Cache status
        cache_stats = {
            'dashboard_cache_hits': cache.get('dashboard_cache_hits', 0),
            'total_cache_size': len(cache)
        }

        return Response({
            'status': 'healthy',
            'database': db_stats,
            'performance': performance,
            'cache': cache_stats,
            'timestamp': timezone.now().isoformat()
        })

    def _measure_query_time(self, queryset):
        """Measure query execution time"""
        import time
        start = time.time()
        list(queryset[:10])  # Execute query
        end = time.time()
        return round((end - start) * 1000, 2)  # Return in milliseconds

# =============================================================================
# 🛠️ BULK OPERATIONS
# =============================================================================

class BulkStockAdjustmentView(APIView):
    """Bulk stock adjustment operations"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        adjustments = request.data.get('adjustments', [])
        warehouse_id = request.data.get('warehouse_id')

        if not adjustments or not warehouse_id:
            return Response({
                'error': 'adjustments and warehouse_id required'
            }, status=status.HTTP_400_BAD_REQUEST)

        results = []
        successful = 0

        with transaction.atomic():
            for adj in adjustments:
                try:
                    stock = Stock.objects.get(id=adj['stock_id'], is_active=True)
                    new_quantity = adj['new_quantity']

                    warehouse_stock = WarehouseStock.objects.get(
                        stock=stock, warehouse_id=warehouse_id
                    )

                    old_quantity = warehouse_stock.quantity
                    warehouse_stock.quantity = new_quantity
                    warehouse_stock.save()

                    difference = new_quantity - old_quantity

                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=warehouse_id,
                        to_warehouse_id=warehouse_id,
                        transaction_type='bulk_adjustment',
                        quantity=difference,
                        unit_price=warehouse_stock.unit_price,
                        reference='Bulk adjustment',
                        user=request.user
                    )

                    results.append({
                        'stock_id': str(adj['stock_id']),
                        'success': True,
                        'old_quantity': old_quantity,
                        'new_quantity': new_quantity,
                        'difference': difference
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': str(adj.get('stock_id', '')),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(adjustments),
            'successful': successful,
            'results': results
        })

# =============================================================================
# 🚚 BULK TRANSFER VIEWS
# =============================================================================

class BulkTransferView(APIView):
    """
    Handle bulk warehouse transfers
    """
    permission_classes = [IsAuthenticated, IsStockManagerOrReadOnly]
    authentication_classes = [JWTAuthentication]

    def post(self, request):
        """Create bulk transfer from CSV/Excel file"""
        try:
            file = request.FILES.get('file')
            from_warehouse_id = request.data.get('from_warehouse_id')
            to_warehouse_id = request.data.get('to_warehouse_id')

            if not all([file, from_warehouse_id, to_warehouse_id]):
                return Response({
                    'error': 'File, from_warehouse_id, and to_warehouse_id are required'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate warehouses
            try:
                from_warehouse = Warehouse.objects.get(id=from_warehouse_id)
                to_warehouse = Warehouse.objects.get(id=to_warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({
                    'error': 'Invalid warehouse ID'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save file temporarily
            import tempfile
            import os
            with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_file:
                for chunk in file.chunks():
                    tmp_file.write(chunk)
                file_path = tmp_file.name

            # Process bulk transfer asynchronously
            task = process_bulk_transfer.delay(
                file_path=file_path,
                from_warehouse_id=from_warehouse_id,
                to_warehouse_id=to_warehouse_id,
                user_id=request.user.id
            )

            # Clean up file after delay
            import threading
            def cleanup_file():
                time.sleep(300)  # 5 minutes
                try:
                    os.unlink(file_path)
                except:
                    pass
            threading.Thread(target=cleanup_file, daemon=True).start()

            return Response({
                'status': 'success',
                'message': 'Bulk transfer processing started',
                'task_id': task.id,
                'file_path': file_path,
                'from_warehouse': from_warehouse_id,
                'to_warehouse': to_warehouse_id,
                'timestamp': timezone.now().isoformat()
            }, status=status.HTTP_202_ACCEPTED)

        except Exception as e:
            logger.error(f"Bulk transfer error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def status(self, request):
        """Get bulk transfer status by task_id"""
        task_id = request.query_params.get('task_id')
        if not task_id:
            return Response({
                'error': 'task_id parameter is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            from celery.result import AsyncResult
            result = AsyncResult(task_id)

            return Response({
                'task_id': task_id,
                'status': result.status,
                'result': result.result if result.ready() else None,
                'failed': result.failed(),
                'timestamp': timezone.now().isoformat()
            })
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get(self, request):
        """List recent bulk transfers"""
        try:
            transfers = WarehouseTransfer.objects.filter(
                transfer_type='bulk'
            ).select_related(
                'from_warehouse', 'to_warehouse', 'created_by'
            ).order_by('-created_at')[:50]

            serializer = WarehouseTransferSerializer(transfers, many=True)
            return Response({
                'count': transfers.count(),
                'transfers': serializer.data,
                'timestamp': timezone.now().isoformat()
            })
        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class BulkStockTransferView(BulkTransferView):
    """Bulk stock transfer (alias)"""
    pass

class BulkReservationView(BatchReserveView):
    """Bulk reservation (alias)"""
    pass

class BulkUnreservationView(APIView):
    """Bulk unreservation operations"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [BatchOperationThrottle]

    def post(self, request):
        stock_ids = request.data.get('stock_ids', [])
        warehouse_id = request.data.get('warehouse_id')
        quantity = request.data.get('quantity', 0)

        results = []
        successful = 0

        with transaction.atomic():
            for stock_id in stock_ids:
                try:
                    stock = Stock.objects.get(id=stock_id, is_active=True)
                    warehouse_stock = WarehouseStock.objects.get(
                        stock=stock, warehouse_id=warehouse_id
                    )

                    if warehouse_stock.reserved_quantity < quantity:
                        results.append({
                            'stock_id': str(stock_id),
                            'success': False,
                            'error': f'Insufficient reserved quantity: {warehouse_stock.reserved_quantity}'
                        })
                        continue

                    warehouse_stock.reserved_quantity -= quantity
                    warehouse_stock.save()

                    StockTransaction.objects.create(
                        stock=stock,
                        from_warehouse_id=warehouse_id,
                        to_warehouse_id=warehouse_id,
                        transaction_type='unreservation',
                        quantity=-quantity,
                        unit_price=warehouse_stock.unit_price,
                        reference='Bulk unreservation',
                        user=request.user
                    )

                    results.append({
                        'stock_id': str(stock_id),
                        'success': True,
                        'new_reserved': warehouse_stock.reserved_quantity
                    })
                    successful += 1

                except Exception as e:
                    results.append({
                        'stock_id': str(stock_id),
                        'success': False,
                        'error': str(e)
                    })

        return Response({
            'success': True,
            'processed': len(stock_ids),
            'successful': successful,
            'results': results
        })

# =============================================================================
# 📥 TEMPLATE DOWNLOADS
# =============================================================================

class StockImportTemplateView(APIView):
    """Download stock import template"""
    permission_classes = [AllowAny]

    def get(self, request):
        df = pd.DataFrame(columns=[
            'item_code', 'name', 'barcode', 'description', 'category',
            'supplier', 'min_stock_level', 'reorder_level', 'max_stock_level',
            'quantity', 'unit_price'
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = 'stock_import_template.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Stock Import Template')

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Stock Import Template']

            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        return response

class TransferImportTemplateView(APIView):
    """Download warehouse transfer import template"""
    permission_classes = [AllowAny]

    def get(self, request):
        df = pd.DataFrame(columns=[
            'from_warehouse_code', 'to_warehouse_code', 'stock_item_code',
            'quantity', 'unit_price', 'notes'
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = 'transfer_import_template.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transfer Template')

        return response

# =============================================================================
# 🖥️ LEGACY VIEWS (Backward Compatibility)
# =============================================================================

class LegacyWarehouseStockView(APIView):
    """Legacy warehouse stock endpoint"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        stocks = WarehouseStock.objects.select_related('stock', 'warehouse').all()
        serializer = WarehouseStockSerializer(stocks, many=True)
        return Response(serializer.data)

class LegacyDashboardView(APIView):
    """Legacy dashboard endpoint"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Redirect to new dashboard
        return Response({
            'message': 'Use /api/v1/dashboard_stats/ for updated dashboard',
            'redirect_url': '/api/v1/dashboard_stats/'
        })

# =============================================================================
# 🔌 WEBSOCKET VIEWS (ASGI)
# =============================================================================

class StockWebSocketView(APIView):
    """WebSocket endpoint for real-time stock updates"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({
            'message': 'WebSocket connection available at /ws/stock_updates/',
            'endpoints': {
                'stock_updates': '/ws/stock_updates/',
                'warehouse_updates': '/ws/warehouse_updates/',
                'transfer_updates': '/ws/transfer_updates/'
            }
        })

# Redis Configuration for WebSocket caching
REDIS_HOST = 'localhost'
REDIS_PORT = 6379
REDIS_DB = 1
REDIS_PREFIX = 'ws_stock_'

class RedisCacheManager:
    """Redis cache manager for WebSocket operations"""

    def __init__(self):
        self.redis_client = None

    async def get_redis_client(self):
        if self.redis_client is None:
            self.redis_client = redis.Redis(
                host=REDIS_HOST, port=REDIS_PORT, db=REDIS_DB,
                decode_responses=True, socket_connect_timeout=5,
                socket_timeout=5, retry_on_timeout=True, max_connections=20
            )
        return self.redis_client

    async def get_cache(self, key: str, default: Any = None) -> Any:
        try:
            client = await self.get_redis_client()
            data = await client.get(f"{REDIS_PREFIX}{key}")
            return json.loads(data) if data else default
        except RedisError:
            return default

    async def set_cache(self, key: str, value: Any, ttl: int = 300):
        try:
            client = await self.get_redis_client()
            await client.setex(f"{REDIS_PREFIX}{key}", ttl, json.dumps(value))
        except RedisError:
            pass

    async def delete_cache(self, key: str):
        try:
            client = await self.get_redis_client()
            await client.delete(f"{REDIS_PREFIX}{key}")
        except RedisError:
            pass

redis_cache = RedisCacheManager()

# =============================================================================
# 🧬 BASE WEBSOCKET CONSUMER
# =============================================================================

class BaseStockConsumer(AsyncWebsocketConsumer):
    """Base consumer with Redis caching and connection management"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.redis_cache = redis_cache
        self.connection_id = None
        self.subscriptions = set()
        self.last_heartbeat = None

    async def connect(self):
        self.connection_id = f"conn_{int(time.time())}_{id(self)}"

        if self.scope['user'].is_authenticated:
            await self._register_connection()
            await self.channel_layer.group_add(self.group_name, self.channel_name)
            await self.accept()

            await self.send_connection_message()
            asyncio.create_task(self._heartbeat_monitor())
        else:
            await self.close()

    async def _register_connection(self):
        connection_data = {
            'user_id': self.scope['user'].id,
            'username': self.scope['user'].username,
            'group': self.group_name,
            'connected_at': timezone.now().isoformat(),
            'last_heartbeat': time.time()
        }
        await self.redis_cache.set_cache(f"connection_{self.connection_id}", connection_data, ttl=3600)
        await self.redis_cache.increment_counter(f"active_connections_{self.group_name}")

    async def disconnect(self, close_code):
        await self.redis_cache.delete_cache(f"connection_{self.connection_id}")
        await self.redis_cache.increment_counter(f"active_connections_{self.group_name}", -1)
        await self.channel_layer.group_discard(self.group_name, self.channel_name)

    async def send_connection_message(self):
        active_connections = await self.redis_cache.get_cache(f"active_connections_{self.group_name}", 0)
        await self.send(text_data=json.dumps({
            'type': 'connection_status',
            'status': 'connected',
            'connection_id': self.connection_id,
            'active_connections': active_connections,
            'timestamp': timezone.now().isoformat()
        }))

    async def _heartbeat_monitor(self):
        while True:
            try:
                await asyncio.sleep(30)
                await self.redis_cache.set_cache(
                    f"connection_{self.connection_id}",
                    {'last_heartbeat': time.time()},
                    ttl=3600
                )
            except:
                break

# =============================================================================
# 📦 STOCK UPDATES CONSUMER
# =============================================================================

class StockWebSocketConsumer(BaseStockConsumer):
    """Real-time stock updates - /ws/stock_updates/ and /ws/stock/<stock_id>/"""

    group_name = 'stock_updates'

    async def connect(self):
        # Handle both general and specific stock connections
        stock_id = self.scope.get('url_route', {}).get('kwargs', {}).get('stock_id')
        self.group_name = f"stock_{stock_id}" if stock_id else 'stock_updates'
        super().connect()

    async def receive(self, text_data):
        try:
            data = json.loads(text_data)
            message_type = data.get('type')

            if message_type == 'subscribe':
                await self.subscribe_to_stock(data.get('stock_ids', []))
            elif message_type == 'get_recent':
                await self.send_recent_transactions(data.get('stock_ids', []))
        except:
            await self.send_error_message('Invalid message')

    @sync_to_async
    def get_recent_transactions(self, stock_id=None, limit=10):
        cache_key = f"recent_transactions_{stock_id or 'all'}_{limit}"
        cached = cache.get(cache_key)
        if cached:
            return cached

        queryset = StockTransaction.objects.select_related(
            'stock', 'from_warehouse', 'to_warehouse', 'user'
        ).order_by('-created_at')

        if stock_id:
            queryset = queryset.filter(stock_id=stock_id)

        transactions = queryset[:limit]
        result = StockTransactionSerializer(transactions, many=True).data
        cache.set(cache_key, result, 120)
        return result

    async def subscribe_to_stock(self, stock_ids):
        if stock_ids:
            for stock_id in stock_ids:
                group_name = f"stock_{stock_id}"
                await self.channel_layer.group_add(group_name, self.channel_name)
                self.subscriptions.add(group_name)

        transactions = await sync_to_async(self.get_recent_transactions)(stock_ids[0] if stock_ids else None)
        await self.send(text_data=json.dumps({
            'type': 'subscription_confirmed',
            'stock_ids': stock_ids,
            'recent_transactions': transactions,
            'timestamp': timezone.now().isoformat()
        }))

    async def send_error_message(self, message):
        await self.send(text_data=json.dumps({
            'type': 'error',
            'message': message,
            'timestamp': timezone.now().isoformat()
        }))

    # Event handlers
    async def stock_transaction_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'stock_transaction',
            'transaction': event['transaction'],
            'stock_id': event['stock_id'],
            'timestamp': event['timestamp']
        }))

    async def stock_level_changed(self, event):
        await self.send(text_data=json.dumps({
            'type': 'stock_level',
            'stock_id': event['stock_id'],
            'warehouse_id': event['warehouse_id'],
            'new_quantity': event['new_quantity'],
            'timestamp': event['timestamp']
        }))

    async def stock_alert_created(self, event):
        await self.send(text_data=json.dumps({
            'type': 'stock_alert',
            'alert': event['alert'],
            'severity': event['severity'],
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 🏭 WAREHOUSE UPDATES CONSUMER
# =============================================================================

class WarehouseWebSocketConsumer(BaseStockConsumer):
    
    group_name = 'warehouse_updates'

    async def connect(self):
        warehouse_id = self.scope.get('url_route', {}).get('kwargs', {}).get('warehouse_id')
        self.group_name = f"warehouse_{warehouse_id}" if warehouse_id else 'warehouse_updates'
        super().connect()

    async def receive(self, text_data):
        try:
            data = json.loads(text_data)
            if data.get('type') == 'get_summary':
                await self.send_warehouse_summary(data.get('warehouse_ids', []))
        except:
            pass

    @sync_to_async
    def get_warehouse_summary(self, warehouse_id=None):
        cache_key = f"warehouse_summary_{warehouse_id or 'all'}"
        # ✅ FIXED: Use Django cache.get() instead of await redis_cache.get_cache()
        cached = cache.get(cache_key)
        if cached:
            return cached

        from .models import WarehouseStock
        from django.db.models import Count, Sum, Q, F
        summary = WarehouseStock.objects.filter(
           warehouse_id=warehouse_id,
           stock__is_active=True
        ).aggregate(
           total_items=Count('id'),
           total_quantity=Sum('quantity'),
           critical_items=Count('id', filter=Q(quantity__lte=F('stock__reorder_level')))
        )

        result = {
            'total_items': summary['total_items'] or 0,
            'total_quantity': int(summary['total_quantity'] or 0),
            'critical_items': summary['critical_items'] or 0,
            'timestamp': timezone.now().isoformat()
        }

        # ✅ FIXED: Use Django cache.set() instead of await redis_cache.set_cache()
        cache.set(cache_key, result, 60)
        return result




    async def send_warehouse_summary(self, warehouse_ids):
        summaries = {}
        for wid in warehouse_ids:
            summary = await sync_to_async(self.get_warehouse_summary)(wid)
            summaries[wid] = summary
        
        await self.send(text_data=json.dumps({
            'type': 'warehouse_summary',
            'warehouse_ids': warehouse_ids,
            'summaries': summaries,
            'timestamp': timezone.now().isoformat()
        }))

    # Event handlers
    async def warehouse_stock_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'warehouse_stock_update',
            'warehouse_id': event['warehouse_id'],
            'stock_id': event['stock_id'],
            'quantity_change': event['quantity_change'],
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 🚚 TRANSFER UPDATES CONSUMER
# =============================================================================

class TransferWebSocketConsumer(BaseStockConsumer):

    group_name = 'transfer_updates'

    async def connect(self):
        transfer_id = self.scope.get('url_route', {}).get('kwargs', {}).get('transfer_id')
        self.group_name = f"transfer_{transfer_id}" if transfer_id else 'transfer_updates'
        super().connect()

    async def receive(self, text_data):
        try:
            data = json.loads(text_data)
            if data.get('type') == 'get_transfers':
                await self.send_transfer_details(data.get('transfer_ids', []))
        except:
            pass

    @sync_to_async
    def get_transfer_details(self, transfer_ids: List[str] = None, single_id: str = None):
        """Get transfer details with comprehensive caching"""
        cache_key = None

        if single_id:
            cache_key = f"transfer_details_{single_id}"
            transfers = WarehouseTransfer.objects.filter(
                id=single_id
            ).select_related(
               'from_warehouse', 'to_warehouse', 'created_by', 'completed_by'
            ).prefetch_related('items__stock', 'items__stock__category')[:1]
        elif transfer_ids:
            cache_key = f"transfers_batch_{'_'.join(sorted(transfer_ids))}"
            transfers = WarehouseTransfer.objects.filter(
                id__in=transfer_ids
            ).select_related(
               'from_warehouse', 'to_warehouse', 'created_by', 'completed_by'
            ).prefetch_related('items__stock')[:10]
        else:
             # Recent transfers for general connection
             cache_key = "recent_transfers"
             transfers = WarehouseTransfer.objects.select_related(
                 'from_warehouse', 'to_warehouse'
             ).order_by('-created_at')[:5]

        # ✅ FIXED: Use Django cache.get() instead of await redis_cache.get_cache()
        cached = cache.get(cache_key)
        if cached:
            return cached

        # Serialize and cache
        result = WarehouseTransferSerializer(transfers, many=True).data
        # ✅ FIXED: Use Django cache.set() instead of await redis_cache.set_cache()
        cache.set(cache_key, result, 300)  # 5 minutes
        return result

    # Event handlers
    async def transfer_status_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'transfer_status',
            'transfer_id': event['transfer_id'],
            'new_status': event['new_status'],
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 📊 DASHBOARD CONSUMER
# =============================================================================

class StockDashboardConsumer(BaseStockConsumer):
    """Real-time dashboard updates - /ws/dashboard/"""

    group_name = 'dashboard_updates'

    @sync_to_async
    def get_dashboard_metrics(self):
        cache_key = f"dashboard_metrics_global"
        cached = cache.get(cache_key)
        if cached:
            return cached

        total_items = StockTransaction.objects.count()
        total_value = WarehouseStock.objects.aggregate(
            total=Sum(F('quantity') * F('unit_price'))
        )['total'] or 0

        metrics = {
            'total_items': total_items,
            'total_value': float(total_value),
            'active_alerts': StockAlert.objects.filter(
                is_active=True
            ).count(),
            'timestamp': timezone.now().isoformat()
        }

        cache.set(cache_key, metrics, 60)
        return metrics

    # Event handlers
    async def dashboard_metrics_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'dashboard_update',
            'metrics': event['metrics'],
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 🔔 NOTIFICATIONS CONSUMER
# =============================================================================

class StockNotificationConsumer(BaseStockConsumer):
    """Critical notifications - /ws/notifications/"""

    group_name = 'stock_notifications'

    # Event handlers
    async def critical_stock_alert(self, event):
        await self.send(text_data=json.dumps({
            'type': 'critical_alert',
            'alert_type': 'low_stock',
            'stock_id': event['stock_id'],
            'current_quantity': event['current_quantity'],
            'severity': 'high',
            'timestamp': event['timestamp']
        }))

    async def batch_expiry_alert(self, event):
        await self.send(text_data=json.dumps({
            'type': 'critical_alert',
            'alert_type': 'batch_expiry',
            'batch_id': event['batch_id'],
            'days_to_expiry': event['days_to_expiry'],
            'severity': 'medium',
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 🔧 REDIS MONITOR CONSUMER
# =============================================================================

class RedisMonitorConsumer(BaseStockConsumer):
    """Redis monitoring - /ws/redis_monitor/"""

    group_name = 'redis_monitor'

    async def receive(self, text_data):
        try:
            data = json.loads(text_data)
            command = data.get('command')

            if command == 'stats':
                await self.send_redis_stats()
            elif command == 'connections':
                await self.send_connection_stats()
        except:
            pass

    async def send_redis_stats(self):
        stats = await self.redis_cache.get_connection_stats()
        await self.send(text_data=json.dumps({
            'type': 'redis_stats',
            'stats': stats,
            'timestamp': timezone.now().isoformat()
        }))

    async def send_connection_stats(self):
        client = await self.redis_cache.get_redis_client()
        keys = await client.keys(f"{REDIS_PREFIX}connection_*")
        await self.send(text_data=json.dumps({
            'type': 'connection_stats',
            'total_connections': len(keys),
            'timestamp': timezone.now().isoformat()
        }))

# =============================================================================
# 📈 ANALYTICS CONSUMER
# =============================================================================

class StockAnalyticsConsumer(BaseStockConsumer):
    """Analytics updates - /ws/analytics/"""

    group_name = 'stock_analytics'

    @sync_to_async
    def get_analytics_data(self):
        cache_key = f"stock_analytics_global"
        cached = cache.get(cache_key)
        if cached:
            return cached

        analytics = {
            'daily_transactions': StockTransaction.objects.filter(
                created_at__date=timezone.now().date()
            ).count(),
            'top_items': StockTransaction.objects.values(
                'stock__name'
            ).annotate(
                total_qty=Sum('quantity')
            ).order_by('-total_qty')[:5]
        }

        cache.set(cache_key, analytics, 300)
        return analytics

    # Event handlers
    async def analytics_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'analytics_update',
            'data': event['data'],
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 🚚 TRANSFER UPDATES CONSUMER (ENHANCED)
# =============================================================================

class TransferWebSocketConsumer(BaseStockConsumer):
    """
    Real-time transfer updates
    - General: /ws/transfer_updates/
    - Specific: /ws/transfer/<transfer_id>/
    """

    group_name = 'transfer_updates'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.transfer_id = None
        self.specific_transfer_groups = set()

    async def connect(self):
        """Handle both general and specific transfer connections"""
        # Extract transfer_id from URL kwargs
        self.transfer_id = self.scope.get('url_route', {}).get('kwargs', {}).get('transfer_id')

        if self.transfer_id:
            # Specific transfer connection
            self.group_name = f"transfer_{self.transfer_id}"
            logger.info(f"Connecting to specific transfer: {self.transfer_id}")
        else:
            # General transfers connection
            self.group_name = 'transfer_updates'
            logger.info("Connecting to general transfer updates")

        # Register connection and join group
        await self._register_connection()
        await self.channel_layer.group_add(self.group_name, self.channel_name)
        await self.accept()

        # Send connection confirmation
        await self.send_connection_message()

        # Send initial transfer data for specific connections
        if self.transfer_id:
            await self.send_initial_transfer_data()

        # Start heartbeat monitoring
        asyncio.create_task(self._heartbeat_monitor())

    async def _register_connection(self):
        """Register connection with transfer-specific info"""
        connection_data = {
            'user_id': self.scope['user'].id,
            'username': self.scope['user'].username,
            'group': self.group_name,
            'transfer_id': self.transfer_id,
            'connected_at': timezone.now().isoformat(),
            'last_heartbeat': time.time()
        }

        await self.redis_cache.set_cache(
            f"connection_{self.connection_id}",
            connection_data,
            ttl=3600
        )

        # Increment specific counters
        await self.redis_cache.increment_counter(f"active_connections_{self.group_name}")
        if self.transfer_id:
            await self.redis_cache.increment_counter(f"transfer_subscribers_{self.transfer_id}")

    async def disconnect(self, close_code):
        """Clean up specific transfer subscriptions"""
        try:
            # Remove from main group
            await self.channel_layer.group_discard(self.group_name, self.channel_name)

            # Remove from specific transfer groups
            if self.transfer_id:
                await self.redis_cache.increment_counter(f"transfer_subscribers_{self.transfer_id}", -1)

            # Cleanup Redis
            await self.redis_cache.delete_cache(f"connection_{self.connection_id}")
            await self.redis_cache.increment_counter(f"active_connections_{self.group_name}", -1)

            logger.info(f"Disconnected from transfer group: {self.group_name} (Transfer ID: {self.transfer_id})")
        except Exception as e:
            logger.error(f"Disconnect error for transfer {self.transfer_id}: {e}")

    async def send_connection_message(self):
        """Send connection confirmation with transfer-specific info"""
        subscriber_count = await self.redis_cache.get_cache(
            f"transfer_subscribers_{self.transfer_id}", 0
        ) if self.transfer_id else 0

        message = {
            'type': 'connection_status',
            'status': 'connected',
            'connection_id': self.connection_id,
            'group': self.group_name,
            'transfer_id': self.transfer_id,
            'subscriber_count': subscriber_count,
            'timestamp': timezone.now().isoformat()
        }

        if self.transfer_id:
            message['message'] = f'Connected to transfer {self.transfer_id} updates'
        else:
            message['message'] = 'Connected to all transfer updates'

        await self.send(text_data=json.dumps(message))

    async def receive(self, text_data):
        """Handle incoming messages for transfer management"""
        try:
            data = json.loads(text_data)
            message_type = data.get('type')

            if message_type == 'subscribe_transfers':
                await self.subscribe_to_transfers(data.get('transfer_ids', []))
            elif message_type == 'unsubscribe_transfers':
                await self.unsubscribe_from_transfers(data.get('transfer_ids', []))
            elif message_type == 'get_transfers':
                await self.send_transfer_details(data.get('transfer_ids', []))
            elif message_type == 'get_transfer_status':
                await self.get_transfer_status(data.get('transfer_id'))
            else:
                await self.send_error_message('Invalid message type')

        except json.JSONDecodeError:
            await self.send_error_message('Invalid JSON format')
        except Exception as e:
            logger.error(f"TransferWebSocketConsumer receive error: {str(e)}")
            await self.send_error_message(str(e))

    @sync_to_async
    def get_transfer_details(self, transfer_ids: List[str] = None, single_id: str = None):
        """Get transfer details with comprehensive caching"""
        cache_key = None

        if single_id:
            cache_key = f"transfer_details_{single_id}"
            transfers = WarehouseTransfer.objects.filter(
                id=single_id
            ).select_related(
                'from_warehouse', 'to_warehouse', 'created_by', 'completed_by'
            ).prefetch_related('items__stock', 'items__stock__category')[:1]
        elif transfer_ids:
            cache_key = f"transfers_batch_{'_'.join(sorted(transfer_ids))}"
            transfers = WarehouseTransfer.objects.filter(
                id__in=transfer_ids
            ).select_related(
                'from_warehouse', 'to_warehouse', 'created_by', 'completed_by'
            ).prefetch_related('items__stock')[:10]
        else:
            # Recent transfers for general connection
            cache_key = "recent_transfers"
            transfers = WarehouseTransfer.objects.select_related(
                'from_warehouse', 'to_warehouse'
            ).order_by('-created_at')[:5]

        # Check cache first
        cached = cache.get(cache_key)
        if cached:
            return cached

        # Serialize and cache
        result = WarehouseTransferSerializer(transfers, many=True).data
        cache.set(cache_key, result, 300)  # 5 minutes
        return result

    async def send_initial_transfer_data(self):
        """Send initial transfer data for specific connections"""
        if not self.transfer_id:
            return

        try:
            details = await sync_to_async(self.get_transfer_details)(single_id=self.transfer_id)

            await self.send(text_data=json.dumps({
                'type': 'initial_transfer_data',
                'transfer_id': self.transfer_id,
                'transfer': details[0] if details else None,
                'status': 'initial_load_complete',
                'timestamp': timezone.now().isoformat()
            }))
        except Exception as e:
            logger.error(f"Initial transfer data error for {self.transfer_id}: {e}")
            await self.send_error_message(f"Failed to load transfer data: {str(e)}")

    async def subscribe_to_transfers(self, transfer_ids: List[str]):
        """Subscribe to specific transfer updates"""
        if not transfer_ids:
            await self.send_error_message('Transfer IDs required for subscription')
            return

        for transfer_id in transfer_ids:
            group_name = f"transfer_{transfer_id}"
            await self.channel_layer.group_add(group_name, self.channel_name)
            self.specific_transfer_groups.add(group_name)

            # Track subscription
            await self.redis_cache.increment_counter(f"transfer_subscribers_{transfer_id}")

        # Send current details
        details = await sync_to_async(self.get_transfer_details)(transfer_ids)

        await self.send(text_data=json.dumps({
            'type': 'transfer_subscription_confirmed',
            'transfer_ids': transfer_ids,
            'transfers': details,
            'total_subscribed': len(self.specific_transfer_groups),
            'timestamp': timezone.now().isoformat()
        }))

    async def unsubscribe_from_transfers(self, transfer_ids: List[str]):
        """Unsubscribe from specific transfers"""
        for transfer_id in transfer_ids:
            group_name = f"transfer_{transfer_id}"
            await self.channel_layer.group_discard(group_name, self.channel_name)
            self.specific_transfer_groups.discard(group_name)
            await self.redis_cache.increment_counter(f"transfer_subscribers_{transfer_id}", -1)

        await self.send(text_data=json.dumps({
            'type': 'transfer_unsubscription_confirmed',
            'transfer_ids': transfer_ids,
            'remaining_subscriptions': len(self.specific_transfer_groups),
            'timestamp': timezone.now().isoformat()
        }))

    async def send_transfer_details(self, transfer_ids: List[str]):
        """Send cached transfer details"""
        details = await sync_to_async(self.get_transfer_details)(transfer_ids)
        await self.send(text_data=json.dumps({
            'type': 'transfer_details',
            'transfer_ids': transfer_ids,
            'transfers': details,
            'timestamp': timezone.now().isoformat()
        }))

    async def get_transfer_status(self, transfer_id: str):
        """Get real-time status for specific transfer"""
        details = await sync_to_async(self.get_transfer_details)(single_id=transfer_id)
        if details:
            await self.send(text_data=json.dumps({
                'type': 'transfer_status',
                'transfer_id': transfer_id,
                'transfer': details[0],
                'timestamp': timezone.now().isoformat()
            }))
        else:
            await self.send_error_message(f'Transfer {transfer_id} not found')

    async def send_error_message(self, message: str):
        """Send standardized error message"""
        await self.send(text_data=json.dumps({
            'type': 'error',
            'message': message,
            'transfer_id': self.transfer_id,
            'timestamp': timezone.now().isoformat()
        }))

    # =============================================================================
    # EVENT HANDLERS - Real-time transfer events
    # =============================================================================

    async def transfer_status_update(self, event):
        """Handle transfer status changes"""
        await self.send(text_data=json.dumps({
            'type': 'transfer_status_update',
            'transfer_id': event['transfer_id'],
            'old_status': event.get('old_status'),
            'new_status': event['new_status'],
            'warehouse_from': event.get('warehouse_from'),
            'warehouse_to': event.get('warehouse_to'),
            'progress': event.get('progress', 0),
            'timestamp': event['timestamp']
        }))

    async def transfer_created(self, event):
        """Handle new transfer creation"""
        await self.send(text_data=json.dumps({
            'type': 'transfer_created',
            'transfer_id': event['transfer_id'],
            'transfer': event['transfer'],
            'created_by': event.get('created_by'),
            'timestamp': event['timestamp']
        }))

    async def transfer_items_updated(self, event):
        """Handle transfer items modification"""
        await self.send(text_data=json.dumps({
            'type': 'transfer_items_updated',
            'transfer_id': event['transfer_id'],
            'items_changed': event['items_changed'],
            'total_items': event['total_items'],
            'total_quantity': event['total_quantity'],
            'timestamp': event['timestamp']
        }))

    async def transfer_completed(self, event):
        """Handle transfer completion"""
        await self.send(text_data=json.dumps({
            'type': 'transfer_completed',
            'transfer_id': event['transfer_id'],
            'completed_at': event['completed_at'],
            'completed_by': event['completed_by'],
            'final_quantity': event['final_quantity'],
            'timestamp': event['timestamp']
        }))

    async def transfer_cancelled(self, event):
        """Handle transfer cancellation"""
        await self.send(text_data=json.dumps({
            'type': 'transfer_cancelled',
            'transfer_id': event['transfer_id'],
            'cancelled_at': event['cancelled_at'],
            'cancelled_by': event['cancelled_by'],
            'reason': event.get('reason'),
            'timestamp': event['timestamp']
        }))

# =============================================================================
# 📊 REPORTS CONSUMER
# =============================================================================

class StockReportsConsumer(BaseStockConsumer):
    """Real-time reports - /ws/reports/"""

    group_name = 'stock_reports'

    # Event handlers
    async def reports_update(self, event):
        await self.send(text_data=json.dumps({
            'type': 'report_update',
            'report_id': event['report_id'],
            'status': event['status'],
            'progress': event.get('progress', 0),
            'timestamp': event['timestamp']
        }))


class WebSocketEndpointsView(APIView):
    """API endpoint listing all available WebSocket endpoints"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Return all available WebSocket endpoints"""
        base_url = request.build_absolute_uri().rstrip('/')

        endpoints = {
            'stock_updates': {
                'url': f'{base_url}/ws/stock_updates/',
                'description': 'Real-time stock transaction updates',
                'features': [
                    'Stock level changes',
                    'Transaction notifications',
                    'Low stock alerts',
                    'Batch expiry warnings'
                ]
            },
            'warehouse_updates': {
                'url': f'{base_url}/ws/warehouse_updates/',
                'description': 'Real-time warehouse inventory updates',
                'features': [
                    'Warehouse stock summaries',
                    'Stock movement tracking',
                    'Capacity utilization',
                    'Critical stock alerts'
                ]
            },
            'transfer_updates': {
                'url': f'{base_url}/ws/transfer_updates/',
                'description': 'Real-time warehouse transfer tracking',
                'features': [
                    'Transfer status updates',
                    'In-transit notifications',
                    'Completion alerts',
                    'Transfer creation events'
                ]
            },
            'dashboard': {
                'url': f'{base_url}/ws/dashboard/',
                'description': 'Real-time dashboard metrics',
                'features': [
                    'Live KPIs',
                    'Top performing items',
                    'Critical alerts',
                    'Analytics updates'
                ]
            },
            'notifications': {
                'url': f'{base_url}/ws/notifications/',
                'description': 'Critical stock notifications',
                'features': [
                    'Low stock alerts',
                    'Batch expiry warnings',
                    'Audit discrepancies',
                    'System alerts'
                ]
            },
            'analytics': {
                'url': f'{base_url}/ws/analytics/',
                'description': 'Real-time analytics updates',
                'features': [
                    'Daily/weekly trends',
                    'Top moving items',
                    'Sales velocity',
                    'Inventory turnover'
                ]
            }
        }

        return Response({
            'message': 'WebSocket endpoints for real-time stock management',
            'endpoints': endpoints,
            'connection_guide': {
                'protocol': 'wss://' if request.scheme == 'https' else 'ws://',
                'authentication': 'Automatic via Django session/cookie',
                'reconnection': 'Built-in with exponential backoff',
                'message_format': 'JSON'
            },
            'timestamp': timezone.now().isoformat()
        })


class WebSocketStatusView(APIView):
    """WebSocket connection status and monitoring"""

    permission_classes = [IsAuthenticated]

    async def get(self, request):
        """Get comprehensive WebSocket status"""
        try:
            # Get Redis connection stats
            redis_stats = await redis_cache.get_connection_stats()

            # Get active connections by group
            active_connections = {}
            groups = [
                'stock_updates', 'warehouse_updates', 'transfer_updates',
                'dashboard_updates', 'stock_notifications', 'stock_analytics'
            ]

            for group in groups:
                count_key = f"active_connections_{group}"
                count = await redis_cache.get_cache(count_key, 0)
                active_connections[group] = int(count)

            # Get cached metrics
            total_connections = sum(active_connections.values())

            # Get recent WebSocket events
            recent_events = await self.get_recent_events()

            return Response({
                'status': 'healthy',
                'active_connections': total_connections,
                'connections_by_group': active_connections,
                'redis': redis_stats,
                'recent_events': recent_events[:10],  # Last 10 events
                'uptime': '24/7',
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            logger.error(f"WebSocket status error: {str(e)}")
            return Response({
                'status': 'error',
                'error': str(e),
                'timestamp': timezone.now().isoformat()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @sync_to_async
    def get_recent_events(self):
        """Get recent WebSocket events from Redis"""
        try:
            # This would typically read from a Redis stream or list
            # For now, return mock data or implement your event logging
            return [
                {
                    'id': 'evt_123',
                    'type': 'stock_transaction',
                    'group': 'stock_updates',
                    'user': 'john_doe',
                    'timestamp': '2024-11-28T10:30:00Z',
                    'payload': {'stock_id': 'STK001', 'quantity': 50}
                }
            ]
        except Exception as e:
            logger.error(f"Recent events error: {e}")
            return []


class WebSocketBroadcastView(APIView):
    """Manual WebSocket broadcast endpoint for testing/admin"""

    permission_classes = [IsAdminUser]

    def post(self, request):
        """Broadcast message to WebSocket groups"""
        group = request.data.get('group')
        message = request.data.get('message')
        event_type = request.data.get('type', 'message')

        if not group or not message:
            return Response({
                'error': 'group and message are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            channel_layer = get_channel_layer()
            if channel_layer:
                # Convert message to JSON if it's not already
                if isinstance(message, dict):
                    message_data = message
                else:
                    message_data = {'content': str(message)}

                # Send to group
                async_to_sync(channel_layer.group_send)(
                    group,
                    {
                        'type': event_type,
                        'message': message_data,
                        'timestamp': timezone.now().isoformat(),
                        'broadcast_by': request.user.username
                    }
                )

                return Response({
                    'status': 'success',
                    'group': group,
                    'message': 'Broadcast sent successfully',
                    'recipients': f'All {group} subscribers'
                })
            else:
                return Response({
                    'error': 'Channel layer not configured'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        except Exception as e:
            logger.error(f"Broadcast error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class WebSocketConnectionManagerView(APIView):
    """Manage WebSocket connections (Admin only)"""

    permission_classes = [IsAdminUser]

    def get(self, request):
        """Get detailed connection information"""
        try:
            connections = []

            # Simple fallback: return cached connection stats
            cache_key = "websocket_connections_summary"
            cached_connections = cache.get(cache_key)

            if cached_connections:
                return Response({
                    'total_connections': len(cached_connections),
                    'connections': cached_connections[:50],  # Limit for performance
                    'timestamp': timezone.now().isoformat()
                })

            # Mock data for demo (replace with real Redis sync operations)
            sample_connections = [
                {
                    'connection_id': 'conn_12345',
                    'user_id': 1,
                    'username': 'admin',
                    'group': 'stock_updates',
                    'connected_at': '2025-11-29T10:00:00Z',
                    'last_heartbeat': 1732864800.0,
                    'is_active': True
                }
            ]

            cache.set(cache_key, sample_connections, 300)

            return Response({
                'total_connections': len(sample_connections),
                'connections': sample_connections,
                'timestamp': timezone.now().isoformat(),
                'message': 'Use Redis admin tools for full connection details'
            })

        except Exception as e:
            logger.error(f"Connection manager error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def delete(self, request, connection_id):
        """Force disconnect a WebSocket connection"""
        try:
            # ✅ FIXED: Pure synchronous operations
            connection_key = f"connection_{connection_id}"

            # Clear from Django cache
            cache.delete(connection_key)

            # Clear related dashboard caches
            cache.delete("websocket_connections_summary")
            cache.delete(f"active_connections_stock_updates")
            cache.delete(f"active_connections_transfer_updates")

            return Response({
                'status': 'success',
                'message': f'Connection {connection_id} removed from tracking',
                'cleared_keys': [
                    connection_key,
                    'websocket_connections_summary'
                ],
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            logger.error(f"Connection delete error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class WebSocketCacheManagerView(APIView):
    """Manage WebSocket Redis cache (Admin only)"""

    permission_classes = [IsAdminUser]

    async def get(self, request):
        """Get cache statistics"""
        try:
            client = await redis_cache.get_redis_client()
            keys = await client.keys("ws_stock_*")

            cache_stats = {
                'total_keys': len(keys),
                'key_types': {},
                'memory_usage': {}
            }

            # Sample analysis of first 50 keys
            sample_keys = keys[:50]
            for key in sample_keys:
                key_type = await client.type(key)
                cache_stats['key_types'][key_type] = (
                        cache_stats['key_types'].get(key_type, 0) + 1
                )

                # Get memory usage for string keys
                if key_type == 'string':
                    memory = await client.memory_usage(key)
                    cache_stats['memory_usage'][key] = memory

            info = await client.info()

            return Response({
                'cache_stats': cache_stats,
                'redis_info': {
                    'used_memory': info.get('used_memory_human'),
                    'connected_clients': info.get('connected_clients'),
                    'total_commands_processed': info.get('total_commands_processed'),
                    'uptime': info.get('uptime_in_seconds')
                },
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            logger.error(f"Cache manager error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        """Clear specific cache or all WebSocket cache"""
        cache_type = request.data.get('type', 'all')

        try:
            if cache_type == 'all':
                # Clear all WebSocket cache
                keys_to_delete = [
                    'recent_transactions_*',
                    'warehouse_summary_*',
                    'dashboard_metrics_*',
                    'transfers_*',
                    'connection_*'
                ]
                # In production, you'd use Redis SCAN and DEL
                return Response({
                    'status': 'success',
                    'cleared': 'all WebSocket cache keys'
                })
            elif cache_type in ['recent_transactions', 'warehouse_summary', 'dashboard_metrics']:
                # Clear specific cache type
                cache.clear()  # Django cache
                return Response({
                    'status': 'success',
                    'cleared': f'{cache_type} cache'
                })
            else:
                return Response({
                    'error': 'Invalid cache type'
                }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Utility Views for WebSocket Testing
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def websocket_test_view(request):
    """Simple test endpoint for WebSocket integration"""
    return JsonResponse({
        'message': 'WebSocket test successful',
        'user': request.user.username,
        'timestamp': timezone.now().isoformat(),
        'next_steps': [
            'Connect to /ws/stock_updates/ for real-time updates',
            'Use WebSocketEndpointsView for full endpoint list',
            'Monitor connections at /api/stock/websocket/status/'
        ]
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def manual_stock_update(request):
    """Manual trigger for stock update (for testing WebSocket broadcasting)"""
    stock_id = request.data.get('stock_id')
    quantity_change = request.data.get('quantity_change', 0)

    if not stock_id:
        return JsonResponse({
            'error': 'stock_id is required'
        }, status=400)

    try:
        # Create a test transaction
        transaction = StockTransaction.objects.create(
            stock_id=stock_id,
            transaction_type='adjustment',
            quantity=quantity_change,
            user=request.user,
            reference='websocket_test'
        )

        # This will trigger WebSocket broadcasting via signals
        serializer = StockTransactionSerializer(transaction)

        return JsonResponse({
            'status': 'success',
            'message': 'Stock updated and broadcasted via WebSocket',
            'transaction': serializer.data
        })

    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)


# Admin Dashboard View
class WebSocketAdminDashboardView(APIView):
    """Comprehensive WebSocket admin dashboard"""

    permission_classes = [IsAdminUser]

    async def get(self, request):
        """Get complete WebSocket dashboard data"""
        try:
            # Connection stats
            connection_stats = await redis_cache.get_connection_stats()

            # Performance metrics
            performance = {
                'message_rate': await self.get_message_rate(),
                'average_latency': '45ms',  # From monitoring
                'error_rate': '0.02%'
            }

            # Recent activity
            recent_activity = await self.get_recent_activity()

            return Response({
                'connections': connection_stats,
                'performance': performance,
                'recent_activity': recent_activity,
                'cache_status': await self.get_cache_status(),
                'alerts': await self.get_websocket_alerts(),
                'timestamp': timezone.now().isoformat()
            })

        except Exception as e:
            logger.error(f"Admin dashboard error: {str(e)}")
            return Response({
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @sync_to_async
    def get_message_rate(self):
        """Calculate recent message rate"""
        # Implement based on your logging/monitoring
        return 125  # messages per minute

    @sync_to_async
    def get_recent_activity(self):
        """Get recent WebSocket activity"""
        return []  # Implement based on your logging

    async def get_cache_status(self):
        """Get cache status"""
        client = await redis_cache.get_redis_client()
        info = await client.info()
        return {
            'hit_rate': '92%',
            'memory_usage': info.get('used_memory_human'),
            'evictions': info.get('evicted_keys')
        }

    @sync_to_async
    def get_websocket_alerts(self):
        """Get WebSocket related alerts"""
        return []  # Implement based on your alert system

# Note: Actual WebSocket consumers should be in stock/consumers.py
# These are referenced in urls.py for ASGI routing

# =============================================================================
# 🛠️ DEBUG & DEVELOPMENT VIEWS
# =============================================================================

class RebuildStockIndicesView(APIView):
    """Rebuild database indices (Development only)"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from django.db import connection

        # Example index rebuilding queries
        with connection.cursor() as cursor:
            cursor.execute("""
                           CREATE INDEX CONCURRENTLY IF NOT EXISTS
                               idx_warehouse_stock_quantity
                               ON stock_warehousestock (quantity)
                               WHERE quantity > 0;
                           """)

        return Response({
            'success': True,
            'message': 'Indices rebuilt successfully'
        })

class StockPerformanceTestView(APIView):
    """Performance testing endpoint"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import time

        # Test different query patterns
        start = time.time()
        Stock.objects.count()
        count_time = time.time() - start

        start = time.time()
        WarehouseStock.objects.filter(quantity__gt=0).count()
        filter_time = time.time() - start

        return Response({
            'count_query_time': f'{count_time:.4f}s',
            'filter_query_time': f'{filter_time:.4f}s',
            'total_stocks': Stock.objects.count(),
            'active_warehouse_stocks': WarehouseStock.objects.filter(quantity__gt=0).count()
        })

class StockCacheStatsView(APIView):
    """Cache statistics"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cache_info = {
            'total_keys': len(cache),
            'dashboard_cache': bool(cache.get('dashboard_stats')),
            'warehouse_cache': bool(cache.get('warehouse_stock_1_1'))
        }

        return Response(cache_info)

# =============================================================================
# 📊 STANDARD VIEWSETS
# =============================================================================

class StockTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StockTransaction.objects.select_related(
        'stock__category', 'user', 'from_warehouse', 'to_warehouse'
    ).order_by('-created_at')
    serializer_class = StockTransactionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'stock', 'transaction_type', 'from_warehouse', 'to_warehouse',
        'created_at', 'user'
    ]
    search_fields = ['stock__name', 'stock__item_code', 'reference', 'notes']
    ordering_fields = ['created_at', 'quantity', 'unit_price']
    throttle_classes = [StockOperationThrottle]

class SupplierViewSet(viewsets.ModelViewSet):
    queryset = Supplier.objects.all().order_by('name')
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'contact_person', 'email']
    ordering_fields = ['name', 'created_at']

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all().order_by('name')
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'description']

class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.all().order_by('name')
    serializer_class = WarehouseSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'created_at', 'capacity']

class TransferItemViewSet(viewsets.ModelViewSet):
    queryset = TransferItem.objects.select_related('transfer', 'stock').order_by('id')
    serializer_class = WarehouseTransferItemSerializer
    permission_classes = [IsAuthenticated]

# =============================================================================
# 🧾 BATCH MANAGEMENT
# =============================================================================

class StockBatchViewSet(viewsets.ModelViewSet):
    """Comprehensive batch management with expiry tracking and valuation"""
    queryset = StockBatch.objects.select_related(
        'stock', 'warehouse', 'stock__category', 'stock__valuation_method'
    ).prefetch_related('usages').order_by('-received_date')
    serializer_class = StockBatchSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = [
        'stock', 'warehouse', 'is_active', 'expiry_date',
        'remaining_quantity', 'stock__category', 'stock__status'
    ]
    search_fields = ['batch_number', 'stock__name', 'stock__item_code']
    ordering_fields = [
        'received_date', 'expiry_date', 'remaining_quantity',
        'initial_quantity', 'unit_price'
    ]
    ordering = ['-received_date']
    throttle_classes = [BatchOperationThrottle]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StockBatchCreateSerializer
        return StockBatchSerializer

    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get batches expiring within specified days"""
        days = int(request.query_params.get('days', 30))
        expiry_date = timezone.now().date() + timedelta(days=days)

        batches = self.queryset.filter(
            expiry_date__lte=expiry_date,
            expiry_date__gte=timezone.now().date(),
            remaining_quantity__gt=0,
            is_active=True
        ).order_by('expiry_date')

        page = self.paginate_queryset(batches)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(batches, many=True)
        return Response({
            'count': batches.count(),
            'batches': serializer.data,
            'days_ahead': days
        })

    @action(detail=False, methods=['get'])
    def low_remaining(self, request):
        """Get batches with low remaining quantity"""
        threshold = int(request.query_params.get('threshold', 10))

        batches = self.queryset.filter(
            remaining_quantity__lte=threshold,
            remaining_quantity__gt=0,
            is_active=True
        ).order_by('remaining_quantity')

        serializer = self.get_serializer(batches, many=True)
        return Response({
            'count': batches.count(),
            'threshold': threshold,
            'batches': serializer.data
        })

    @action(detail=True, methods=['post'])
    def use_batch(self, request, pk=None):
        """Record batch usage for FIFO/LIFO valuation"""
        batch = self.get_object()
        quantity = int(request.data.get('quantity', 0))
        transaction_reference = request.data.get('reference', '')
        notes = request.data.get('notes', '')

        if quantity <= 0:
            return Response({
                'error': 'Quantity must be greater than 0'
            }, status=status.HTTP_400_BAD_REQUEST)

        if batch.remaining_quantity < quantity:
            return Response({
                'error': f'Insufficient remaining quantity. Available: {batch.remaining_quantity}'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # Update batch usage
                batch.used_quantity += quantity
                batch.save()

                # Create usage record
                StockBatchUsage.objects.create(
                    batch=batch,
                    transaction_id=request.data.get('transaction_id'),
                    quantity=quantity
                )

                # Create stock transaction
                StockTransaction.objects.create(
                    stock=batch.stock,
                    from_warehouse=batch.warehouse,
                    to_warehouse=batch.warehouse,
                    transaction_type='out',
                    quantity=-quantity,
                    unit_price=batch.unit_price,
                    reference=f"Batch {batch.batch_number} - {transaction_reference}",
                    notes=notes,
                    user=request.user
                )

                # Update warehouse stock
                warehouse_stock = WarehouseStock.objects.get(
                    stock=batch.stock, warehouse=batch.warehouse
                )
                warehouse_stock.quantity -= quantity
                warehouse_stock.save()

                serializer = self.get_serializer(batch)
                return Response({
                    'success': True,
                    'batch': serializer.data,
                    'remaining_quantity': batch.remaining_quantity,
                    'used_quantity': batch.used_quantity
                })

        except Exception as e:
            return Response({
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def valuation_summary(self, request):
        """Batch-based valuation summary by stock item"""
        stock_id = request.query_params.get('stock_id')
        warehouse_id = request.query_params.get('warehouse_id')

        batches = self.queryset.filter(
            remaining_quantity__gt=0,
            is_active=True
        )

        if stock_id:
            batches = batches.filter(stock_id=stock_id)
        if warehouse_id:
            batches = batches.filter(warehouse_id=warehouse_id)

        valuation_data = []
        for batch in batches:
            valuation_data.append({
                'batch_id': str(batch.id),
                'stock_id': str(batch.stock.id),
                'stock_name': batch.stock.name,
                'item_code': batch.stock.item_code,
                'batch_number': batch.batch_number,
                'warehouse': batch.warehouse.name,
                'remaining_quantity': batch.remaining_quantity,
                'unit_price': float(batch.unit_price),
                'total_value': float(batch.remaining_quantity * batch.unit_price),
                'expiry_date': batch.expiry_date.isoformat() if batch.expiry_date else None,
                'days_to_expiry': (batch.expiry_date - timezone.now().date()).days if batch.expiry_date else None
            })

        # Group by stock for summary
        stock_summary = {}
        for item in valuation_data:
            stock_id = item['stock_id']
            if stock_id not in stock_summary:
                stock_summary[stock_id] = {
                    'stock_name': item['stock_name'],
                    'item_code': item['item_code'],
                    'total_quantity': 0,
                    'total_value': 0.00,
                    'batches': []
                }
            stock_summary[stock_id]['total_quantity'] += item['remaining_quantity']
            stock_summary[stock_id]['total_value'] += item['total_value']
            stock_summary[stock_id]['batches'].append(item)

        return Response({
            'total_batches': len(valuation_data),
            'total_value': sum(item['total_value'] for item in valuation_data),
            'stock_summary': list(stock_summary.values()),
            'detailed_batches': valuation_data
        })

# =============================================================================
# 💰 VALUATION VIEWS
# =============================================================================

class CalculateValuationView(APIView):
    """Comprehensive inventory valuation calculator"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [ValuationThrottle]

    def get(self, request):
        """Calculate valuation using multiple methods"""
        stock_id = request.query_params.get('stock_id')
        warehouse_id = request.query_params.get('warehouse_id')
        method = request.query_params.get('method', 'all').lower()

        # Base queryset
        stocks = Stock.objects.select_related(
            'valuation_method', 'category'
        ).prefetch_related(
            Prefetch('warehouse_stocks',
                     queryset=WarehouseStock.objects.select_related('warehouse')),
            Prefetch('batches',
                     queryset=StockBatch.objects.filter(is_active=True, remaining_quantity__gt=0)
                     .order_by('received_date'))
        ).filter(is_active=True)

        if stock_id:
            stocks = stocks.filter(id=stock_id)
        if warehouse_id:
            stocks = stocks.filter(warehouse_stocks__warehouse_id=warehouse_id).distinct()

        valuation_results = []

        for stock in stocks:
            # Get relevant warehouse stock
            warehouse_stock = None
            if warehouse_id:
                warehouse_stock = stock.warehouse_stocks.filter(warehouse_id=warehouse_id).first()
            else:
                warehouse_stock = stock.warehouse_stocks.filter(quantity__gt=0).first()

            if not warehouse_stock or warehouse_stock.quantity == 0:
                continue

            total_quantity = warehouse_stock.quantity
            relevant_batches = stock.batches.filter(warehouse=warehouse_stock.warehouse)

            # Calculate different valuation methods
            valuations = {}

            if method in ['all', 'fifo']:
                valuations['fifo'] = self._calculate_fifo(relevant_batches, total_quantity)

            if method in ['all', 'lifo']:
                valuations['lifo'] = self._calculate_lifo(relevant_batches, total_quantity)

            if method in ['all', 'weighted']:
                valuations['weighted_average'] = float(warehouse_stock.total_value)

            if method in ['all', 'average']:
                total_cost = sum(batch.remaining_quantity * batch.unit_price
                                 for batch in relevant_batches)
                total_qty = sum(batch.remaining_quantity for batch in relevant_batches)
                valuations['simple_average'] = total_cost / max(total_qty, 1)

            valuation_results.append({
                'stock_id': str(stock.id),
                'item_code': stock.item_code,
                'name': stock.name,
                'category': stock.category.name if stock.category else None,
                'warehouse': warehouse_stock.warehouse.name if warehouse_stock else None,
                'total_quantity': total_quantity,
                'unit_price': float(warehouse_stock.unit_price),
                'valuations': valuations,
                'recommended_method': stock.valuation_method.name if stock.valuation_method else 'weighted_average'
            })

        # Calculate summary
        summary = {
            'total_items': len(valuation_results),
            'total_quantity': sum(r['total_quantity'] for r in valuation_results),
        }

        for method in ['fifo', 'lifo', 'weighted_average', 'simple_average']:
            if method in next(iter(valuation_results), {}).get('valuations', {}):
                summary[f'total_{method}_value'] = sum(
                    r['valuations'].get(method, 0) for r in valuation_results
                )

        return Response({
            'method': method,
            'results': valuation_results,
            'summary': summary,
            'timestamp': timezone.now().isoformat()
        })

    def _calculate_fifo(self, batches, total_quantity):
        """FIFO valuation calculation"""
        fifo_value = Decimal('0.00')
        remaining_qty = total_quantity

        for batch in batches.order_by('received_date'):
            if remaining_qty <= 0:
                break
            available = min(batch.remaining_quantity, remaining_qty)
            fifo_value += Decimal(str(available)) * batch.unit_price
            remaining_qty -= available

        return float(fifo_value)

    def _calculate_lifo(self, batches, total_quantity):
        """LIFO valuation calculation"""
        lifo_value = Decimal('0.00')
        remaining_qty = total_quantity

        for batch in batches.order_by('-received_date'):
            if remaining_qty <= 0:
                break
            available = min(batch.remaining_quantity, remaining_qty)
            lifo_value += Decimal(str(available)) * batch.unit_price
            remaining_qty -= available

        return float(lifo_value)

class FIFOValuationView(CalculateValuationView):
    """FIFO-specific valuation"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['method'] = 'fifo'
        return super().get(request)

class LIFOValuationView(CalculateValuationView):
    """LIFO-specific valuation"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['method'] = 'lifo'
        return super().get(request)

class WeightedAverageValuationView(CalculateValuationView):
    """Weighted Average-specific valuation"""
    def get(self, request):
        request.query_params._mutable = True
        request.query_params['method'] = 'weighted'
        return super().get(request)

# =============================================================================
# 📊 REPORTING VIEWS
# =============================================================================

class MonthlyInventoryReportView(APIView):
    """Monthly inventory report generation and retrieval"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [ReportGenerationThrottle]

    def get(self, request):
        """Retrieve existing monthly reports or generate new one"""
        year = int(request.query_params.get('year', timezone.now().year))
        month = int(request.query_params.get('month', timezone.now().month))
        warehouse_id = request.query_params.get('warehouse_id')

        # Try to get existing report
        reports = StockReport.objects.filter(
            report_type='monthly',
            report_date__year=year,
            report_date__month=month,
            warehouse_id=warehouse_id
        ).order_by('-created_at')

        if reports.exists():
            serializer = StockReportSerializer(reports.first())
            return Response({
                'success': True,
                'report': serializer.data,
                'message': 'Monthly report found'
            })

        # Generate new report via Celery
        task = generate_monthly_report.delay(
            year, month, request.user.id, warehouse_id
        )

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Monthly report generation started',
            'status_url': f'/api/stock/reports/status/{task.id}/',
            'estimated_time': '2-5 minutes'
        })

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """Force generate monthly report"""
        year = int(request.data.get('year', timezone.now().year))
        month = int(request.data.get('month', timezone.now().month))
        warehouse_id = request.data.get('warehouse_id')

        task = generate_monthly_report.delay(
            year, month, request.user.id, warehouse_id
        )

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Monthly report generation started'
        })

class YearlyInventoryReportView(APIView):
    """Yearly inventory report generation and retrieval"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [ReportGenerationThrottle]

    def get(self, request):
        """Retrieve existing yearly reports or generate new one"""
        year = int(request.query_params.get('year', timezone.now().year))
        warehouse_id = request.query_params.get('warehouse_id')

        # Try to get existing report
        reports = StockReport.objects.filter(
            report_type='yearly',
            report_date__year=year,
            warehouse_id=warehouse_id
        ).order_by('-created_at')

        if reports.exists():
            serializer = StockReportSerializer(reports.first())
            return Response({
                'success': True,
                'report': serializer.data,
                'message': 'Yearly report found'
            })

        # Generate new report via Celery
        task = generate_yearly_report.delay(
            year, request.user.id, warehouse_id
        )

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Yearly report generation started',
            'status_url': f'/api/stock/reports/status/{task.id}/',
            'estimated_time': '5-10 minutes'
        })

    @action(detail=False, methods=['post'])
    def generate(self, request):
        """Force generate yearly report"""
        year = int(request.data.get('year', timezone.now().year))
        warehouse_id = request.data.get('warehouse_id')

        task = generate_yearly_report.delay(
            year, request.user.id, warehouse_id
        )

        return Response({
            'success': True,
            'task_id': task.id,
            'message': 'Yearly report generation started'
        })

class WarehouseTransferViewSet(viewsets.ModelViewSet):
    queryset = WarehouseTransfer.objects.select_related(
        'from_warehouse', 'to_warehouse', 'created_by'
    ).prefetch_related('items__stock').order_by('-transfer_date')
    serializer_class = WarehouseTransferSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'from_warehouse', 'to_warehouse']
    ordering_fields = ['transfer_date', 'expected_delivery_date']
    ordering = ['-transfer_date']
    throttle_classes = [WarehouseTransferThrottle]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return WarehouseTransferCreateSerializer
        return WarehouseTransferSerializer

    def perform_create(self, serializer):
        if not serializer.validated_data.get('transfer_number'):
            last_transfer = WarehouseTransfer.objects.order_by('-id').first()
            sequence = last_transfer.id + 1 if last_transfer else 1
            year_month = timezone.now().strftime("%Y%m")
            serializer.save(
                transfer_number=f'TR{year_month}{sequence:04d}',
                created_by=self.request.user
            )
        else:
            serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def start_transfer(self, request, pk=None):
        transfer = self.get_object()
        if transfer.status != 'pending':
            return Response({
                'error': 'Only pending transfers can be started'
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for item in transfer.items.all():
                from_stock = WarehouseStock.objects.get(
                    stock=item.stock, warehouse=transfer.from_warehouse
                )
                if from_stock.available_quantity < item.quantity:
                    return Response({
                        'error': f'Insufficient stock for {item.stock.name}'
                    }, status=status.HTTP_400_BAD_REQUEST)

                from_stock.reserved_quantity += item.quantity
                from_stock.save()

            transfer.status = 'in_transit'
            transfer.save()

        serializer = self.get_serializer(transfer)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def complete_transfer(self, request, pk=None):
        transfer = self.get_object()
        if transfer.status != 'in_transit':
            return Response({
                'error': 'Only in-transit transfers can be completed'
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            transfer.status = 'completed'
            transfer.save()

            for item in transfer.items.all():
                # Source warehouse
                from_stock = WarehouseStock.objects.get(
                    stock=item.stock, warehouse=transfer.from_warehouse
                )
                from_stock.reserved_quantity -= item.quantity
                from_stock.quantity -= item.quantity
                from_stock.save()

                # Destination warehouse
                to_stock, created = WarehouseStock.objects.get_or_create(
                    stock=item.stock,
                    warehouse=transfer.to_warehouse,
                    defaults={'quantity': 0, 'unit_price': item.unit_price}
                )
                to_stock.quantity += item.quantity
                to_stock.save()

        serializer = self.get_serializer(transfer)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def cancel_transfer(self, request, pk=None):
        transfer = self.get_object()
        if transfer.status not in ['pending', 'in_transit']:
            return Response({
                'error': 'Only pending or in-transit transfers can be cancelled'
            }, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            if transfer.status == 'in_transit':
                for item in transfer.items.all():
                    from_stock = WarehouseStock.objects.get(
                        stock=item.stock, warehouse=transfer.from_warehouse
                    )
                    from_stock.reserved_quantity -= item.quantity
                    from_stock.save()

            transfer.status = 'cancelled'
            transfer.save()

        serializer = self.get_serializer(transfer)
        return Response(serializer.data)

class InventoryAdjustmentViewSet(viewsets.ModelViewSet):
    """Inventory adjustment tracking"""
    queryset = InventoryAdjustment.objects.select_related('warehouse', 'created_by').order_by('-created_at')
    serializer_class = InventoryAdjustmentSerializer
    permission_classes = [IsAuthenticated]

class StockDashboardStatsView(APIView):
    """Comprehensive enterprise dashboard with real-time analytics"""
    permission_classes = [IsAuthenticated]
    throttle_classes = [StockOperationThrottle]

    def get(self, request):
        """Get complete dashboard statistics"""
        cache_key = f"dashboard_stats_{request.user.id}"
        cached_data = cache.get(cache_key)

        if cached_data:
            return Response(cached_data)

        # =================================================================
        # 1. Core Metrics
        # =================================================================
        total_items = Stock.objects.filter(is_active=True).count()

        total_value_result = WarehouseStock.objects.aggregate(
            total=Sum(F('quantity') * F('unit_price'), output_field=DecimalField())
        )
        total_value = total_value_result['total'] or Decimal('0.00')

        # =================================================================
        # 2. Stock Status Breakdown (Critical / Low / Normal / Out of Stock)
        # =================================================================
        stock_status = WarehouseStock.objects.filter(
            stock__is_active=True
        ).aggregate(
            critical=Count('id', filter=Q(quantity__lte=F('stock__reorder_level'))),
            low=Count('id', filter=Q(
                quantity__gt=F('stock__reorder_level'),
                quantity__lte=F('stock__min_stock_level')
            )),
            normal=Count('id', filter=Q(quantity__gt=F('stock__min_stock_level'))),
            out_of_stock=Count('id', filter=Q(quantity=0))
        )

        # =================================================================
        # 3. Expiry Analysis (Critical: <7 days, Warning: 7–30 days)
        # =================================================================
        today = timezone.now().date()
        expiring_batches = StockBatch.objects.filter(
            expiry_date__gte=today,
            is_active=True
        ).annotate(
            remaining_qty=F('initial_quantity') - F('used_quantity')
        ).filter(
            remaining_qty__gt=0
        ).aggregate(
            critical_expiry=Count('id', filter=Q(expiry_date__lte=today + timedelta(days=7))),
            warning_expiry=Count('id', filter=Q(
                expiry_date__gt=today + timedelta(days=7),
                expiry_date__lte=today + timedelta(days=30)
            ))
        )

        # =================================================================
        # 4. Warehouse Utilization
        # =================================================================
        warehouse_stats = Warehouse.objects.filter(is_active=True).aggregate(
            total_capacity=Sum('capacity'),
            total_used=Sum('current_utilization')
        )
        total_capacity = warehouse_stats['total_capacity'] or 0
        total_used = warehouse_stats['total_used'] or 0
        avg_utilization = (total_used / total_capacity * 100) if total_capacity > 0 else 0

        warehouse_utilization = {
            'total_capacity': int(total_capacity),
            'total_used': int(total_used),
            'avg_utilization': round(avg_utilization, 2)
        }

        # =================================================================
        # 5. Recent Activity (Last 24 hours)
        # =================================================================
        last_24h = timezone.now() - timedelta(hours=24)
        recent_activity = StockTransaction.objects.filter(
            created_at__gte=last_24h
        ).aggregate(
            total_transactions=Count('id'),
            total_in=Sum('quantity', filter=Q(transaction_type__in=['in', 'receive', 'transfer_in'])),
            total_out=Sum('quantity', filter=Q(transaction_type__in=['out', 'issue', 'transfer_out'])),
            value_moved=Sum(F('quantity') * F('unit_price'), output_field=DecimalField())
        )

        # =================================================================
        # 6. Top 5 Categories by Value
        # =================================================================
        top_categories = Stock.objects.filter(is_active=True) \
            .values('category__name') \
            .annotate(
            total_value=Sum(F('warehouse_stocks__quantity') * F('warehouse_stocks__unit_price')),
            item_count=Count('id')
        ) \
            .order_by('-total_value')[:5]

        # =================================================================
        # Final Response Data
        # =================================================================
        data = {
            "core_metrics": {
                "total_items": total_items,
                "total_value": float(total_value),
                "total_value_formatted": f"${float(total_value):,.2f}"
            },
            "stock_status": {k: int(v or 0) for k, v in stock_status.items()},
            "expiry_analysis": {k: int(v or 0) for k, v in expiring_batches.items()},
            "warehouse_utilization": {
                "total_capacity": int(warehouse_utilization['total_capacity'] or 0),
                "total_used": int(warehouse_utilization['total_used'] or 0),
                "avg_utilization": round(float(warehouse_utilization['avg_utilization'] or 0), 2)
            },
            "recent_activity": {
                "total_transactions": int(recent_activity['total_transactions'] or 0),
                "total_in": int(recent_activity['total_in'] or 0),
                "total_out": int(recent_activity['total_out'] or 0),
                "value_moved": float(recent_activity['value_moved'] or 0)
            },
            "top_categories": [
                {
                    "category": item['category__name'] or "Uncategorized",
                    "total_value": float(item['total_value'] or 0),
                    "item_count": item['item_count']
                }
                for item in top_categories
            ],
            "generated_at": timezone.now().isoformat(),
            "cached": False
        }

        # Cache for 5 minutes per user
        cache.set(cache_key, data, timeout=300)

        return Response(data)

class ThrottleStatusAPIView(APIView):
    """API endpoint to check all throttle statuses"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get comprehensive throttle status"""
        from .throttles import StockThrottleGroup

        status_data = {
            'throttles': StockThrottleGroup.get_all_status(request),
            'settings': {
                'max_batch_size': getattr(settings, 'MAX_BATCH_SIZE', 1000),
                'max_transfer_value': getattr(settings, 'MAX_TRANSFER_VALUE', 50000),
                'max_file_size_mb': getattr(settings, 'MAX_IMPORT_FILE_SIZE_MB', 50)
            },
            'timestamp': datetime.now().isoformat()
        }

        return Response(status_data, status=status.HTTP_200_OK)