# backend/tenants/views.py
from reportlab.lib.units import inch
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from .models import Company, Branch
from .serializers import (
    CompanySerializer, BranchSerializer,
    CompanyMinimalSerializer, BranchMinimalSerializer
)
from django.contrib.auth import get_user_model
from django.http import HttpResponse
import csv
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO

User = get_user_model()

# Optional: Custom pagination class
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class CompanyViewSet(viewsets.ModelViewSet):
    """
    API endpoint for viewing, creating, updating, and deleting companies (tenants)
    """
    queryset = Company.objects.all().order_by('name')
    serializer_class = CompanySerializer
    permission_classes = [permissions.IsAdminUser]  # Only admins manage tenants
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active', 'currency', 'country', 'city']
    search_fields = ['name', 'slug', 'email', 'phone']
    pagination_class = StandardResultsSetPagination

    def get_serializer_class(self):
        """
        Use minimal serializer for list/selects
        """
        if self.action in ['list', 'retrieve']:
            if self.request.query_params.get('minimal') == 'true':
                return CompanyMinimalSerializer
        return super().get_serializer_class()

    def perform_create(self, serializer):
        """
        Auto-assign current user as creator
        """
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['get'], url_path='branches')
    def branches(self, request, pk=None):
        """Get all active branches for this company"""
        company = self.get_object()
        branches = company.branches.filter(is_active=True).order_by('name')
        serializer = BranchSerializer(branches, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='analytics')
    def analytics(self, request):
        """
        Advanced analytics for companies/tenants
        Returns data ready for charts (total, growth, users, etc.)
        """
        now = timezone.now()
        month_ago = now - timedelta(days=30)

        # Basic counts
        total_companies = Company.objects.count()
        active_companies = Company.objects.filter(is_active=True).count()
        inactive_companies = total_companies - active_companies

        # New companies this month
        new_this_month = Company.objects.filter(created_at__gte=month_ago).count()

        # Growth rate (simple % change vs last month)
        last_month_start = month_ago - timedelta(days=30)
        last_month_new = Company.objects.filter(
            created_at__gte=last_month_start,
            created_at__lt=month_ago
        ).count()
        growth_rate = ((new_this_month - last_month_new) / last_month_new * 100) if last_month_new else 0

        # User stats per company (aggregated)
        user_stats = Company.objects.annotate(
            user_count=Count('users'),
            new_users_this_month=Count('users', filter=Q(users__date_joined__gte=month_ago))
        ).values('name', 'user_count', 'new_users_this_month')

        # Data for frontend charts (e.g. Chart.js)
        chart_data = {
            'labels': [c['name'] for c in user_stats],
            'datasets': [
                {
                    'label': 'Total Users',
                    'data': [c['user_count'] for c in user_stats],
                    'backgroundColor': 'rgba(59, 130, 246, 0.6)',
                },
                {
                    'label': 'New Users (30 days)',
                    'data': [c['new_users_this_month'] for c in user_stats],
                    'backgroundColor': 'rgba(16, 185, 129, 0.6)',
                }
            ]
        }

        return Response({
            'total_companies': total_companies,
            'active_companies': active_companies,
            'inactive_companies': inactive_companies,
            'new_this_month': new_this_month,
            'growth_rate_percent': round(growth_rate, 2),
            'user_stats': list(user_stats),
            'chart_data': chart_data,
        })


class BranchViewSet(viewsets.ModelViewSet):
    """
    API endpoint for viewing, creating, updating, and deleting branches
    """
    queryset = Branch.objects.all().order_by('company__name', 'name')
    serializer_class = BranchSerializer
    permission_classes = [permissions.IsAdminUser]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['company', 'company_id', 'is_active', 'city', 'region']
    search_fields = ['name', 'slug', 'city', 'region', 'address', 'company__name']
    pagination_class = StandardResultsSetPagination

    def get_serializer_class(self):
        if self.action in ['list', 'retrieve']:
            if self.request.query_params.get('minimal') == 'true':
                return BranchMinimalSerializer
        return super().get_serializer_class()

    def get_queryset(self):
        queryset = super().get_queryset()
        company_id = self.request.query_params.get('company_id')
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        return queryset.select_related('company', 'manager')

    def perform_create(self, serializer):
        # Optional: add creator if needed
        serializer.save()

    @action(detail=False, methods=['get'], url_path='analytics')
    def analytics(self, request):
        """
        Branch-level analytics (total, active, by company, etc.)
        """
        total_branches = Branch.objects.count()
        active_branches = Branch.objects.filter(is_active=True).count()
        branches_by_company = Branch.objects.values('company__name').annotate(count=Count('id'))

        return Response({
            'total_branches': total_branches,
            'active_branches': active_branches,
            'inactive_branches': total_branches - active_branches,
            'branches_by_company': list(branches_by_company),
        })

class CompanyCSVExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Company.objects.all()
        # Apply filters (copy from your existing filter logic)
        # ... add your filter code here ...

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="companies_export_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Name', 'Slug', 'Email', 'Phone', 'Currency', 'Timezone',
            'Active', 'Created At', 'Branch Count', 'User Count'
        ])

        for company in queryset:
            writer.writerow([
                company.id,
                company.name,
                company.slug,
                company.email or '—',
                company.phone or '—',
                company.currency,
                company.timezone,
                'Yes' if company.is_active else 'No',
                company.created_at.strftime('%Y-%m-%d %H:%M'),
                company.branches.count(),
                company.users.count()  # Adjust related_name if needed
            ])

        return response


class BranchCSVExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Branch.objects.all()
        # Apply filters (copy your filter code)
        # ... add your filter code here ...

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="branches_export_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Company', 'Name', 'City', 'Region', 'Phone', 'Email',
            'Manager', 'Active', 'Created At', 'User Count'
        ])

        for branch in queryset:
            manager_name = branch.manager.get_full_name() if branch.manager else '—'
            writer.writerow([
                branch.id,
                branch.company.name,
                branch.name,
                branch.city or '—',
                branch.region or '—',
                branch.phone or '—',
                branch.email or '—',
                manager_name,
                'Yes' if branch.is_active else 'No',
                branch.created_at.strftime('%Y-%m-%d %H:%M'),
                branch.users.count()  # Adjust related_name
            ])

        return response

class CompanyExcelExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Company.objects.all()
        # ... filters ...

        wb = Workbook()
        ws = wb.active
        ws.title = "Companies Report"

        headers = [
            'ID', 'Name', 'Slug', 'Email', 'Phone', 'Currency', 'Timezone',
            'Active', 'Created At', 'Branches', 'Users'
        ]
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for company in queryset:
            ws.append([
                company.id,
                company.name,
                company.slug,
                company.email or '—',
                company.phone or '—',
                company.currency,
                company.timezone,
                'Yes' if company.is_active else 'No',
                company.created_at.strftime('%Y-%m-%d %H:%M'),
                company.branches.count(),
                company.users.count()
            ])

        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="companies_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response


class BranchExcelExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Branch.objects.all()
        # ... filters ...

        wb = Workbook()
        ws = wb.active
        ws.title = "Branches Report"

        headers = [
            'ID', 'Company', 'Name', 'City', 'Region', 'Phone', 'Email',
            'Manager', 'Active', 'Created At', 'Users'
        ]
        ws.append(headers)

        # Header style
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        for branch in queryset:
            manager_name = branch.manager.get_full_name() if branch.manager else '—'
            ws.append([
                branch.id,
                branch.company.name,
                branch.name,
                branch.city or '—',
                branch.region or '—',
                branch.phone or '—',
                branch.email or '—',
                manager_name,
                'Yes' if branch.is_active else 'No',
                branch.created_at.strftime('%Y-%m-%d %H:%M'),
                branch.users.count()
            ])

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="branches_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        wb.save(response)
        return response

class CompanyPDFExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Company.objects.all()
        # ... filters ...

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph("Companies Report - High Prosper", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 24))

        # Summary
        summary_data = [
            ["Total Companies", queryset.count()],
            ["Active Companies", queryset.filter(is_active=True).count()],
            ["Inactive Companies", queryset.filter(is_active=False).count()],
        ]
        summary_table = Table(summary_data, colWidths=[4*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 24))

        # Table
        data = [['Name', 'Slug', 'Email', 'Phone', 'Currency', 'Active', 'Branches', 'Users']]
        for company in queryset:
            data.append([
                company.name,
                company.slug,
                company.email or '—',
                company.phone or '—',
                company.currency,
                'Yes' if company.is_active else 'No',
                company.branches.count(),
                company.users.count()
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="companies_export_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        response.write(buffer.getvalue())
        return response


class BranchPDFExportView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def get(self, request):
        queryset = Branch.objects.all()
        # ... filters ...

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []

        styles = getSampleStyleSheet()
        title = Paragraph("Branches Report - High Prosper", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 24))

        summary_data = [
            ["Total Branches", queryset.count()],
            ["Active Branches", queryset.filter(is_active=True).count()],
            ["Inactive Branches", queryset.filter(is_active=False).count()],
        ]
        summary_table = Table(summary_data, colWidths=[4*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.black),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 24))

        data = [['Company', 'Branch', 'City', 'Region', 'Phone', 'Manager', 'Active', 'Created']]
        for branch in queryset:
            manager_name = branch.manager.get_full_name() if branch.manager else '—'
            data.append([
                branch.company.name,
                branch.name,
                branch.city or '—',
                branch.region or '—',
                branch.phone or '—',
                manager_name,
                'Yes' if branch.is_active else 'No',
                branch.created_at.strftime('%Y-%m-%d')
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="branches_export_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
        response.write(buffer.getvalue())
        return response