# backend/fleet/admin.py
# FULL ADMIN PANEL – 2025 HIGH PROSPER FLEET SYSTEM

from django.contrib import admin
from django.urls import reverse
from django.utils.html import format_html
from django.utils import timezone
from django.http import HttpResponse, response
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
import qrcode
from io import BytesIO
from django.core.mail import EmailMessage
from django.conf import settings
from datetime import timedelta

from .models import (
    Vehicle, Driver, Branch, Customer, WasteCollection,
    FuelLog, Route, VehiclePhoto, Repair, Consumption,
    FuelEntry, OilEntry, FuelRecord, MaintenanceRecord,
    FuelEfficiencyRecord, Compliance, WorkshopRecord,
    DriverPerformanceHistory
)


# ====================== VEHICLE ADMIN – FINAL 2025 EDITION ======================
@admin.register(Vehicle)
class VehicleAdmin(admin.ModelAdmin):
    list_display = ['registration_number', 'brand', 'model', 'status_colored', 'odometer_reading', 'photo_tag']
    list_filter = ['status', 'vehicle_type', 'manufacture_year', 'brand']
    search_fields = ['registration_number', 'chassis_number', 'engine_number', 'brand', 'model']
    readonly_fields = ['photo_preview', 'created_at', 'updated_at']
    fieldsets = (
        ("Basic Information", {
            "fields": ("registration_number", "vehicle_type", "brand", "model", "manufacture_year", "registration_date")
        }),
        ("Technical Details", {
            "fields": ("chassis_number", "engine_number", "bdm_kg", "odometer_reading")
        }),
        ("Financial", {
            "fields": ("purchase_date", "purchase_price", "current_value")
        }),
        ("Status & GPS", {
            "fields": ("status", "lat", "lng", "last_location_update")
        }),
        ("Media & Notes", {
            "fields": ("photo", "photo_preview", "notes")
        }),
        ("AI Damage Inspection", {
            "fields": ("damage_photo", "damage_detected", "damage_confidence", "damage_report"),
            "classes": ("collapse",)
        }),
        ("Timestamps", {
            "fields": ("created_at", "updated_at"),
            "classes": ("collapse",)
        }),
    )

    # Photo preview in list and change view
    def photo_tag(self, obj):
        if obj.photo:
            return format_html(
                '<img src="{}" width="100" height="70" style="border-radius:8px;object-fit:cover;box-shadow:0 4px 10px rgba(0,0,0,0.1);" />',
                obj.photo.url
            )
        return "No Photo"
    photo_tag.short_description = "Photo"

    def photo_preview(self, obj):
        if obj.photo:
            return format_html('<img src="{}" width="700" style="border-radius:16px;" />', obj.photo.url)
        return "No Image Uploaded"
    photo_preview.short_description = "Full Preview"

    # Color-coded status badge
    def status_colored(self, obj):
        colors = {
            'active': '#10b981',
            'on_road': '#3b82f6',
            'workshop': '#f97316',
            'standby': '#a855f7',
            'maintenance': '#eab308',
            'retired': '#ef4444'
        }
        color = colors.get(obj.status, '#6b7280')
        return format_html(
            '<span style="color:white;background:{};padding:8px 16px;border-radius:999px;font-weight:bold;font-size:13px;">{}</span>',
            color, obj.get_status_display().upper()
        )
    status_colored.short_description = "Status"

    # ====================== ACTIONS ======================

    actions = [
        'export_csv',
        'export_excel',
        'generate_vehicle_pdf',
        'generate_pdf_with_qr',
        'email_pdf_report',
        'run_ai_damage_check'
    ]

    # 1. Export CSV
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="fleet_{timezone.now().strftime("%Y%m%d")}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Reg No', 'Brand', 'Model', 'Type', 'Status', 'Odometer', 'BDM kg', 'Year', 'Value'])
        for v in queryset:
            writer.writerow([
                v.registration_number, v.brand, v.model,
                v.get_vehicle_type_display(), v.get_status_display(),
                v.odometer_reading, v.bdm_kg, v.manufacture_year,
                v.current_value or v.purchase_price or 'N/A'
            ])
        return response
    export_csv.short_description = "Export to CSV"

    # 2. Export Excel (XLSX) - Beautiful & Colored
    def export_excel(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="fleet_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Fleet Report"

        # Styles
        header_font = Font(name='Calibri', bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        columns = ['Reg No', 'Brand', 'Model', 'Type', 'Status', 'Odometer (km)', 'BDM (kg)', 'Year', 'Value (RM)']
        for i, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=i, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        status_colors = {
            'Active': '10b981', 'On Road': '3b82f6', 'Workshop': 'f97316',
            'Standby': 'a855f7', 'Maintenance': 'eab308', 'Retired': 'ef4444'
        }

        for r_idx, obj in enumerate(queryset, 2):
            row = [
                obj.registration_number, obj.brand, obj.model,
                obj.get_vehicle_type_display(), obj.get_status_display(),
                obj.odometer_reading, obj.bdm_kg, obj.manufacture_year,
                obj.current_value or obj.purchase_price or 0
            ]
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border
                if c_idx == 5:  # Status column
                    color_hex = status_colors.get(value, "6b7280")
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

        # Auto column width
        for col in ws.columns:
            max_len = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_len + 2, 50)

        wb.save(response)
        return response
    export_excel.short_description = "Export to Excel (Styled)"

    # 3. PDF with QR Code
    def generate_pdf_with_qr(self, request, queryset):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.8*inch, bottomMargin=0.8*inch)
        elements = []
        styles = getSampleStyleSheet()

        for vehicle in queryset:
            # QR Code
            qr_buffer = BytesIO()
            qr = qrcode.QRCode(version=1, box_size=6, border=4)
            qr.add_data(f"https://yourdomain.com/fleet/vehicle/{vehicle.id}/")  # Change URL
            qr.make(fit=True)
            img = qr.make_image(fill_color="#1e40af", back_color="white")
            img.save(qr_buffer, "PNG")
            qr_buffer.seek(0)

            elements.append(Paragraph(f"<font size=20><b>{vehicle.registration_number}</b></font>", styles['Title']))
            elements.append(Paragraph(f"{vehicle.brand} {vehicle.model} • {vehicle.get_vehicle_type_display()}", styles['Heading2']))
            elements.append(Spacer(1, 20))
            elements.append(Image(qr_buffer, width=130, height=130))
            elements.append(Paragraph("<i>Scan for live vehicle details</i>", styles['Italic']))
            elements.append(Spacer(1, 20))

            data = [
                ["Field", "Details"],
                ["Status", vehicle.get_status_display()],
                ["Odometer", f"{vehicle.odometer_reading:,.0f} km"],
                ["BDM", f"{vehicle.bdm_kg} kg"],
                ["Manufacture Year", str(vehicle.manufacture_year)],
                ["Current Value", f"RM {vehicle.current_value or 'N/A'}"],
            ]
            table = Table(data, colWidths=[2.5*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('GRID', (0,0), (-1,-1), 1, colors.lightgrey),
                ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
            ]))
            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fleet_qr_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
        return response
    generate_pdf_with_qr.short_description = "Generate PDF with QR Code"

    # 4. Email PDF Report
    def email_pdf_report(self, request, queryset):
        if not request.user.email:
            self.message_user(request, "Add your email in profile to receive reports.", level='error')
            return

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        for v in queryset:
            elements.append(Paragraph(f"<b>{v.registration_number}</b> - {v.brand} {v.model}", styles['Heading1']))
            elements.append(Paragraph(f"Status: {v.get_status_display()} | Odometer: {v.odometer_reading:,} km", styles['Normal']))
            elements.append(Spacer(1, 20))

        doc.build(elements)
        buffer.seek(0)

        email = EmailMessage(
            subject=f"Fleet Report - {queryset.count()} Vehicles",
            body="Your fleet report is attached.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[request.user.email],
        )
        email.attach("fleet_report.pdf", buffer.getvalue(), 'application/pdf')
        email.send()

        self.message_user(request, f"Report sent to {request.user.email}!", level='success')
    email_pdf_report.short_description = "Email PDF Report"

    # 5. AI Damage Detection (YOLOv8)
    def run_ai_damage_check(self, request, queryset):
        from .utils import detect_vehicle_damage
        for vehicle in queryset:
            if not vehicle.damage_photo:
                self.message_user(request, f"{vehicle} → No damage photo", level='warning')
                continue
            try:
                result = detect_vehicle_damage(vehicle.damage_photo.path)
                vehicle.damage_detected = result["damage_detected"]
                vehicle.damage_confidence = result["confidence"]
                vehicle.damage_report = " | ".join(result["labels"])
                vehicle.save()
                status = "DETECTED" if result["damage_detected"] else "CLEAR"
                self.message_user(request, f"{vehicle.registration_number}: {status} ({result['confidence']})", level='success')
            except Exception as e:
                self.message_user(request, f"{vehicle}: AI Error - {e}", level='error')
    run_ai_damage_check.short_description = "Run AI Damage Detection (YOLOv8)"



# ====================== DRIVER ADMIN ======================
@admin.register(Driver)
class DriverAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'phone', 'license_number', 'license_expiry', 'assigned_vehicle_link', 'rating']
    list_filter = ['license_expiry', 'assigned_vehicle__status']
    search_fields = ['user__first_name', 'user__last_name', 'user__username', 'license_number']
    readonly_fields = ['profile_picture_preview']

    def full_name(self, obj):
        return obj.full_name
    full_name.short_description = "Driver Name"

    def phone(self, obj):
        return obj.phone

    def assigned_vehicle_link(self, obj):
        if obj.assigned_vehicle:
            url = reverse("admin:fleet_vehicle_change", args=[obj.assigned_vehicle.id])
            return format_html('<a href="{}">{}</a>', url, obj.assigned_vehicle.registration_number)
        return "—"
    assigned_vehicle_link.short_description = "Vehicle"

    def profile_picture_preview(self, obj):
        if obj.profile_picture:
            return format_html('<img src="{}" width="150" style="border-radius:50%;" />', obj.profile_picture)
        return "No Photo"
    profile_picture_preview.short_description = "Photo"


# ====================== FUEL EFFICIENCY (NEW) ======================
@admin.register(FuelEfficiencyRecord)
class FuelEfficiencyRecordAdmin(admin.ModelAdmin):
    list_display = ['vehicle_link', 'date', 'distance_km', 'liters', 'km_per_liter', 'cost', 'cost_per_km']
    list_filter = ['date', 'vehicle__brand']
    search_fields = ['vehicle__registration_number']
    date_hierarchy = 'date'

    def vehicle_link(self, obj):
        url = reverse("admin:fleet_vehicle_change", args=[obj.vehicle.id])
        return format_html('<a href="{}"><b>{}</b></a>', url, obj.vehicle.registration_number)
    vehicle_link.short_description = "Vehicle"


# ====================== COMPLIANCE (NEW) ======================
@admin.register(Compliance)
class ComplianceAdmin(admin.ModelAdmin):
    list_display = ['vehicle_link', 'compliance_type', 'expiry_date', 'days_left_badge', 'status_colored']
    list_filter = ['compliance_type', 'expiry_date']
    search_fields = ['vehicle__registration_number']
    readonly_fields = ['days_left']

    def vehicle_link(self, obj):
        url = reverse("admin:fleet_vehicle_change", args=[obj.vehicle.id])
        return format_html('<a href="{}"><b>{}</b></a>', url, obj.vehicle.registration_number)
    vehicle_link.short_description = "Vehicle"

    def days_left_badge(self, obj):
        if obj.expiry_date is None:
            return format_html('<span style="color:#888">Not set</span>')
        days = obj.days_left()
        if days < 0:
            return format_html('<span style="color:#dc2626;font-weight:bold;">EXPIRED</span>')
        elif days <= 5:
            return format_html('<span style="color:#b91c1c;">{} days (CRITICAL)</span>', days)
        elif days <= 30:
            return format_html('<span style="color:#f97316;">{} days</span>', days)
        else:
            return format_html('<span style="color:#16a34a;">{} days</span>', days)
        days_left_badge.short_description = "Days Left"

    def status_colored(self, obj):
        if obj.expiry_date is None:
            return format_html('<span style="color:#888;padding:6px 12px;border-radius:999px;background:#333;">NOT SET</span>')

        status = obj.status()
        colors = {
            "expired": "#dc2626",
            "critical": "#b91c1c",
            "warning": "#f97316",
            "valid": "#16a34a",
            "not_set": "#6b7280"
        }
        return format_html(
            '<span style="color:white;background:{};padding:6px 12px;border-radius:999px;font-weight:bold;">{}</span>',
            colors.get(status, "#6b7280"), status.upper()
        )
    status_colored.short_description = "Status"


# ====================== WORKSHOP (NEW) ======================
@admin.register(WorkshopRecord)
class WorkshopRecordAdmin(admin.ModelAdmin):
    list_display = ['vehicle_link', 'date_in', 'status_colored', 'mechanic', 'cost']
    list_filter = ['status', 'date_in', 'mechanic']
    search_fields = ['vehicle__registration_number', 'issue_description']

    def vehicle_link(self, obj):
        url = reverse("admin:fleet_vehicle_change", args=[obj.vehicle.id])
        return format_html('<a href="{}"><b>{}</b></a>', url, obj.vehicle.registration_number)
    vehicle_link.short_description = "Vehicle"

    def status_colored(self, obj):
        colors = {'pending': '#6b7280', 'in_progress': '#f59e0b', 'completed': '#10b981'}
        color = colors.get(obj.status, '#6b7280')
        return format_html(
            '<span style="color:white;background:{};padding:6px 12px;border-radius:9999px;">{}</span>',
            color, obj.get_status_display().upper()
        )
    status_colored.short_description = "Status"


# ====================== OTHER MODELS (Quick & Clean) ======================

@admin.register(Branch)
class BranchAdmin(admin.ModelAdmin):
    list_display = ['name', 'city', 'district']
    search_fields = ['name', 'city']

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ['full_name', 'customer_type', 'phone', 'branch']
    list_filter = ['customer_type', 'branch']
    search_fields = ['full_name', 'phone']

@admin.register(WasteCollection)
class WasteCollectionAdmin(admin.ModelAdmin):
    list_display = ['customer', 'vehicle', 'driver', 'date', 'waste_weight', 'collection_status']
    list_filter = ['collection_status', 'date']
    date_hierarchy = 'date'

@admin.register(FuelLog)
class FuelLogAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'date', 'fuel_amount', 'cost', 'odometer_reading']
    list_filter = ['date']
    search_fields = ['vehicle__registration_number']

@admin.register(Route)
class RouteAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'driver', 'date', 'start_location', 'end_location']
    list_filter = ['date']

@admin.register(VehiclePhoto)
class VehiclePhotoAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'uploaded_at', 'image_tag']
    readonly_fields = ['image_preview']

    def image_tag(self, obj):
        return format_html('<img src="{}" width="120" />', obj.image.url)
    image_tag.short_description = "Thumbnail"

    def image_preview(self, obj):
        return format_html('<img src="{}" width="600" />', obj.image.url)

@admin.register(Repair)
class RepairAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'description', 'cost', 'status', 'date']
    list_filter = ['status', 'date']

@admin.register(Consumption)
class ConsumptionAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'month', 'fuel', 'oil']

@admin.register(FuelEntry)
class FuelEntryAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'date', 'liters', 'cost']
    list_filter = ['date']

@admin.register(OilEntry)
class OilEntryAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'date', 'liters', 'cost']

@admin.register(FuelRecord)
class FuelRecordAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'date', 'liters', 'cost']
    list_filter = ['date']

@admin.register(MaintenanceRecord)
class MaintenanceRecordAdmin(admin.ModelAdmin):
    list_display = ['vehicle', 'maintenance_type', 'date', 'next_due', 'cost']
    list_filter = ['maintenance_type', 'date']

@admin.register(DriverPerformanceHistory)
class DriverPerformanceAdmin(admin.ModelAdmin):
    list_display = ['driver', 'month', 'score', 'collections', 'fuel', 'rating']
    list_filter = ['month', 'branch']
    search_fields = ['driver__user__first_name', 'driver__user__last_name']