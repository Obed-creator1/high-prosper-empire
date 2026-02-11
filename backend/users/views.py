import json
import secrets
import traceback
from decimal import Decimal

from asgiref.sync import async_to_sync
from django.contrib.auth.hashers import make_password
from django.contrib.contenttypes.models import ContentType
from django.contrib.messages import get_messages
from django.core.exceptions import PermissionDenied
from django.db.models.functions import TruncMonth, TruncDay, ExtractHour, ExtractIsoWeekDay
from django.db import models
from django.db.models import Q, Sum, Count, Value, CharField, Func, F, OuterRef, Exists
from django.contrib.auth import get_user_model
from django.shortcuts import get_object_or_404
from rest_framework import status, viewsets, views, permissions, generics
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from channels.layers import get_channel_layer
import io

from reportlab.lib.pagesizes import A4
from django.http import HttpResponse, JsonResponse
import csv
from django.utils import timezone
from django.core.mail import send_mail, EmailMessage
import random
from datetime import timedelta, datetime
from django.conf import settings
from payments.models import Payment, PaymentMethod
from pywebpush import webpush, WebPushException
from .models import Sticker, BlockedUser
from .permissions import IsOwnerOrAdmin
from django.core.files.storage import default_storage
from hr.services import MTNSMSService, EmailService
from .utils import require_group_admin
from customers.models import Customer
from django.core.cache import cache
from collector.models import Collector
from notifications.models import Notification
from fleet.models import Vehicle
from hr.models import Staff
from payments.models import Invoice
from procurement.models import Item, Supplier
from django.db.models.functions import Lower, Concat
from django_filters.rest_framework import DjangoFilterBackend
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.pagination import PageNumberPagination
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage, \
    PageBreak, PageTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
import os

from .models import (
    CustomUser, UserProfile, ChatMessage, Sticker, MessageReaction,
    ChatRoom, RoomMember, Post, Comment, Reaction, Share, Friendship, Activity, OTP, SearchAnalytics
)
from .serializers import (
    UserSerializer, LoginSerializer, UserUpdateSerializer,
    SidebarUserSerializer, ChatRoomSerializer, UserListSerializer,
    ChatMessageSerializer, StickerSerializer,
    PostSerializer, CommentSerializer, ReactionSerializer,
    ShareSerializer, FriendshipSerializer, ActivitySerializer, BlockedUserListSerializer,
    BlockedUserSerializer, BlockedUserCreateSerializer, ActivityMinimalSerializer
)

User = get_user_model()

# Temporary OTP store (in-memory). For production use Redis or database table.
OTP_STORE = {}

# ─── Configuration ──────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STATIC_DIR = os.path.join(BASE_DIR, 'static')
LOGO_PATH = os.path.join(STATIC_DIR, 'logo.png')  # Update to your real logo path

SOCIAL_ICONS = {
    "Twitter/X": {
        "icon": os.path.join(STATIC_DIR, 'social', 'twitter.png'),
        "url": "https://twitter.com/highprosper_rw"
    },
    "Facebook": {
        "icon": os.path.join(STATIC_DIR, 'social', 'facebook.png'),
        "url": "https://facebook.com/highprosper"
    },
    "LinkedIn": {
        "icon": os.path.join(STATIC_DIR, 'social', 'linkedin.png'),
        "url": "https://linkedin.com/company/highprosper"
    },
    "Instagram": {
        "icon": os.path.join(STATIC_DIR, 'social', 'instagram.png'),
        "url": "https://instagram.com/highprosper_rw"
    },
    "WhatsApp": {
        "icon": os.path.join(STATIC_DIR, 'social', 'whatsapp.png'),
        "url": "https://wa.me/250788123456"
    },
    "Telegram": {
        "icon": os.path.join(STATIC_DIR, 'social', 'telegram.png'),
        "url": "https://t.me/highprosper_rw"
    }
}

COMPANY_NAME = "High Prosper Services Ltd"
COMPANY_LOCATION = "Kigali, Rwanda"
COMPANY_PHONE = "+250 788 123 456"
COMPANY_EMAIL = "info@highprosper.com"

# Custom Pagination
class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class ActivityPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


# ──────────────────────────────────────────────────────────────
# PERMISSIONS
# ──────────────────────────────────────────────────────────────

class IsAdminOrCEO(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role in ['admin', 'ceo']

class IsOwnerOrReadOnly(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True
        return obj.user == request.user

class IsPostOwnerOrAdmin(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.method in permissions.SAFE_METHODS:
            return True
        return obj.user == request.user or request.user.role in ['admin', 'ceo']

class CheckUniqueAPIView(APIView):
    """
    POST /api/v1/users/check-unique/

    Body:
    {
        "field": "username" | "email" | "phone",
        "value": "the_value_to_check"
    }

    Returns:
    200 { "available": true/false }
    400 { "detail": "Invalid field" }
    """
    permission_classes = []  # public endpoint

    def post(self, request):
        field = request.data.get("field")
        value = request.data.get("value")

        valid_fields = ["username", "email", "phone"]
        if field not in valid_fields:
            return Response(
                {"detail": f"Invalid field. Must be one of: {', '.join(valid_fields)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not value:
            return Response(
                {"detail": "Value is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Build query (case-insensitive for username/email, exact for phone)
        query = Q()
        if field == "username":
            query = Q(username__iexact=value)
        elif field == "email":
            query = Q(email__iexact=value)
        elif field == "phone":
            query = Q(phone__exact=value)  # phone usually exact match

        exists = User.objects.filter(query).exists()

        return Response({"available": not exists})

# ---------------------------------------------------------------------------
# 👥 User Management (Admin)
# ---------------------------------------------------------------------------

class UserViewSet(viewsets.ModelViewSet):
    """
    Full User Management API
    - Admins/CEOs: full CRUD on all users
    - Authenticated users: read/update own profile only (via /me/)
    - POST /api/v1/users/ → create new user (admin only)
    """
    queryset = CustomUser.objects.all().order_by('first_name', 'last_name')
    serializer_class = UserSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'is_active', 'company', 'branch']
    search_fields = ['username', 'first_name', 'last_name', 'email', 'phone']
    ordering_fields = ['date_joined', 'last_login', 'role', 'first_name']

    def get_permissions(self):
        """
        Custom permissions per action:
        - list/retrieve: authenticated users (but filtered to own data if not admin)
        - create/update/delete: admin or CEO only
        - me: authenticated user only
        """
        if self.action in ['me', 'retrieve', 'update', 'partial_update']:
            return [IsAuthenticated()]
        if self.action in ['create', 'destroy', 'list', 'collectors']:
            return [IsAdminOrCEO()]
        return super().get_permissions()

    def get_serializer_class(self):
        if self.action in ['list', 'collectors', 'me']:
            return UserListSerializer
        return UserSerializer

    def get_queryset(self):
        """
        Filter queryset based on user role
        - Admins/CEOs see everything
        - Regular users only see their own profile
        """
        qs = super().get_queryset()
        user = self.request.user

        if user.is_superuser or user.is_ceo:  # assuming is_ceo is a property/method
            return qs

        if self.action in ['retrieve', 'update', 'partial_update', 'me']:
            return qs.filter(id=user.id)

        # For list/collectors: non-admins get empty queryset
        return qs.none()

    def perform_create(self, serializer):
        """
        Handle password hashing on create + set created_by
        """
        if 'password' in serializer.validated_data:
            serializer.validated_data['password'] = make_password(
                serializer.validated_data['password']
            )
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        """
        Handle password update only if provided
        """
        if 'password' in serializer.validated_data:
            serializer.validated_data['password'] = make_password(
                serializer.validated_data['password']
            )
        serializer.save()

    @action(detail=False, methods=['get'], url_path='me', permission_classes=[IsAuthenticated])
    def me(self, request):
        """Return current user's full profile"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    def list(self, request, *args, **kwargs):
        """Always return paginated format with count"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "count": queryset.count(),
            "results": serializer.data
        })

    @action(detail=False, methods=['get'], url_path='collectors', permission_classes=[IsAuthenticated])
    def collectors(self, request):
        """
        Dedicated endpoint: /api/v1/users/collectors/
        Returns only active collectors — perfect for village assignment dropdown
        """
        collectors_qs = CustomUser.objects.filter(
            role='collector',
            is_active=True
        ).order_by('first_name', 'last_name')

        page = self.paginate_queryset(collectors_qs)
        if page is not None:
            serializer = UserListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = UserListSerializer(collectors_qs, many=True)
        return Response({
            "count": collectors_qs.count(),
            "results": serializer.data
        })

class UpdateProfileView(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request):
        user = request.user
        serializer = UserUpdateSerializer(user, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Profile updated successfully",
                "user": UserSerializer(user, context={'request': request}).data
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        new_password = request.data.get("new_password")

        if not new_password:
            return Response({"error": "New password required"}, status=400)

        user.set_password(new_password)
        user.must_change_password = False
        user.save()
        return Response({"success": "Password updated successfully"})


# ──────────────────────────────────────────────────────────────
# SOCIAL FEATURES VIEWSETS
# ──────────────────────────────────────────────────────────────

class PostViewSet(viewsets.ModelViewSet):
    """
    Posts API - Create, list, detail, update, delete
    """
    queryset = Post.objects.all().order_by('-created_at')
    serializer_class = PostSerializer
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['privacy', 'is_announcement', 'user']
    search_fields = ['content']

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAuthenticated(), IsPostOwnerOrAdmin()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=True, methods=['post'], url_path='like')
    def like(self, request, pk=None):
        """
        Toggle like on a post.
        Expects: {"action": "like" | "unlike"}
        Returns updated likes_count and user_has_liked
        """
        post = self.get_object()

        # Get or create content type for Post
        content_type = ContentType.objects.get_for_model(Post)

        # Check if user already liked
        existing = Reaction.objects.filter(
            user=request.user,
            content_type=content_type,
            object_id=post.id,
            reaction_type='like'  # or your like emoji/reaction type
        ).first()

        if request.data.get('action') == 'unlike':
            if existing:
                existing.delete()
        else:  # like
            if not existing:
                Reaction.objects.create(
                    user=request.user,
                    content_type=content_type,
                    object_id=post.id,
                    reaction_type='like'
                )

        # Recalculate total likes
        likes_count = Reaction.objects.filter(
            content_type=content_type,
            object_id=post.id,
            reaction_type='like'
        ).count()

        user_has_liked = Reaction.objects.filter(
            user=request.user,
            content_type=content_type,
            object_id=post.id,
            reaction_type='like'
        ).exists()

        return Response({
            'likes_count': likes_count,
            'user_has_liked': user_has_liked
        })

    @action(detail=True, methods=['post'], url_path='comments')
    def comments(self, request, pk=None):
        """
        POST /api/v1/users/posts/<post_id>/comments/
        Create a new comment on this post.
        Expected payload: { "content": "Your comment here" }
        """
        post = self.get_object()

        # Pass post & request to serializer context
        serializer = CommentSerializer(
            data=request.data,
            context={'request': request, 'post': post}  # ← pass context!
        )

        serializer.is_valid(raise_exception=True)

        # Auto-assign user & post
        comment = serializer.save()  # create() will use context

        # Return full serialized comment (with user info)
        return Response(CommentSerializer(comment).data, status=status.HTTP_201_CREATED)


class CommentViewSet(viewsets.ModelViewSet):
    """
    Comment ViewSet - Handles CRUD for comments with proper nesting under posts.
    """
    queryset = Comment.objects.all().order_by('-created_at')  # Newest first
    serializer_class = CommentSerializer
    pagination_class = StandardResultsSetPagination

    # Permissions: authenticated for create, owner-only for update/delete
    def get_permissions(self):
        if self.action in ['create', 'list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsOwnerOrReadOnly()]

    # ─── Nested under Post ────────────────────────────────────────────────────
    # This allows /api/v1/posts/<post_pk>/comments/
    def get_queryset(self):
        """
        Filter comments by post if we're in a nested route
        """
        queryset = super().get_queryset()

        # If accessed via /posts/<id>/comments/
        post_pk = self.kwargs.get('post_pk')
        if post_pk is not None:
            queryset = queryset.filter(post_id=post_pk)

        return queryset

    def perform_create(self, serializer):
        """
        Auto-assign user and post when creating a comment
        """
        post_pk = self.kwargs.get('post_pk')
        if post_pk is None:
            raise PermissionDenied("Comments must be created under a specific post.")

        try:
            post = Post.objects.get(id=post_pk)
        except Post.DoesNotExist:
            raise PermissionDenied("Post does not exist.")

        # Save with user and post
        serializer.save(
            user=self.request.user,
            post=post
        )

    # Optional: Custom create response (more detailed)
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    # Extra action: Get comments count for a post (if needed)
    @action(detail=False, methods=['get'], url_path='count/(?P<post_id>[^/.]+)')
    def count(self, request, post_id=None):
        try:
            count = Comment.objects.filter(post_id=post_id).count()
            return Response({'comments_count': count})
        except:
            return Response({'error': 'Invalid post ID'}, status=400)


class ReactionViewSet(viewsets.ModelViewSet):
    queryset = Reaction.objects.all()
    serializer_class = ReactionSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ShareViewSet(viewsets.ModelViewSet):
    queryset = Share.objects.all().order_by('-shared_at')
    serializer_class = ShareSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(sharer=self.request.user)


class FriendshipViewSet(viewsets.ModelViewSet):
    """
    Friendship ViewSet - Handles friend requests & friendships
    """
    queryset = Friendship.objects.all()
    serializer_class = FriendshipSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Users can only see their own friendships/requests
        Admins/CEOs see everything
        """
        user = self.request.user
        if user.is_superuser or user.role in ['admin', 'ceo']:
            return Friendship.objects.all()

        return Friendship.objects.filter(
            Q(from_user=user) | Q(to_user=user)
        )

    def perform_create(self, serializer):
        """
        When creating → it's always a new friend request
        Auto-set from_user = current user
        """
        serializer.save(from_user=self.request.user)

    @action(detail=True, methods=['post'], url_path='accept')
    def accept(self, request, pk=None):
        """Accept a friend request (only recipient can do this)"""
        friendship = self.get_object()

        if friendship.to_user != request.user:
            return Response(
                {"detail": "You can only accept requests sent to you."},
                status=status.HTTP_403_FORBIDDEN
            )

        if friendship.status != 'pending':
            return Response(
                {"detail": "This request is not pending."},
                status=status.HTTP_400_BAD_REQUEST
            )

        friendship.status = 'accepted'
        friendship.save()

        return Response(FriendshipSerializer(friendship).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Reject a friend request (only recipient)"""
        friendship = self.get_object()

        if friendship.to_user != request.user:
            return Response(
                {"detail": "You can only reject requests sent to you."},
                status=status.HTTP_403_FORBIDDEN
            )

        if friendship.status != 'pending':
            return Response(
                {"detail": "This request is not pending."},
                status=status.HTTP_400_BAD_REQUEST
            )

        friendship.status = 'rejected'
        friendship.save()

        return Response(FriendshipSerializer(friendship).data)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        """Cancel a pending friend request (only sender)"""
        friendship = self.get_object()

        if friendship.from_user != request.user:
            return Response(
                {"detail": "You can only cancel requests you sent."},
                status=status.HTTP_403_FORBIDDEN
            )

        if friendship.status != 'pending':
            return Response(
                {"detail": "This request is not pending."},
                status=status.HTTP_400_BAD_REQUEST
            )

        friendship.status = 'cancelled'
        friendship.save()

        return Response(FriendshipSerializer(friendship).data)

    # List pending received requests
    @action(detail=False, methods=['get'], url_path='pending-received')
    def pending_received(self, request):
        friendships = Friendship.objects.filter(
            to_user=request.user,
            status='pending'
        )
        serializer = FriendshipSerializer(friendships, many=True)
        return Response(serializer.data)

    # List pending sent requests
    @action(detail=False, methods=['get'], url_path='pending-sent')
    def pending_sent(self, request):
        friendships = Friendship.objects.filter(
            from_user=request.user,
            status='pending'
        )
        serializer = FriendshipSerializer(friendships, many=True)
        return Response(serializer.data)

    # List accepted friends
    @action(detail=False, methods=['get'], url_path='friends')
    def friends(self, request):
        friendships = Friendship.objects.filter(
            Q(from_user=request.user) | Q(to_user=request.user),
            status='accepted'
        )
        friends = []
        for f in friendships:
            if f.from_user == request.user:
                friends.append(f.to_user)
            else:
                friends.append(f.from_user)

        serializer = UserListSerializer(friends, many=True)
        return Response(serializer.data)


class ActivityViewSet(viewsets.ModelViewSet):
    """
    Activity ViewSet - complete CRUD + custom actions

    Permissions:
    - Admins / staff → full access
    - Authenticated users → read-only access to their own activities
    - Anonymous → no access
    """
    queryset = Activity.objects.all()
    serializer_class = ActivitySerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = {
        'action_type': ['exact', 'in'],
        'user__id': ['exact'],
        'created_at': ['gte', 'lte', 'date__gte', 'date__lte'],
    }
    search_fields = ['action_type', 'user__username', 'extra_data']
    ordering_fields = ['created_at', 'action_type']
    ordering = ['-created_at']

    def get_queryset(self):
        """
        Admins see everything
        Regular users see only their own activities
        """
        qs = super().get_queryset().select_related('user')

        if not self.request.user.is_staff and not self.request.user.is_superuser:
            qs = qs.filter(user=self.request.user)

        return qs

    def get_serializer_class(self):
        """
        Use minimal serializer for list views to reduce payload size
        """
        if self.action in ['list', 'my_activities', 'recent']:
            return ActivityMinimalSerializer
        return ActivitySerializer

    # ─── Custom Actions ─────────────────────────────────────────────────────

    @action(detail=False, methods=['get'], url_path='my')
    def my_activities(self, request):
        """Shortcut for current user's activities"""
        queryset = self.filter_queryset(
            self.get_queryset().filter(user=request.user)
        )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='recent')
    def recent(self, request):
        """Last 50 activities (for dashboard / notification center)"""
        queryset = self.filter_queryset(
            self.get_queryset()
        )[:50]
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='today')
    def today(self, request):
        """Activities from today only"""
        today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        queryset = self.filter_queryset(
            self.get_queryset().filter(created_at__gte=today_start)
        )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    # ─── Create / Update / Delete restrictions ─────────────────────────────

    def create(self, request, *args, **kwargs):
        """Activities should usually be created via signals, not direct POST"""
        return Response(
            {"detail": "Activities are created automatically via system events."},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def update(self, request, *args, **kwargs):
        """Activities are read-only after creation"""
        return Response(
            {"detail": "Activities cannot be updated."},
            status=status.HTTP_405_METHOD_NOT_ALLOWED
        )

    def destroy(self, request, *args, **kwargs):
        """Only admins can delete activities (cleanup)"""
        if not request.user.is_staff:
            return Response(
                {"detail": "Only staff can delete activities."},
                status=status.HTTP_403_FORBIDDEN
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['patch'], url_path='read')
    def mark_read(self, request, pk=None):
        activity = self.get_object()
        if activity.user != request.user:
            return Response(status=403)
        # Assuming you add is_read field to Activity model
        activity.is_read = True
        activity.save(update_fields=['is_read'])
        return Response({'status': 'read'})

    @action(detail=False, methods=['post'], url_path='mark-all-read')
    def mark_all_read(self, request):
        Activity.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'status': 'all marked as read'})

class BlockListCreateAPIView(generics.ListCreateAPIView):
    """
    GET: List all blocks created by the current user
    POST: Create a new block
    """
    serializer_class = BlockedUserCreateSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'GET':
            return BlockedUserListSerializer
        return BlockedUserCreateSerializer

    def get_queryset(self):
        return BlockedUser.objects.filter(blocker=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(blocker=self.request.user)


class BlockDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    """
    GET: Retrieve a specific block
    PATCH: Update reason or expires_at (owner or admin)
    DELETE: Unblock / delete (owner or admin)
    """
    serializer_class = BlockedUserSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrAdmin]
    lookup_field = 'id'
    lookup_url_kwarg = 'block_id'

    def get_queryset(self):
        return BlockedUser.objects.all()

    def perform_destroy(self, instance):
        # Soft unblock instead of hard delete
        instance.is_active = False
        instance.save(update_fields=['is_active'])
        # Optional: broadcast via signal


class GlobalSearchView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        query = request.GET.get('q', '').strip()
        type_filter = request.GET.get('type')
        status_filter = request.GET.get('status')
        date_from = request.GET.get('from')
        date_to = request.GET.get('to')
        page = max(1, int(request.GET.get('page', 1)))
        limit = min(50, int(request.GET.get('limit', 15)))
        suggest = request.GET.get('suggest') == 'true'

        filters = {
            "type": type_filter,
            "status": status_filter,
            "from": date_from,
            "to": date_to,
        }
        filters_clean = {k: v for k, v in filters.items() if v}

        # Log search analytics (except suggestions)
        analytics_entry = None
        if not suggest and (query or filters_clean):
            analytics_entry = SearchAnalytics.objects.create(
                user=user,
                query=query,
                filters=filters_clean,
                results_count=0,
                has_results=False,
            )

        if len(query) < 2 and not suggest:
            return Response({"results": []})

        is_admin = user.role in ['ceo', 'admin', 'manager']
        is_collector = user.role == 'collector'

        # Parse dates
        date_from_obj = None
        date_to_obj = None
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                pass

        if suggest:
            # Real-time suggestions — unique titles only
            titles = set()

            # Customers
            customer_qs = Customer.objects.all()
            if is_collector:
                customer_qs = customer_qs.filter(collector=user)
            titles.update(customer_qs.filter(name__icontains=query).values_list('name', flat=True)[:8])

            # Invoices
            invoice_qs = Invoice.objects.all()
            if is_collector:
                invoice_qs = invoice_qs.filter(customer__collector=user)
            titles.update(
                invoice_qs.annotate(
                    title=Concat(
                        Value('Invoice '),
                        'period_year',
                        Value('-'),
                        Func(F('period_month'), function='LPAD', arg2=Value(2), arg3=Value('0'))
                    )
                ).values_list('title', flat=True)[:8]
            )

            # Vehicles (admin only)
            if is_admin:
                titles.update(
                    Vehicle.objects.filter(
                        Q(registration_number__icontains=query) |
                        Q(brand__icontains=query) |
                        Q(model__icontains=query)
                    )
                    .annotate(full_title=Concat('registration_number', Value(' — '), 'brand', Value(' '), 'model'))
                    .values_list('full_title', flat=True)[:8]
                )

            # Items
            titles.update(Item.objects.filter(Q(name__icontains=query) | Q(sku__icontains=query)).values_list('name', flat=True)[:8])

            # Staff (admin only)
            if is_admin:
                titles.update(
                    Staff.objects.select_related('user')
                    .filter(
                        Q(user__username__icontains=query) |
                        Q(user__first_name__icontains=query) |
                        Q(user__last_name__icontains=query)
                    )
                    .annotate(full_name=Concat('user__first_name', Value(' '), 'user__last_name'))
                    .values_list('full_name', flat=True)[:8]
                )

            # Suppliers (admin only)
            if is_admin:
                titles.update(Supplier.objects.filter(name__icontains=query).values_list('name', flat=True)[:8])

            suggestions = list(titles)[:8]
            return Response({"suggestions": suggestions})

        # Full search mode
        all_results = []
        total_count = 0
        offset = (page - 1) * limit

        # Helper to add results
        def add_results(qs, result_type, title_expr, subtitle_template, href_template):
            nonlocal total_count
            count = qs.count()
            total_count += count
            for item in qs:
                title = title_expr(item)
                relevance = 0 if query and query.lower() in title.lower() else 1
                all_results.append({
                    "id": item.id,
                    "title": title,
                    "subtitle": subtitle_template(item),
                    "type": result_type,
                    "href": href_template(item),
                    "relevance": relevance
                })

        # 1. Users (Admins only)
        if is_admin and (not type_filter or type_filter == 'user'):
            qs = CustomUser.objects.filter(
                Q(username__icontains=query) |
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(phone__icontains=query) |
                Q(email__icontains=query)
            )
            add_results(
                qs, "user",
                lambda u: u.get_full_name() or u.username,
                lambda u: f"{u.get_role_display()} • {u.phone or 'No phone'}",
                lambda u: f"/users/{u.id}"
            )

        # 2. Customers
        if not type_filter or type_filter == 'customer':
            qs = Customer.objects.all()
            if is_collector:
                qs = qs.filter(collector=user)
            if query:
                qs = qs.filter(
                    Q(name__icontains=query) |
                    Q(phone__icontains=query) |
                    Q(village__name__icontains=query)
                )

            add_results(
                qs, "customer",
                lambda c: c.name,
                lambda c: f"Customer • {c.phone} • Balance: RWF {c.balance:,}",
                lambda c: f"/customers/{c.id}"
            )

        # 3. Invoices
        if not type_filter or type_filter == 'invoice':
           qs = Invoice.objects.all()
           if is_collector:
               qs = qs.filter(customer__collector=user)
           if query:
               # Search in customer name or period (as string)
               qs = qs.filter(
                   Q(customer__name__icontains=query) |  # assuming customer has 'name'
                   Q(period_year__icontains=query) |
                   Q(period_month__icontains=query)
               )
           if status_filter:
               qs = qs.filter(status__iexact=status_filter)
           if date_from_obj:
               qs = qs.filter(due_date__gte=date_from_obj)
           if date_to_obj:
               qs = qs.filter(due_date__lte=date_to_obj)

               add_results(
                   qs, "invoice",
                   lambda i: f"Invoice {i.period_year}-{i.period_month:02d}",  # e.g., "Invoice 2025-12"
                   lambda i: f"Invoice • RWF {i.amount:,} • {i.status}",
                   lambda i: f"/payments/invoices/{i.id}"
               )

        # 4. Vehicles (Admins only)
        if is_admin and (not type_filter or type_filter == 'vehicle'):
            qs = Vehicle.objects.all()
            if query:
                qs = qs.filter(
                    Q(registration_number__icontains=query) |
                    Q(brand__icontains=query) |
                    Q(model__icontains=query)
                )
            if status_filter:
                qs = qs.filter(status__iexact=status_filter)
            add_results(
                qs, "vehicle",
                lambda v: f"{v.registration_number} — {v.brand} {v.model}",
                lambda v: f"Vehicle • {v.vehicle_type} • {v.status}",
                lambda v: f"/fleet/vehicles/{v.id}"
            )

        # 5. Items
        if not type_filter or type_filter == 'item':
            qs = Item.objects.all()
            if query:
                qs = qs.filter(
                    Q(name__icontains=query) |
                    Q(sku__icontains=query)
                )
            add_results(
                qs, "item",
                lambda i: i.name,
                lambda i: f"{i.get_item_type_display()} • {i.sku}",
                lambda i: f"/procurement/items/{i.id}"
            )

        # 6. Staff (Admins only)
        if is_admin and (not type_filter or type_filter == 'staff'):
            qs = Staff.objects.select_related('user')
            if query:
                qs = qs.filter(
                    Q(user__username__icontains=query) |
                    Q(user__first_name__icontains=query) |
                    Q(user__last_name__icontains=query) |
                    Q(position__icontains=query)
                )
            add_results(
                qs, "staff",
                lambda s: s.user.get_full_name() or s.user.username,
                lambda s: f"Staff • {s.position} • {s.department}",
                lambda s: f"/hr/staff/{s.id}"
            )

        # 7. Suppliers (Admins only)
        if is_admin and (not type_filter or type_filter == 'supplier'):
            qs = Supplier.objects.all()
            if query:
                qs = qs.filter(
                    Q(name__icontains=query) |
                    Q(code__icontains=query) |
                    Q(phone__icontains=query)
                )
            add_results(
                qs, "supplier",
                lambda s: s.name,
                lambda s: f"Supplier • {s.code}",
                lambda s: f"/procurement/suppliers/{s.id}"
            )

        # Sort by relevance (exact title matches first)
        all_results.sort(key=lambda x: x['relevance'])

        # Pagination
        paginated = all_results[offset:offset + limit]
        total_pages = (total_count + limit - 1) // limit if total_count > 0 else 1

        # Update analytics
        if analytics_entry:
            analytics_entry.results_count = len(paginated)
            analytics_entry.has_results = len(paginated) > 0
            analytics_entry.save(update_fields=['results_count', 'has_results'])

        return Response({
            "results": paginated,
            "count": len(paginated),
            "total_count": total_count,
            "page": page,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_previous": page > 1
        })

class SearchClickView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        data = request.data

        query = data.get('search_term', '')
        result = data.get('result', {})

        # Find the most recent search by this user with matching query
        recent_search = SearchAnalytics.objects.filter(
            user=user,
            query=query
        ).order_by('-timestamp').first()

        if recent_search:
            recent_search.clicked_result = {
                "title": result.get('title'),
                "type": result.get('type'),
                "href": result.get('href'),
                "position": result.get('position')
            }
            recent_search.save(update_fields=['clicked_result'])

        return Response({"status": "logged"})

def send_notification_to_user(user_id: int, title: str, message: str, notification_type="info", action_url=None, image=None):
    channel_layer = get_channel_layer()
    if channel_layer:
        async_to_sync(channel_layer.group_send)(
            f"user_notify_{user_id}",
            {
                "type": "send_notification",
                "content": {
                    "title": title,
                    "message": message,
                    "notification_type": notification_type,
                    "action_url": action_url,
                    "image": image,
                }
            }
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_notifications(request):
    notifications = Notification.objects.filter(user=request.user)[:50]
    data = [{
        "id": n.id,
        "title": n.title,
        "message": n.message,
        "is_read": n.is_read,
        "created_at": n.created_at_time.isoformat(),
        "notification_type": n.notification_type,
        "action_url": n.action_url,
        "image": n.image,
    } for n in notifications]
    return Response({"results": data})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_as_read(request, pk):
    notification = get_object_or_404(Notification, id=pk, user=request.user)
    notification.is_read = True
    notification.save()
    return Response({"status": "ok"})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def mark_all_as_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return Response({"status": "ok"})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def save_push_subscription(request):
    subscription = request.data.get("subscription")
    if not subscription:
        return Response({"error": "No subscription"}, status=400)

    # Save to user profile (create field in model or use separate model)
    request.user.push_subscription = json.dumps(subscription)
    request.user.save()

    return Response({"success": True})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_push_notification(request):
    title = request.data.get("title", "New Message")
    body = request.data.get("body", "You have a new message")
    icon = request.data.get("icon", "/icon-192x192.png")
    url = request.data.get("url", "/chat")

    # Send to all users with subscription
    users = CustomUser.objects.exclude(push_subscription__isnull=True).exclude(push_subscription="")

    for user in users:
        try:
            subscription = json.loads(user.push_subscription)
            webpush(
                subscription_info=subscription,
                data=json.dumps({
                    "title": title,
                    "body": body,
                    "icon": icon,
                    "url": url
                }),
                vapid_private_key="YOUR_PRIVATE_KEY",
                vapid_claims={"sub": "mailto:admin@yourapp.com"}
            )
        except Exception as e:
            print(f"Push failed for {user.username}: {e}")

    return Response({"sent": users.count()})

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me(request):
    """
    Return current logged-in user info.
    """
    serializer = UserSerializer(request.user)
    return Response(serializer.data)

class AdminResetPasswordView(APIView):
    permission_classes = [IsAdminUser]

    def post(self, request, user_id):
        try:
            user = CustomUser.objects.get(id=user_id, role='customer')
        except CustomUser.DoesNotExist:
            return Response({"error": "Customer not found"}, status=404)

        # Reset password
        user.password = make_password("default123")
        user.must_change_password = True
        user.save()

        # Send notifications
        if user.phone:
            MTNSMSService(
                phone_number=user.phone,
                message=f"Your password has been reset. Username: {user.username}, Password: default123. Please change it after login."
            )
        if user.email:
            EmailService(
                to_email=user.email,
                subject="Password Reset Notification",
                message=f"Hello {user.username},\nYour password has been reset to default123. Please change it after login."
            )

        return Response({"success": f"Password reset for {user.username} and notifications sent"})


class UploadAttachment(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=400)

        path = default_storage.save(f"chat_attachments/{file.name}", file)
        url = request.build_absolute_uri(default_storage.url(path))

        return Response({"url": url})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_stickers(request):
    stickers = Sticker.objects.all().order_by("-uploaded_at")
    serializer = StickerSerializer(stickers, many=True)
    return Response(serializer.data)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_sticker(request):
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided"}, status=400)
    sticker = Sticker.objects.create(name=file.name, file=file)
    return Response({"id": sticker.id, "url": sticker.url})


@api_view(["POST"])
@permission_classes([AllowAny])
def password_reset_request(request):
    """Handles sending an OTP to the user's email for password reset."""
    email = request.data.get("email")

    if not email:
        return Response({"error": "Email is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return Response({"message": "If your email exists, an OTP has been sent."}, status=status.HTTP_200_OK)

    # Generate OTP
    otp = str(random.randint(100000, 999999))
    OTP_STORE[email] = {
        "otp": otp,
        "expires_at": timezone.now() + timedelta(minutes=10)
    }

    # Send email
    send_mail(
        subject="Password Reset OTP - High Prosper Services",
        message=f"Hello {user.username},\n\nYour OTP for password reset is: {otp}\nIt will expire in 10 minutes.\n\nThank you.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[email],
        fail_silently=True,
    )

    return Response({"message": "OTP has been sent to your email."}, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([AllowAny])
def password_reset_confirm(request):
    """Verifies OTP and sets a new password."""
    email = request.data.get("email")
    otp = request.data.get("otp")
    new_password = request.data.get("new_password")

    if not all([email, otp, new_password]):
        return Response({"error": "Email, OTP, and new password are required."},
                        status=status.HTTP_400_BAD_REQUEST)

    otp_data = OTP_STORE.get(email)

    if not otp_data:
        return Response({"error": "No OTP found. Request a new one."}, status=status.HTTP_400_BAD_REQUEST)

    if timezone.now() > otp_data["expires_at"]:
        OTP_STORE.pop(email, None)
        return Response({"error": "OTP expired. Request a new one."}, status=status.HTTP_400_BAD_REQUEST)

    if otp_data["otp"] != otp:
        return Response({"error": "Invalid OTP."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.get(email=email)
        user.set_password(new_password)
        user.save()

        OTP_STORE.pop(email, None)
        return Response({"message": "Password reset successful."}, status=status.HTTP_200_OK)
    except User.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_400_BAD_REQUEST)


# STEP 1: Send OTP
class LoginSendOTPView(APIView):
    def post(self, request):
        username = request.data.get("username")

        if not username:
            return Response({"error": "Username is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        # Remove previous OTPs
        OTP.objects.filter(user=user, is_used=False).delete()

        # Generate new OTP
        code = str(random.randint(100000, 999999))
        expires_at = timezone.now() + timedelta(minutes=5)
        otp = OTP.objects.create(user=user, code=code, expires_at=expires_at)

        print(f"OTP for {user.username}: {otp.code}")  # Debug only; replace with email/SMS send

        return Response({
            "message": "OTP sent successfully.",
            "session_id": str(otp.session_id),
            "expires_in": 300,  # 5 minutes
        }, status=status.HTTP_200_OK)


# STEP 2: Verify OTP
class LoginVerifyOTPView(APIView):
    def post(self, request):
        session_id = request.data.get("session_id")
        code = request.data.get("code")

        if not session_id or not code:
            return Response({"error": "Session ID and OTP code are required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            otp = OTP.objects.get(session_id=session_id, code=code)
        except OTP.DoesNotExist:
            return Response({"error": "Invalid OTP"}, status=status.HTTP_400_BAD_REQUEST)

        if otp.is_used or otp.expires_at < timezone.now():
            return Response({"error": "OTP expired or already used"}, status=status.HTTP_400_BAD_REQUEST)

        otp.is_used = True
        otp.save()

        # Generate token for frontend
        token, _ = Token.objects.get_or_create(user=otp.user)

        return Response({
            "message": "OTP verified successfully",
            "token": token.key,
            "user": {
                "id": otp.user.id,
                "username": otp.user.username,
                "email": otp.user.email,
                "role": getattr(otp.user, "role", "user")
            }
        }, status=status.HTTP_200_OK)

# ---------------- CSV EXPORT FOR USERS ----------------
@api_view(["GET"])
@permission_classes([IsAdminUser])
def admin_users_export_csv(request):
    users = CustomUser.objects.all().order_by("id")

    # Create HTTP response
    response = HttpResponse(content_type="text/csv")
    timestamp = timezone.now().strftime("%Y-%m-%d_%H-%M-%S")
    response["Content-Disposition"] = f'attachment; filename="users_report_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Username", "Email", "Role", "Date Joined"])

    for user in users:
        writer.writerow([user.id, user.username, user.email, user.role, user.date_joined.strftime("%Y-%m-%d")])

    return response


# ---------------- PDF EXPORT FOR USERS ----------------
@api_view(["GET"])
@permission_classes([IsAdminUser])
def admin_users_export_pdf(request):
    users = CustomUser.objects.all().order_by("id")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Users Report", styles["Title"]), Spacer(1, 12)]

    # Table data
    data = [["ID", "Username", "Email", "Role"]]
    for u in users:
        data.append([u.id, u.username, u.email, u.role])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="users_report.pdf"'
    return response


# ---------------------------------------------------------------------------
# 🔐 Permissions
# ---------------------------------------------------------------------------
class IsAdminOrReadOnly(permissions.BasePermission):
    """
    Custom permission: full access for admin/ceo, read-only for others.
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user and request.user.is_authenticated
        return request.user and request.user.is_authenticated and request.user.role in ("admin", "ceo")

# ---------------------------------------------------------------------------
# 📊 Admin Dashboard Stats
# ---------------------------------------------------------------------------
@api_view(["GET"])
@permission_classes([IsAdminUser])
def admin_dashboard_stats(request):
    """
    Admin-only stats for dashboard cards.
    Used by frontend: /api/users/stats/
    """
    total_users = CustomUser.objects.count()
    active_users = CustomUser.objects.filter(is_active=True).count()
    admin_users = CustomUser.objects.filter(role="admin").count()
    collectors = CustomUser.objects.filter(role="collector").count()

    return Response({
        "totalUsers": total_users,
        "activeUsers": active_users,
        "adminUsers": admin_users,
        "collectors": collectors,
    })


# ---------------------------------------------------------------------------
# 🔑 Token-based Login
# ---------------------------------------------------------------------------
@method_decorator(csrf_exempt, name='dispatch')
class LoginView(views.APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data["user"]
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                "token": token.key,
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": user.role,
                },
            }, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ---------------------------------------------------------------------------
# 👤 Authenticated User Profile
# ---------------------------------------------------------------------------
class UserProfileView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Returns the logged-in user's profile data.
        Used by frontend: /api/users/profile/
        """
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


# ---------------------------------------------------------------------------
# 🚚 Collector Stats
# ---------------------------------------------------------------------------
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def collector_stats(request):
    """
    Collector-only dashboard stats.
    Used by frontend: /api/users/collector-stats/
    """
    user = request.user
    if user.role != "collector":
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

    # Example placeholder stats — replace with your real DB queries
    stats = {
        "collectedPayments": 120,
        "pendingPayments": 30,
        "totalCustomers": 45,
    }
    return Response(stats)

class AdminUserListView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        users = CustomUser.objects.all()
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)

class IsAdminOrCEO(IsAuthenticated):
    """Only Admin or CEO can manage users."""

    def has_permission(self, request, view):
        return (
            super().has_permission(request, view)
            and request.user.role in ["admin", "ceo"]
        )

class AdminUserViewSet(viewsets.ModelViewSet):
    """
    CRUD for admin/CEO to manage users
    GET: List all users
    POST: Create new user
    PUT/PATCH: Update user
    DELETE: Remove user
    """
    queryset = CustomUser.objects.all().order_by("id")
    serializer_class = UserSerializer
    permission_classes = [IsAdminOrCEO]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)  # ← return plain list, no pagination

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if serializer.is_valid():
            user = serializer.save()
            return Response(UserSerializer(user).data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({"detail": "User deleted successfully"}, status=status.HTTP_204_NO_CONTENT)

# Get all users except self
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_users(request):
    users = CustomUser.objects.exclude(id=request.user.id)
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)

# Get chat history with a user
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chat_history(request):
    other_id = request.query_params.get('user_id')
    if not other_id:
        return Response([])

    # For demo purposes, store messages in memory
    # In production, create a ChatMessage model to store messages
    if not hasattr(request.user, "_chat_messages"):
        request.user._chat_messages = []

    history = [
        m for m in request.user._chat_messages
        if m['sender'] == int(other_id) or m['receiver'] == int(other_id)
    ]
    return Response(history)

class UserListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = CustomUser.objects.exclude(id=request.user.id)  # exclude self
        serializer = UserSerializer(users, many=True, context={'request': request})
        return Response(serializer.data)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def all_users(request):
    """
    Return all users for chat (excluding current user)
    """
    users = CustomUser.objects.exclude(id=request.user.id).order_by("username")
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)

# ===================================================================
# MAIN CHAT MESSAGES API — LIST + CREATE (used by frontend)
# ===================================================================
class ChatMessageListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        recipient_id = request.query_params.get("recipient")
        room_id = request.query_params.get("room_id")

        if recipient_id:
            try:
                recipient_id = int(recipient_id)
            except (ValueError, TypeError):
                return Response([])

            messages = ChatMessage.objects.filter(
                Q(sender=user, receiver_id=recipient_id) |
                Q(sender_id=recipient_id, receiver=user)
            ).order_by("timestamp")

            serializer = ChatMessageSerializer(messages, many=True, context={'request': request})
            return Response(serializer.data)

        elif room_id:
            messages = ChatMessage.objects.filter(room__room_id=room_id).order_by("timestamp")
            serializer = ChatMessageSerializer(messages, many=True, context={'request': request})
            return Response(serializer.data)

        return Response([])

    def post(self, request):
        serializer = ChatMessageSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            message = serializer.save(sender=request.user)

            # Send via WebSocket
            channel_layer = get_channel_layer()
            room_name = f"chat_{min(request.user.id, int(request.data['receiver_id']))}_{max(request.user.id, int(request.data['receiver_id']))}"

            async_to_sync(channel_layer.group_send)(
                room_name,
                {
                    "type": "chat_message",
                    "payload": ChatMessageSerializer(message, context={'request': request}).data
                }
            )
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_last_seen(request):
    request.user.update_last_seen()
    return Response({"status": "ok"})

# users/views.py
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def collector_locations(request):
    collectors = CustomUser.objects.filter(role="collector")
    data = []

    for collector in collectors:
        # Get all villages linked to this collector through VillageCollector
        villages = collector.villages.values("name", "cell__name", "cell__sector__name")

        data.append({
            "username": collector.username,
            "branch": collector.branch,
            "villages": [
                {
                    "village": v["name"],
                    "cell": v["cell__name"],
                    "sector": v["cell__sector__name"]
                } for v in villages
            ],
            # Optional: simple generated coords for map display
            "lat": -1.95 + (hash(collector.username) % 100) * 0.001,
            "lng": 30.05 + (hash(collector.username[::-1]) % 100) * 0.001,
        })

    return Response(data)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def collector_leaderboard(request):
    collectors = CustomUser.objects.filter(role="collector")
    data = []
    for collector in collectors:
        collected = Payment.objects.filter(collector=collector, status="completed").count()
        pending = Payment.objects.filter(collector=collector, status="pending").count()
        efficiency = (collected / (collected + pending + 1e-9)) * 100
        data.append({
            "username": collector.username,
            "collected": collected,
            "pending": pending,
            "efficiency": round(efficiency, 2),
        })
    data.sort(key=lambda x: x["efficiency"], reverse=True)
    return Response(data)


class AllUsersView(generics.ListAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return User.objects.exclude(id=self.request.user.id)


# ✅ List & create messages
class MessageListCreateView(generics.ListCreateAPIView):
    serializer_class = ChatMessageSerializer

    def get_queryset(self):
        recipient_id = self.request.query_params.get('recipient')
        if recipient_id:
            return ChatMessage.objects.filter(
                models.Q(sender=self.request.user, receiver_id=recipient_id) |
                models.Q(sender_id=recipient_id, receiver=self.request.user)
            ).order_by('timestamp')
        return ChatMessage.objects.none()

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)


# ✅ Update user last_seen
class UpdateLastSeenView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        request.user.last_seen = timezone.now()
        request.user.save(update_fields=["last_seen"])
        return Response({"status": "updated"}, status=status.HTTP_200_OK)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mark_delivered(request, pk):
    try:
        m = ChatMessage.objects.get(pk=pk)
        m.delivered_at = timezone.now()
        m.save(update_fields=["delivered_at"])
        return Response({"status":"ok"})
    except ChatMessage.DoesNotExist:
        return Response({"error":"not found"}, status=404)

@api_view(["POST"])
@permission_classes([IsAuthenticated])
def mark_seen(request, pk):
    try:
        m = ChatMessage.objects.get(pk=pk)
        m.seen_at = timezone.now()
        m.save(update_fields=["seen_at"])
        return Response({"status":"ok"})
    except ChatMessage.DoesNotExist:
        return Response({"error":"not found"}, status=404)

class UploadAudioMessage(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        sender = request.user
        receiver_id = request.data.get("receiver_id")
        if not receiver_id:
            return Response({"error": "receiver_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            receiver = sender.__class__.objects.get(id=receiver_id)
        except sender.__class__.DoesNotExist:
            return Response({"error": "Receiver not found"}, status=status.HTTP_404_NOT_FOUND)

        audio_file = request.FILES.get("audio")
        if not audio_file:
            return Response({"error": "No audio file provided"}, status=status.HTTP_400_BAD_REQUEST)

        msg = ChatMessage.objects.create(
            sender=sender,
            receiver=receiver,
            attachment=audio_file,
            attachment_type="audio"
        )

        serializer = ChatMessageSerializer(msg)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def chat_users(request):
    """Return list of users available for chatting, excluding self"""
    users = User.objects.exclude(id=request.user.id).values(
        "id", "username", "profile_picture", "last_seen"
    )
    return Response(list(users))

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def sidebar_users(request):
    """
    Returns a list of users for the sidebar with last message and unread count.
    """
    current_user = request.user

    users = CustomUser.objects.exclude(id=current_user.id)
    data = []

    for u in users:
        # Last message
        last_msg = ChatMessage.objects.filter(
            sender__in=[current_user, u],
            receiver__in=[current_user, u]
        ).order_by("-timestamp").first()

        # Unread count
        unread_count = ChatMessage.objects.filter(
            sender=u,
            receiver=current_user,
            seen_at__isnull=True
        ).count()

        data.append({
            "id": u.id,
            "username": u.username,
            "profile_picture": u.profile_picture.url if u.profile_picture else None,
            "is_online": u.is_online,
            "last_message": last_msg.message if last_msg else "",
            "unread_count": unread_count,
        })

    return Response(data)

class SidebarUsersView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        users = CustomUser.objects.exclude(id=request.user.id)
        serializer = SidebarUserSerializer(users, many=True, context={"current_user_id": request.user.id})
        return Response(serializer.data)

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def mark_delivered(request):
    message_ids = request.data.get("message_ids", [])
    updated = ChatMessage.objects.filter(id__in=message_ids, receiver=request.user, delivered_at__isnull=True)
    updated.update(delivered_at=timezone.now())
    return Response({"updated": updated.count()})

@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def mark_seen(request):
    message_ids = request.data.get("message_ids", [])
    updated = ChatMessage.objects.filter(id__in=message_ids, receiver=request.user, seen_at__isnull=True)
    updated.update(seen_at=timezone.now())
    return Response({"updated": updated.count()})

# views.py
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def react_to_message(request):
    message_id = request.data.get('message_id')
    emoji = request.data.get('emoji')
    # toggle reaction
    reaction, created = MessageReaction.objects.get_or_create(
        message_id=message_id, user=request.user, emoji=emoji
    )
    if not created:
        reaction.delete()
        action = "removed"
    else:
        action = "added"

    return Response({"status": action, "emoji": emoji})

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def delete_message(request):
    msg = ChatMessage.objects.get(id=request.data['message_id'], sender=request.user)
    msg.is_deleted = True
    msg.deleted_for_everyone = request.data.get('for_everyone', False)
    msg.deleted_at = timezone.now()
    msg.save()
    return Response({"status": "deleted"})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_chat(request):
    q = request.query_params.get('q')
    user_id = request.query_params.get('user_id')
    messages = ChatMessage.objects.filter(
        Q(sender=request.user, receiver_id=user_id) |
        Q(receiver=request.user, sender_id=user_id),
        message__icontains=q,
        is_deleted=False
    ).order_by('-timestamp')[:100]
    serializer = ChatMessageSerializer(messages, many=True, context={'request': request})
    return Response(serializer.data)

# views.py
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def search_messages(request):
    q = request.query_params.get('q', '').strip()
    room_id = request.query_params.get('room_id')  # group or 1-on-1
    user_id = request.query_params.get('user_id')  # for 1-on-1

    if not q:
        return Response([])

    queryset = ChatMessage.objects.filter(
        models.Q(message__icontains=q) |
        models.Q(search_vector=q)
    ).order_by('-timestamp')

    if room_id:
        queryset = queryset.filter(room__room_id=room_id)
    elif user_id:
        other = CustomUser.objects.get(id=user_id)
        queryset = queryset.filter(
            models.Q(sender=request.user, receiver=other) |
            models.Q(sender=other, receiver=request.user)
        )

    results = ChatMessageSerializer(queryset[:100], many=True, context={'request': request}).data
    return Response(results)

# views.py
class CreateGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        name = request.data.get("name")
        member_ids = request.data.get("members", [])  # list of user IDs

        room = ChatRoom.objects.create(
            name=name,
            type='group',
            creator=request.user
        )

        # Add creator as admin
        RoomMember.objects.create(room=room, user=request.user, role='admin')

        # Add members
        for user_id in member_ids:
            try:
                user = CustomUser.objects.get(id=user_id)
                RoomMember.objects.create(room=room, user=user, role='member')
            except CustomUser.DoesNotExist:
                continue

        return Response(ChatRoomSerializer(room).data, status=201)


class MyRoomsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # CORRECT: use the reverse relation name "memberships"
        rooms = ChatRoom.objects.filter(memberships__user=request.user).distinct()

        # Optional: add last message preview & unread count
        serialized_rooms = []
        for room in rooms:
            data = ChatRoomSerializer(room).data

            # Last message
            last_msg = room.messages.order_by('-timestamp').first()
            data['last_message'] = last_msg.message if last_msg else None
            data['last_message_time'] = last_msg.timestamp.isoformat() if last_msg else None

            # Unread count
            unread = room.messages.filter(
                seen_at__isnull=True
            ).exclude(sender=request.user).count()
            data['unread_count'] = unread

            serialized_rooms.append(data)

        return Response(serialized_rooms)



class AddMemberToGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        user_id = request.data.get("user_id")
        try:
            user_to_add = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

        membership, created = RoomMember.objects.get_or_create(
            room=room,
            user=user_to_add,
            defaults={'role': 'member', 'added_by': request.user}
        )

        if not created:
            return Response({"error": "User already in group"}, status=400)

        # Broadcast to group
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "member_added",
                "user": UserSerializer(user_to_add).data,
                "added_by": request.user.username
            }
        )

        return Response({"success": f"{user_to_add.username} added to group"}, status=200)


class RemoveMemberFromGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        user_id = request.data.get("user_id")
        if user_id == request.user.id:
            return Response({"error": "You cannot remove yourself as admin"}, status=400)

        try:
            membership = RoomMember.objects.get(room=room, user_id=user_id)
            username = membership.user.username
            membership.delete()

            # Broadcast removal
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{room.room_id}",
                {
                    "type": "member_removed",
                    "user_id": user_id,
                    "removed_by": request.user.username
                }
            )

            return Response({"success": f"{username} removed from group"})
        except RoomMember.DoesNotExist:
            return Response({"error": "User not in group"}, status=404)


class MakeAdminView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        user_id = request.data.get("user_id")
        try:
            membership = RoomMember.objects.get(room=room, user_id=user_id)
            membership.role = 'admin'
            membership.save()

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{room.room_id}",
                {
                    "type": "admin_promoted",
                    "user_id": user_id,
                    "promoted_by": request.user.username
                }
            )

            return Response({"success": f"{membership.user.username} is now admin"})
        except RoomMember.DoesNotExist:
            return Response({"error": "User not in group"}, status=404)




class UpdateGroupInfoView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        name = request.data.get("name")
        description = request.data.get("description")
        image = request.FILES.get("image")

        if name:
            room.name = name
        if description is not None:
            room.description = description
        if image:
            room.image = image

        room.save()

        # Broadcast update
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "group_updated",
                "room": ChatRoomSerializer(room).data
            }
        )

        return Response(ChatRoomSerializer(room).data)

# === LEAVE GROUP ===
class LeaveGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        try:
            membership = RoomMember.objects.get(room=room, user=request.user)
        except RoomMember.DoesNotExist:
            return Response({"error": "You are not in this group"}, status=400)

        # If last admin leaves → auto-promote another member or delete group
        if membership.role == 'admin':
            admin_count = room.memberships.filter(role='admin').count()
            if admin_count == 1:
                # Find another member to promote
                other_member = room.memberships.exclude(user=request.user).first()
                if other_member:
                    other_member.role = 'admin'
                    other_member.save()
                    broadcast_msg = f"{request.user.username} left and transferred admin to {other_member.user.username}"
                else:
                    # No one left → delete group
                    room.delete()
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        f"chat_{room.room_id}",
                        {"type": "group_deleted", "message": "Group deleted: last admin left"}
                    )
                    return Response({"success": "You left and group was deleted"})
            else:
                broadcast_msg = f"{request.user.username} left the group"
        else:
            broadcast_msg = f"{request.user.username} left the group"

        membership.delete()

        # Broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "member_left",
                "user_id": request.user.id,
                "username": request.user.username,
                "message": broadcast_msg
            }
        )

        return Response({"success": "You have left the group"})


# === BAN USER ===
class BanUserView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        user_id = request.data.get("user_id")
        try:
            user_to_ban = CustomUser.objects.get(id=user_id)
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=404)

        # Prevent banning creator
        if room.creator == user_to_ban:
            return Response({"error": "Cannot ban the group creator"}, status=400)

        # Remove from members + add to banned
        RoomMember.objects.filter(room=room, user=user_to_ban).delete()
        room.banned_users.add(user_to_ban)

        # Broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "user_banned",
                "user_id": user_to_ban.id,
                "username": user_to_ban.username,
                "banned_by": request.user.username,
                "message": f"{user_to_ban.username} was banned from the group"
            }
        )

        return Response({"success": f"{user_to_ban.username} has been banned"})


# === UNBAN USER ===
class UnbanUserView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        user_id = request.data.get("user_id")
        try:
            user = CustomUser.objects.get(id=user_id)
            room.banned_users.remove(user)
            return Response({"success": f"{user.username} unbanned and can rejoin"})
        except CustomUser.DoesNotExist:
            return Response({"error": "User not found"}, status=404)


# === DELETE GROUP ===
class DeleteGroupView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        # Only creator or admin can delete
        is_admin = room.memberships.filter(user=request.user, role='admin').exists()
        is_creator = room.creator == request.user

        if not (is_admin or is_creator):
            return Response({"error": "Only group creator or admin can delete"}, status=403)

        room_name = room.name or "this group"
        room.delete()

        # Broadcast deletion
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room_id}",
            {
                "type": "group_deleted",
                "message": f"{room_name} has been deleted by {request.user.username}"
            }
        )

        return Response({"success": "Group deleted permanently"})


# === TOGGLE ADMIN-ONLY MESSAGING ===
class ToggleAdminOnlyMessagingView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        room.only_admins_can_send = not room.only_admins_can_send
        room.save()

        status = "ON" if room.only_admins_can_send else "OFF"
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "admin_only_toggled",
                "only_admins_can_send": room.only_admins_can_send,
                "message": f"Admin-only messaging turned {status}"
            }
        )

        return Response({
            "success": True,
            "only_admins_can_send": room.only_admins_can_send
        })

# users/views.py
class GroupInfoView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id)
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        # Check if user is member
        is_member = room.memberships.filter(user=request.user).exists()
        if not is_member and not room.is_public:
            return Response({"error": "You are not a member"}, status=403)

        # Serialize full info
        serializer = ChatRoomSerializer(room)
        data = serializer.data

        # Add extra info
        data.update({
            "is_member": is_member,
            "my_role": None,
            "invite_link": room.invite_link if room.allow_invite_link and room.is_invite_valid() else None,
            "invite_expires_at": room.invite_expires_at,
            "banned_users": [
                {"id": u.id, "username": u.username, "profile_picture": u.profile.profile_picture.url if u.profile.profile_picture else None}
                for u in room.banned_users.all()
            ],
            "pending_requests": []  # future
        })

        if is_member:
            membership = room.memberships.get(user=request.user)
            data["my_role"] = membership.role

        return Response(data)

# Generate / Regenerate Invite Link
class GenerateInviteLinkView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        expires_in = int(request.data.get("expires_in", 168))  # 7 days default
        max_uses = request.data.get("max_uses", None)
        if max_uses: max_uses = int(max_uses)

        code = room.generate_invite_link(expires_in_hours=expires_in, max_uses=max_uses)

        return Response({
            "invite_link": room.invite_link,
            "code": code,
            "expires_at": room.invite_expires_at,
            "max_uses": max_uses,
            "message": "New invite link generated"
        })


# Join via Invite Link
class JoinViaInviteView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        code = request.data.get("code")
        if not code:
            return Response({"error": "Invite code required"}, status=400)

        try:
            room = ChatRoom.objects.get(invite_code=code)
        except ChatRoom.DoesNotExist:
            return Response({"error": "Invalid or expired invite link"}, status=404)

        if not room.is_invite_valid():
            return Response({"error": "This invite link has expired or reached max uses"}, status=400)

        if room.banned_users.filter(id=request.user.id).exists():
            return Response({"error": "You are banned from this group"}, status=403)

        if room.memberships.filter(user=request.user).exists():
            return Response({"error": "You are already a member"}, status=400)

        # Join group
        RoomMember.objects.create(
            room=room,
            user=request.user,
            role='member'
        )

        # Increment use count
        room.invite_used_count += 1
        room.save(update_fields=['invite_used_count'])

        # Broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "member_joined_via_link",
                "user": UserSerializer(request.user).data,
                "message": f"{request.user.username} joined using invite link"
            }
        )

        return Response({
            "success": True,
            "room": ChatRoomSerializer(room).data,
            "message": "You have joined the group!"
        })

# users/views.py

class SetAnnouncementView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        text = request.data.get("announcement", "").strip()
        room.announcement = text
        room.announcement_updated_at = timezone.now()
        room.announcement_updated_by = request.user
        room.save()

        # Broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "announcement_updated",
                "announcement": text,
                "updated_by": request.user.username,
                "timestamp": timezone.now().isoformat()
            }
        )

        return Response({"success": True, "announcement": text})


class PinMessageView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
            message = ChatMessage.objects.get(id=request.data["message_id"], room=room)
        except (ChatRoom.DoesNotExist, ChatMessage.DoesNotExist):
            return Response({"error": "Invalid group or message"}, status=404)

        require_group_admin(request.user, room)

        # Unpin old one
        if room.pinned_message:
            old = room.pinned_message
            old.is_announcement = False
            old.save()

        room.pinned_message = message
        message.is_announcement = True
        message.save()
        room.save()

        # Broadcast
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"chat_{room.room_id}",
            {
                "type": "message_pinned",
                "message_id": message.id,
                "message": ChatMessageSerializer(message).data
            }
        )

        return Response({"success": True, "pinned_message": ChatMessageSerializer(message).data})


class UnpinMessageView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, room_id):
        try:
            room = ChatRoom.objects.get(room_id=room_id, type='group')
        except ChatRoom.DoesNotExist:
            return Response({"error": "Group not found"}, status=404)

        require_group_admin(request.user, room)

        if room.pinned_message:
            old = room.pinned_message
            old.is_announcement = False
            old.save()
            room.pinned_message = None
            room.save()

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"chat_{room.room_id}",
                {"type": "message_unpinned"}
            )

        return Response({"success": True})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_chat_pdf(request):
    room_id = request.query_params.get('room_id')
    user_id = request.query_params.get('user_id')

    messages = get_messages(room_id, user_id, request.user)[:1000]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="chat_export_{timezone.now().strftime("%Y%m%d")}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph("Chat Export", styles['Title']), Spacer(1, 12)]

    for msg in messages:
        sender = msg.sender.username
        time = msg.timestamp.strftime("%b %d, %I:%M %p")
        text = msg.message or "[Attachment]"
        story.append(Paragraph(f"<b>{sender}</b> • {time}<br/>{text}", styles['Normal']))
        story.append(Spacer(1, 6))

    doc.build(story)
    return response

@api_view(['POST'])
def start_call(request):
    target_id = request.data['to']
    call_id = secrets.token_hex(16)
    # Save call metadata if needed
    return Response({"call_id": call_id, "url": f"/call/{call_id}/"})

# ===================================================================
# DELETE MESSAGE — Supports "Delete for Me" and "Delete for Everyone"
# ===================================================================
class DeleteMessageView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, message_id):
        try:
            message = ChatMessage.objects.get(id=message_id)
        except ChatMessage.DoesNotExist:
            return Response({"error": "Message not found"}, status=404)

        delete_for_everyone = request.query_params.get("for_everyone", "false").lower() == "true"

        # Security: Only sender can delete for everyone
        if delete_for_everyone and message.sender != request.user:
            return Response({"error": "You can only delete your own messages for everyone"}, status=403)

        # Security: Only participants can delete for themselves
        is_participant = (
                message.sender == request.user or
                message.receiver == request.user or
                (message.room and message.room.memberships.filter(user=request.user).exists())
        )
        if not is_participant:
            return Response({"error": "Not authorized"}, status=403)

        if delete_for_everyone:
            # Full delete + notify everyone
            message.is_deleted = True
            message.deleted_for_everyone = True
            message.deleted_at = timezone.now()
            message.message = ""  # Optional: wipe content
            message.attachment = None
            message.save()

            action = "deleted_for_everyone"
        else:
            # Delete only for current user (future: use deleted_for field)
            message.is_deleted = True
            message.deleted_at = timezone.now()
            message.save()
            action = "deleted_for_me"

        # Determine room name for WebSocket broadcast
        if message.room:
            room_name = message.room.room_id
        else:
            room_name = f"chat_{min(message.sender.id, message.receiver.id)}_{max(message.sender.id, message.receiver.id)}"

        # Send real-time delete event
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            room_name,
            {
                "type": "message_deleted",
                "message_id": message_id,
                "deleted_by": request.user.id,
                "action": action  # "deleted_for_me" or "deleted_for_everyone"
            }
        )

        return Response({
            "status": "deleted",
            "message_id": message_id,
            "action": action
        })

@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
def notification_settings(request):
    user = request.user

    if request.method == 'GET':
        return Response({
            "notify_realtime": user.notify_realtime,
            "notify_email": user.notify_email,
            "notify_sms": user.notify_sms,
            "notify_browser": user.notify_browser,
            "notify_sound": user.notify_sound,
            "notify_payment": user.notify_payment,
            "notify_customer_update": user.notify_customer_update,
            "notify_chat": user.notify_chat,
            "notify_task": user.notify_task,
            "notify_leave": user.notify_leave,
            "notify_system": user.notify_system,
        })

    elif request.method == 'PATCH':
        data = request.data
        for field in data:
            if hasattr(user, field):
                setattr(user, field, data[field])
        user.save()
        return Response({"detail": "Notification settings updated"})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def analytics_summary(request):
    """
    Comprehensive analytics endpoint for High Prosper Admin Dashboard (2026)
    Fixed: PostgreSQL-compatible date grouping with TO_CHAR()
    """
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    village = request.GET.get('village')
    method_name = request.GET.get('method')

    # Base querysets
    payments = Payment.objects.all()
    customers = Customer.objects.all()
    collectors = Collector.objects.filter(is_active=True, is_deleted=False)

    # Apply date range filters
    if start_date:
        payments = payments.filter(created_at__date__gte=start_date)
        customers = customers.filter(created_at__date__gte=start_date)
    if end_date:
        payments = payments.filter(created_at__date__lte=end_date)
        customers = customers.filter(created_at__date__lte=end_date)

    # Apply village filter
    if village:
        payments = payments.filter(customer__village__name=village)
        customers = customers.filter(village__name=village)

    # Apply payment method filter
    if method_name:
        try:
            method_obj = PaymentMethod.objects.get(name=method_name)
            payments = payments.filter(method=method_obj)
        except PaymentMethod.DoesNotExist:
            pass

    # === 1. Total Revenue ===
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # === 2. Collections (Successful payments) ===
    collections = payments.filter(status='Successful').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # === 3. Pending Payments ===
    pending = payments.filter(status='Pending').aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # === Previous Period for Growth Comparison ===
    prev_start = (timezone.now().date() - timedelta(days=60)).strftime('%Y-%m-%d')
    prev_end = (timezone.now().date() - timedelta(days=30)).strftime('%Y-%m-%d')
    prev_payments = Payment.objects.filter(created_at__date__gte=prev_start, created_at__date__lte=prev_end)
    prev_revenue = prev_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    revenue_growth = ((total_revenue - prev_revenue) / prev_revenue * 100) if prev_revenue > 0 else Decimal('0')

    # === 4. Payment Methods Breakdown ===
    payment_methods = payments.values('method__name').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-count')
    payment_methods_dict = {
        item['method__name'] or 'Unknown': {
            'count': item['count'],
            'total': item['total'] or Decimal('0')
        }
        for item in payment_methods
    }

    # Most Used Method
    most_used_method = None
    most_used_count = 0
    if payment_methods_dict:
        most_used_method, data = max(payment_methods_dict.items(), key=lambda x: x[1]['count'])
        most_used_count = data['count']

    # === 5. Top Customers ===
    top_customers = payments.values(
        'customer__name', 'customer__village__name'
    ).annotate(
        total_paid=Sum('amount')
    ).order_by('-total_paid')[:10]

    # === 6. Customer Growth (PostgreSQL: TO_CHAR) ===
    customer_growth = customers.extra({
        'month': "TO_CHAR(created_at, 'YYYY-MM')"
    }).values('month').annotate(count=Count('id')).order_by('month')

    # === 7. Daily, Weekly, Monthly Revenue (PostgreSQL) ===
    daily_revenue = payments.extra({
        'date': "DATE(created_at)"
    }).values('date').annotate(amount=Sum('amount')).order_by('date')

    weekly_revenue = payments.extra({
        'week': "EXTRACT(week FROM created_at)"
    }).values('week').annotate(amount=Sum('amount')).order_by('week')

    monthly_revenue = payments.extra({
        'month': "TO_CHAR(created_at, 'YYYY-MM')"
    }).values('month').annotate(amount=Sum('amount')).order_by('month')

    # === 8. Collector Performance with Monthly Trends & Incentives ===
    collector_performance = []

    for collector in collectors:
        collector_payments = payments.filter(customer__village__collectors=collector)

        monthly_trends = collector_payments.extra({
            'month': "TO_CHAR(created_at, 'YYYY-MM')"
        }).values('month').annotate(
            collected=Sum('amount', filter=Q(status='Successful')),
            pending=Sum('amount', filter=Q(status='Pending'))
        ).order_by('month')

        total_collected = collector_payments.filter(status='Successful').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        pending_collected = collector_payments.filter(status='Pending').aggregate(total=Sum('amount'))['total'] or Decimal('0')
        customers_count = Customer.objects.filter(village__collectors=collector).count()
        avg_per_customer = total_collected / Decimal(customers_count) if customers_count > 0 else Decimal('0')

        # Tiered Bonus Calculation
        bonus = Decimal('0')
        if total_collected > collector.target_amount * Decimal('1.2'):
            bonus = total_collected * Decimal('0.15')
        elif total_collected >= collector.target_amount:
            bonus = total_collected * Decimal('0.10')
        elif total_collected >= collector.target_amount * Decimal('0.9'):
            bonus = total_collected * Decimal('0.05')

        collector_performance.append({
            "collector": collector.user.get_full_name() or collector.user.username,
            "villages": ", ".join(v.name for v in collector.villages.all()) or "None",
            "customers": customers_count,
            "total_collected": float(total_collected),
            "pending": float(pending_collected),
            "avg_per_customer": float(avg_per_customer),
            "bonus": float(bonus),
            "target_amount": float(collector.target_amount or 0),
            "monthly_trends": [
                {
                    "month": item['month'],
                    "collected": float(item['collected'] or 0),
                    "pending": float(item['pending'] or 0),
                    "avg": float((item['collected'] or 0) / Decimal(customers_count) if customers_count > 0 else 0)
                }
                for item in monthly_trends
            ]
        })

    # Sort by total collected
    collector_performance.sort(key=lambda x: x['total_collected'], reverse=True)

    # === AI-Powered Insights ===
    insights = []

    # Top Performing Village
    village_performance = payments.values('customer__village__name').annotate(total=Sum('amount')).order_by('-total')[:3]
    if village_performance:
        top_village = village_performance[0]
        insights.append({
            "title": "Top Performing Village",
            "message": f"{top_village['customer__village__name']} generated {top_village['total']:,} RWF — focus marketing here!",
            "type": "success"
        })

    # Fastest Growing Village
    growth_by_village = customers.values('village__name').annotate(count=Count('id')).order_by('-count')[:3]
    if growth_by_village:
        fastest = growth_by_village[0]
        insights.append({
            "title": "Fastest Growing Village",
            "message": f"{fastest['village__name']} added {fastest['count']} new customers this period!",
            "type": "growth"
        })

    # Payment Method Recommendation
    if most_used_method:
        insights.append({
            "title": "Recommended Payment Method",
            "message": f"{most_used_method} is most popular ({most_used_count} transactions) — promote it more!",
            "type": "info"
        })

    # Risk Alert: High Pending
    pending_ratio = (pending / total_revenue * 100) if total_revenue > 0 else 0
    if pending_ratio > 30:
        insights.append({
            "title": "High Pending Risk",
            "message": f"Pending payments are {pending_ratio:.1f}% of revenue — follow up urgently!",
            "type": "warning"
        })

    # Daily Revenue Alert
    today_amount = Decimal('0')
    if daily_revenue:
        last_entry = daily_revenue.order_by('-date').first()
        today_amount = last_entry['amount'] if last_entry else Decimal('0')

    avg_daily = Decimal('0')
    if daily_revenue:
        total_daily = sum(d['amount'] or Decimal('0') for d in daily_revenue)
        avg_daily = total_daily / Decimal(len(daily_revenue)) if daily_revenue else Decimal('0')

    if today_amount < avg_daily * Decimal('0.7'):
        insights.append({
            "title": "Daily Revenue Alert",
            "message": f"Today's revenue ({today_amount:,} RWF) is below average — check activity!",
            "type": "warning"
        })

    # === AI-Generated Report Summary ===
    summary_parts = []
    if revenue_growth > 0:
        summary_parts.append(f"Revenue grew by {revenue_growth:.1f}%")
    else:
        summary_parts.append(f"Revenue declined by {abs(revenue_growth):.1f}%")

    if pending_ratio > 30:
        summary_parts.append("high pending payments detected")
    elif pending_ratio > 15:
        summary_parts.append("moderate pending payments")

    if most_used_method:
        summary_parts.append(f"{most_used_method} is the dominant payment method")

    report_summary = "Summary: " + ", ".join(summary_parts) + "." if summary_parts else "No data available yet."

    # === Email Alerts for Critical Insights ===
    critical_insights = [i for i in insights if i["type"] == "warning"]
    if critical_insights:
        subject = "URGENT: Critical Insights Detected in High Prosper Analytics"
        message = "\n".join([f"- {i['title']}: {i['message']}" for i in critical_insights])
        message += f"\n\nFull Report: {report_summary}"
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [settings.ADMIN_EMAIL],
            fail_silently=False,
        )

    # === Final Response ===
    response_data = {
        "totalRevenue": float(total_revenue),
        "collections": float(collections),
        "pending": float(pending),
        "paymentMethods": payment_methods_dict,
        "topCustomers": list(top_customers),
        "customerGrowth": list(customer_growth),
        "revenue": {
            "daily": list(daily_revenue),
            "weekly": list(weekly_revenue),
            "monthly": list(monthly_revenue),
        },
        "collectorPerformance": collector_performance,
        "insights": insights,
        "reportSummary": report_summary,
        "revenueGrowth": float(revenue_growth)
    }

    return Response(response_data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_report_email(request):
    pdf_file = request.FILES.get('pdf')
    if not pdf_file:
        return Response({"error": "No PDF file provided"}, status=400)

    email = EmailMessage(
        subject=f"High Prosper Analytics Report - {timezone.now().strftime('%Y-%m-%d')}",
        body="Please find attached the latest analytics report.",
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[settings.ADMIN_EMAIL],  # CEO/admin email
    )
    email.attach(pdf_file.name, pdf_file.read(), "application/pdf")
    email.send()

    return Response({"message": "Report email sent successfully"})

class UserAnalyticsAPIView(APIView):
    """
    GET /api/v1/users/admin/analytics/
    Comprehensive analytics for advanced admin dashboard charts
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        now = timezone.now()
        online_threshold = now - timedelta(minutes=5)      # considered online
        inactive_threshold = now - timedelta(days=30)      # inactive if no activity
        today = now.date()
        this_month_start = today.replace(day=1)
        month_ago = now - timedelta(days=30)
        year_ago = now - timedelta(days=365)

        # Base queryset
        all_users = CustomUser.objects.all()

        # ─── Basic Counts ───────────────────────────────────────────────────────
        total_users = all_users.count()
        users_by_role = dict(
            all_users.values('role').annotate(count=Count('id')).values_list('role', 'count')
        )

        total_online = all_users.filter(
            is_active=True,
            last_seen__gte=online_threshold
        ).count()
        total_offline = total_users - total_online

        new_today = all_users.filter(date_joined__date=today).count()
        new_month = all_users.filter(date_joined__gte=this_month_start).count()

        blocked_users = all_users.filter(
            Exists(BlockedUser.objects.filter(blocked=OuterRef('pk')))
        ).count()

        inactive_users = all_users.filter(
            is_active=True,
            last_seen__lt=inactive_threshold
        ).count()

        deleted_users = all_users.filter(is_deleted=True).count()

        # ─── Status Breakdown (Donut Chart) ─────────────────────────────────────
        status_breakdown = {
            'active': all_users.filter(is_active=True).count(),
            'inactive': all_users.filter(is_active=False).count(),
            'verified': all_users.filter(is_verified=True).count(),
            'unverified': all_users.filter(is_verified=False).count(),
            'online': total_online,
            'offline': total_offline,
        }

        # ─── Monthly User Growth (Line Chart) ───────────────────────────────────
        monthly_growth_raw = (
            all_users
            .annotate(month=TruncMonth('date_joined'))
            .values('month')
            .annotate(registrations=Count('id'))
            .order_by('month')
        )

        monthly_growth = [
            {
                'month': entry['month'].strftime('%Y-%m'),
                'registrations': entry['registrations']
            }
            for entry in monthly_growth_raw
        ]

        # ─── Role Trend Over Time (Stacked Area Chart) ──────────────────────────
        role_trend_raw = (
            all_users
            .annotate(month=TruncMonth('date_joined'))
            .values('month', 'role')
            .annotate(count=Count('id'))
            .order_by('month', 'role')
        )

        role_trend = {}
        for entry in role_trend_raw:
            month_str = entry['month'].strftime('%Y-%m')
            role = entry['role'] or 'Unknown'
            if month_str not in role_trend:
                role_trend[month_str] = {}
            role_trend[month_str][role] = entry['count']

        role_trend_formatted = sorted(
            [{'month': month, **counts} for month, counts in role_trend.items()],
            key=lambda x: x['month']
        )

        # Previous period (last 12 months for comparison)
        previous_role_trend_raw = (
            all_users
            .filter(date_joined__lte=year_ago)
            .annotate(month=TruncMonth('date_joined'))
            .values('month', 'role')
            .annotate(count=Count('id'))
            .order_by('month', 'role')
        )

        previous_role_trend = {}
        for entry in previous_role_trend_raw:
            month_str = entry['month'].strftime('%Y-%m')
            role = entry['role'] or 'Unknown'
            if month_str not in previous_role_trend:
                previous_role_trend[month_str] = {}
            previous_role_trend[month_str][role] = entry['count']

        previous_role_trend_formatted = sorted(
            [{'month': month, **counts} for month, counts in previous_role_trend.items()],
            key=lambda x: x['month']
        )

        # ─── Activity Heatmap (Day/Hour Matrix - last 30 days) ──────────────────
        activity_heatmap_raw = (
            all_users
            .filter(last_seen__gte=month_ago)
            .annotate(
                day=ExtractIsoWeekDay('last_seen'), # ← Correct name
                hour=ExtractHour('last_seen')
            )
            .values('day', 'hour')
            .annotate(count=Count('id'))
            .order_by('day', 'hour')
        )

        activity_heatmap = [
            {'day': entry['day'], 'hour': entry['hour'], 'value': entry['count']}
            for entry in activity_heatmap_raw
        ]

        # ─── Logged-in Stats (from Activity model) ──────────────────────────────
        logged_in_total = Activity.objects.filter(action_type='login').count()
        logged_in_month = Activity.objects.filter(
            action_type='login',
            created_at__gte=this_month_start
        ).count()

        # ─── Usage Hours & Performance ──────────────────────────────────────────
        # Placeholder: 30 min average session per login
        usage_hours = round(logged_in_month * 0.5, 1)

        active_7d = all_users.filter(last_seen__gte=now - timedelta(days=7)).count()
        performance_percentage = round((active_7d / total_users * 100), 1) if total_users > 0 else 0

        # ─── Final Response Data ────────────────────────────────────────────────
        data = {
            "total_users": total_users,
            "users_by_role": users_by_role,
            "total_online": total_online,
            "total_offline": total_offline,
            "new_users_today": new_today,
            "new_users_month": new_month,
            "blocked_users": blocked_users,
            "inactive_users": inactive_users,
            "deleted_users": deleted_users,
            "logged_in_total": logged_in_total,
            "logged_in_month": logged_in_month,
            "usage_hours": usage_hours,
            "performance_percentage": performance_percentage,

            # Chart-specific data
            "monthly_growth": monthly_growth,               # Line chart
            "role_trend": role_trend_formatted,             # Stacked area chart
            "role_trend_previous": previous_role_trend_formatted,  # Comparison
            "activity_heatmap": activity_heatmap,           # Heatmap + zoom
            "status_breakdown": status_breakdown,           # Donut chart
        }

        return Response(data)

class ExportUsersPDFAPIView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        print("PDF Export requested by:", request.user, "Is staff:", request.user.is_staff)

        try:
            # Filters
            queryset = CustomUser.objects.all()
            role = request.query_params.get('role')
            if role:
                queryset = queryset.filter(role=role)
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(username__icontains=search) |
                    Q(email__icontains=search) |
                    Q(first_name__icontains=search) |
                    Q(last_name__icontains=search)
                )
            is_active_str = request.query_params.get('is_active')
            if is_active_str is not None:
                is_active_bool = is_active_str.lower() in ('true', '1', 'yes')
                queryset = queryset.filter(is_active=is_active_bool)
            date_from = request.query_params.get('date_from')
            if date_from:
                try:
                    dt_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                    queryset = queryset.filter(date_joined__date__gte=dt_from)
                except ValueError:
                    pass
            date_to = request.query_params.get('date_to')
            if date_to:
                try:
                    dt_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                    queryset = queryset.filter(date_joined__date__lte=dt_to)
                except ValueError:
                    pass

            users = queryset.order_by('-date_joined')

            # Summary stats
            total_users = users.count()
            users_by_role = dict(
                users.values('role').annotate(count=Count('id')).values_list('role', 'count')
            )
            total_online = users.filter(is_online=True).count()
            total_offline = total_users - total_online
            now = timezone.now()
            today = now.date()
            month_start = today.replace(day=1)
            new_today = users.filter(date_joined__date=today).count()
            new_month = users.filter(date_joined__gte=month_start).count()
            inactive_users = users.filter(is_active=True, last_seen__lt=now - timedelta(days=30)).count()

            # PDF response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="high_prosper_users_report.pdf"'

            doc = SimpleDocTemplate(
                response,
                pagesize=landscape(letter),
                rightMargin=0.4*inch,
                leftMargin=0.4*inch,
                topMargin=1.2*inch,
                bottomMargin=1.0*inch
            )

            elements = []
            styles = getSampleStyleSheet()

            title_style = ParagraphStyle(name='Title', fontSize=20, leading=24, textColor=colors.darkblue, spaceAfter=10, alignment=1)
            heading_style = ParagraphStyle(name='Heading2', fontSize=14, leading=18, textColor=colors.darkblue, spaceAfter=6)
            normal_style = styles['Normal']

            def add_header(canvas, doc):
                canvas.saveState()
                if os.path.exists(LOGO_PATH):
                    try:
                        logo = ReportLabImage(LOGO_PATH, width=1.5*inch, height=1.0*inch)
                        logo.drawOn(canvas, 0.4*inch, doc.height + doc.topMargin - 1.2*inch)
                    except Exception as e:
                        print(f"Logo error: {e}")
                canvas.restoreState()

            def add_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont("Helvetica", 8)
                canvas.setFillColor(colors.darkgray)

                page_num = canvas.getPageNumber()
                y = 0.3*inch

                if page_num == 1:
                    canvas.drawString(0.4*inch, y, f"{COMPANY_NAME} • {COMPANY_LOCATION} • {COMPANY_PHONE} • {COMPANY_EMAIL}")

                    x = doc.width + doc.rightMargin - 4.5*inch
                    canvas.drawString(x, y + 0.08*inch, "Follow us:")
                    x += canvas.stringWidth("Follow us:", "Helvetica", 8) + 0.1*inch

                    icon_size = 0.25*inch
                    for platform, data in SOCIAL_ICONS.items():
                        icon_path = data["icon"]
                        url = data["url"]

                        if os.path.exists(icon_path):
                            try:
                                icon = ReportLabImage(icon_path, width=icon_size, height=icon_size)
                                icon.drawOn(canvas, x, y - 0.04*inch)
                            except:
                                canvas.drawString(x, y, platform[:3])
                        else:
                            canvas.drawString(x, y, platform[:3])

                        canvas.linkURL(url, (x, y-0.08*inch, x+icon_size, y+icon_size+0.08*inch))
                        x += icon_size + 0.15*inch

                    canvas.setStrokeColor(colors.lightgrey)
                    canvas.line(0.4*inch, y-0.15*inch, doc.width + doc.rightMargin - 0.4*inch, y-0.15*inch)

                canvas.setFont("Helvetica-Oblique", 7)
                canvas.drawCentredString(doc.width/2 + doc.leftMargin, 0.2*inch,
                                         f"Page {page_num} • Confidential Document")

                canvas.restoreState()

            # Apply templates (no frames needed for full-page default)
            doc.addPageTemplates([
                PageTemplate(
                    id='AllPages',
                    onPage=add_header,
                    onPageEnd=add_footer
                )
            ])

            # Title Page
            elements.append(Paragraph("High Prosper Services", title_style))
            elements.append(Paragraph("Users Management Report", title_style))
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')} by {request.user.get_full_name()}", normal_style))
            elements.append(Spacer(1, 0.6*inch))

            # Summary
            elements.append(Paragraph("Summary", heading_style))
            elements.append(Spacer(1, 0.15*inch))

            summary_data = [
                ["Total Users", total_users],
                ["Online", total_online],
                ["New Today", new_today],
                ["New Month", new_month],
                ["Inactive", inactive_users],
            ]

            summary_table = Table(summary_data, colWidths=[2.8*inch, 2.0*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 10),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 0.4*inch))

            # Pie Chart
            elements.append(Paragraph("Users by Role", heading_style))
            drawing = Drawing(350, 180)
            pie = Pie()
            top_roles = sorted(users_by_role.items(), key=lambda x: x[1], reverse=True)[:8]
            others = sum(count for _, count in sorted(users_by_role.items())[8:])
            pie.data = [count for _, count in top_roles] + [others]
            pie.labels = [role.capitalize()[:12] for role, _ in top_roles] + ['Others']
            pie.x = 80
            pie.y = 40
            pie.width = 160
            pie.height = 160
            drawing.add(pie)
            elements.append(drawing)
            elements.append(PageBreak())

            # Compact Table
            elements.append(Paragraph("Users List", heading_style))

            table_data = [[
                "ID", "User", "Name", "Email", "Phone", "Role", "Company", "Branch", "Login", "Joined", "Online", "Ver.", "Act."
            ]]

            for user in users:
                company_name = getattr(user.company, 'name', '—') if user.company else '—'
                branch_name = user.branch or '—'
                last_login = user.last_login.strftime("%Y-%m-%d") if user.last_login else "—"
                date_joined = user.date_joined.strftime("%Y-%m-%d")

                table_data.append([
                    str(user.id),
                    user.username[:12] + '…' if len(user.username) > 12 else user.username,
                    user.get_full_name()[:15] + '…' if len(user.get_full_name()) > 15 else user.get_full_name(),
                    user.email[:20] + '…' if len(user.email) > 20 else user.email,
                    user.phone[:12] + '...' if user.phone and len(user.phone) > 12 else (user.phone or "—"),
                    user.role[:10].capitalize() if user.role else "—",
                    company_name[:15] + '…' if len(company_name) > 15 else company_name,
                    branch_name[:12] + '…' if len(branch_name) > 12 else branch_name,
                    last_login,
                    date_joined,
                    "Y" if user.is_online else "N",
                    "Y" if user.is_verified else "N",
                    "Y" if user.is_active else "N"
                ])

            from reportlab.platypus import LongTable
            table = LongTable(
                table_data,
                repeatRows=1,
                colWidths=[0.5*inch, 1.0*inch, 1.2*inch, 1.5*inch, 0.8*inch, 0.7*inch, 1.0*inch, 0.8*inch, 0.9*inch, 0.8*inch, 0.5*inch, 0.5*inch, 0.5*inch]
            )

            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 8),
                ('BOTTOMPADDING', (0,0), (-1,0), 4),
                ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
                ('FONTSIZE', (0,1), (-1,-1), 7),
                ('LEFTPADDING', (0,0), (-1,-1), 2),
                ('RIGHTPADDING', (0,0), (-1,-1), 2),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (3,1), (3,-1), 'LEFT'),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))

            elements.append(table)

            # Build PDF
            doc.build(elements)
            print("PDF generated successfully - returning response")
            return response

        except Exception as e:
            print("PDF EXPORT FAILED!")
            print(traceback.format_exc())
            return JsonResponse({"error": str(e)}, status=500)



class ExportUsersExcelAPIView(APIView):
    """
    GET /api/v1/users/admin/users/export_excel/

    Generates professional Excel report with:
    - Users List sheet (filtered data + conditional formatting + role-based colors)
    - Summary Statistics sheet (dashboard with embedded charts)

    Supports filtering via query params (same as list view):
    - role
    - search
    - is_active (true/false)
    - date_from / date_to (YYYY-MM-DD)
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        # ─── Apply Filters (sync with frontend) ─────────────────────────────────
        queryset = CustomUser.objects.all()

        role = request.query_params.get('role')
        if role:
            queryset = queryset.filter(role=role)

        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )

        is_active_str = request.query_params.get('is_active')
        if is_active_str is not None:
            is_active_bool = is_active_str.lower() in ('true', '1', 'yes')
            queryset = queryset.filter(is_active=is_active_bool)

        date_from = request.query_params.get('date_from')
        if date_from:
            try:
                dt_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(date_joined__date__gte=dt_from)
            except ValueError:
                pass

        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                dt_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(date_joined__date__lte=dt_to)
            except ValueError:
                pass

        users = queryset.order_by('-date_joined')

        # ─── Calculate Summary Stats ────────────────────────────────────────────
        total_users = users.count()
        users_by_role = dict(
            users.values('role').annotate(count=Count('id')).values_list('role', 'count')
        )
        total_online = users.filter(is_online=True).count()
        total_offline = total_users - total_online

        now = timezone.now()
        today = now.date()
        month_start = today.replace(day=1)
        new_today = users.filter(date_joined__date=today).count()
        new_month = users.filter(date_joined__gte=month_start).count()

        blocked_users = 0  # Replace with real BlockedUser logic if needed
        inactive_users = users.filter(is_active=True, last_seen__lt=now - timedelta(days=30)).count()
        deleted_users = 0  # Add soft-delete later if needed

        active_last_week = users.filter(last_seen__gte=now - timedelta(days=7)).count()
        performance_percentage = round((active_last_week / total_users * 100), 1) if total_users > 0 else 0

        # Daily new users trend (last 30 days for line chart)
        daily_new = []
        for day in range(29, -1, -1):
            day_date = today - timedelta(days=day)
            count = users.filter(date_joined__date=day_date).count()
            daily_new.append((day_date.strftime('%Y-%m-%d'), count))

        # ─── Create Workbook ────────────────────────────────────────────────────
        wb = Workbook()
        ws_list = wb.active
        ws_list.title = "Users List"

        # ─── SHEET 1: Users List ────────────────────────────────────────────────
        # Title
        ws_list.merge_cells('A1:O1')
        ws_list['A1'] = "High Prosper Services - Users List"
        ws_list['A1'].font = Font(size=10, bold=True, color="FFFFFF")
        ws_list['A1'].fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        ws_list['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws_list.merge_cells('A2:O2')
        filter_summary = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | Filtered: {total_users} users"
        if role: filter_summary += f" | Role: {role}"
        if search: filter_summary += f" | Search: '{search}'"
        ws_list['A2'].value = filter_summary
        ws_list['A2'].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "ID", "Username", "Full Name", "Email", "Phone", "Role",
            "Company", "Branch", "Last Login", "Date Joined",
            "Online", "Verified", "Active"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws_list.cell(row=4, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data rows
        row_num = 5
        for user in users:
            company_name = getattr(user.company, 'name', '—') if user.company else '—'
            branch_name = user.branch or '—'

            last_login = user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else "Never"
            date_joined = user.date_joined.strftime("%Y-%m-%d")

            online_text = "Yes" if user.is_online else "No"
            verified_text = "Yes" if user.is_verified else "No"
            active_text = "Yes" if user.is_active else "No"

            ws_list.append([
                user.id,
                user.username,
                user.get_full_name(),
                user.email,
                user.phone or "—",
                user.role.capitalize(),
                company_name,
                branch_name,
                last_login,
                date_joined,
                online_text,
                verified_text,
                active_text
            ])

            # Alternate row color
            if row_num % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws_list.cell(row=row_num, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            row_num += 1

        # Conditional formatting for status columns
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        last_row = row_num - 1
        ws_list.conditional_formatting.add(f'K5:K{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
        ws_list.conditional_formatting.add(f'K5:K{last_row}', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))
        ws_list.conditional_formatting.add(f'L5:L{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
        ws_list.conditional_formatting.add(f'L5:L{last_row}', CellIsRule(operator='equal', formula=['"No"'], fill=orange_fill))
        ws_list.conditional_formatting.add(f'M5:M{last_row}', CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
        ws_list.conditional_formatting.add(f'M5:M{last_row}', CellIsRule(operator='equal', formula=['"No"'], fill=red_fill))

        # Auto-size columns on list sheet
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(4, row_num):
                cell = ws_list.cell(row=row, column=col)
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_list.column_dimensions[column_letter].width = max_length + 4

        ws_list.freeze_panes = "A5"

        # ─── SHEET 2: Summary Statistics with Charts ────────────────────────────
        ws_summary = wb.create_sheet(title="Summary Statistics")

        # Title
        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'] = "Summary Statistics Dashboard"
        ws_summary['A1'].font = Font(size=18, bold=True, color="1F497D")
        ws_summary['A1'].alignment = Alignment(horizontal="center")

        ws_summary.merge_cells('A2:F2')
        ws_summary['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')} | Total Users: {total_users}"
        ws_summary['A2'].alignment = Alignment(horizontal="center")

        # Key Metrics (A4:B12)
        metrics = [
            ("Total Users", total_users),
            ("Online Users", total_online),
            ("Offline Users", total_offline),
            ("New Today", new_today),
            ("New This Month", new_month),
            ("Blocked/Inactive", blocked_users + inactive_users),
            ("Performance %", f"{performance_percentage}%"),
        ]

        for i, (label, value) in enumerate(metrics, start=4):
            ws_summary[f'A{i}'] = label
            ws_summary[f'A{i}'].font = Font(bold=True)
            ws_summary[f'B{i}'] = value
            ws_summary[f'B{i}'].alignment = Alignment(horizontal="right")

        # Users by Role Pie Chart Data (D4:E12)
        sorted_roles = sorted(users_by_role.items(), key=lambda x: x[1], reverse=True)
        top_roles = sorted_roles[:8]
        others_count = sum(count for _, count in sorted_roles[8:])

        ws_summary['D4'] = "Users by Role"
        ws_summary['D4'].font = Font(bold=True)
        row = 5
        for role, count in top_roles:
            ws_summary[f'D{row}'] = role.capitalize()
            ws_summary[f'E{row}'] = count
            row += 1
        if others_count > 0:
            ws_summary[f'D{row}'] = "Others"
            ws_summary[f'E{row}'] = others_count

        # ─── PIE CHART: Users by Role ───────────────────────────────────────────
        pie = PieChart()
        labels = Reference(ws_summary, min_col=4, min_row=5, max_row=row)
        data = Reference(ws_summary, min_col=5, min_row=5, max_row=row)
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = "Distribution by Role"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showVal = True
        pie.dataLabels.showPercent = True
        pie.legend.position = 'b'
        ws_summary.add_chart(pie, "G4")

        # ─── BAR CHART: Key Metrics ─────────────────────────────────────────────
        bar = BarChart()
        bar.title = "Key User Metrics"
        bar.y_axis.title = "Count"
        bar.x_axis.title = "Category"
        data_ref = Reference(ws_summary, min_col=2, min_row=4, max_row=10)
        cats = Reference(ws_summary, min_col=1, min_row=5, max_row=10)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats)
        bar.legend = None
        ws_summary.add_chart(bar, "A15")

        # ─── LINE CHART: New Users Trend (Last 30 Days) ─────────────────────────
        line = LineChart()
        line.title = "New Users - Last 30 Days"
        line.y_axis.title = "New Users"
        line.x_axis.title = "Date"

        # Write trend data to summary sheet
        ws_summary['G20'] = "Date"
        ws_summary['H20'] = "New Users"
        for i, (date, count) in enumerate(daily_new, start=21):
            ws_summary[f'G{i}'] = date
            ws_summary[f'H{i}'] = count

        line_dates = Reference(ws_summary, min_col=7, min_row=20, max_row=len(daily_new)+20)
        line_data = Reference(ws_summary, min_col=8, min_row=20, max_row=len(daily_new)+20)
        line.add_data(line_data, titles_from_data=False)
        line.set_categories(line_dates)
        ws_summary.add_chart(line, "J20")

        # Auto-size columns on summary sheet
        for col in ['A', 'B', 'D', 'E', 'G', 'H']:
            max_length = 0
            for row in range(1, ws_summary.max_row + 1):
                cell = ws_summary[f'{col}{row}']
                try:
                    if len(str(cell.value or '')) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws_summary.column_dimensions[col].width = max_length + 4

        # Final response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="high_prosper_users_report.xlsx"'
        wb.save(response)
        return response

class BlockedUserAnalyticsAPIView(APIView):
    """
    GET /api/v1/users/admin/blocked-analytics/

    Advanced analytics for blocked users:
    - Total blocks
    - Blocks by time periods (today, week, month)
    - Top blockers/blocked users
    - Block trends over time
    - Percentage of users involved in blocks
    - Recent blocks list
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=7)
        month_start = today_start - timedelta(days=30)

        # All blocks
        all_blocks = BlockedUser.objects.all()
        total_blocks = all_blocks.count()

        # Time-based stats
        blocks_today = all_blocks.filter(created_at__gte=today_start).count()
        blocks_week = all_blocks.filter(created_at__gte=week_start).count()
        blocks_month = all_blocks.filter(created_at__gte=month_start).count()

        # Top blockers (users who blocked the most)
        top_blockers = all_blocks.values('blocker__username').annotate(
            count=Count('id')
        ).order_by('-count')[:10]  # Top 10

        # Top blocked (users blocked by most people)
        top_blocked = all_blocks.values('blocked__username').annotate(
            count=Count('id')
        ).order_by('-count')[:10]

        # Block trends (daily counts last 30 days)
        trends = []
        for day in range(29, -1, -1):
            day_start = today_start - timedelta(days=day)
            day_end = day_start + timedelta(days=1)
            count = all_blocks.filter(created_at__range=(day_start, day_end)).count()
            trends.append({
                'date': day_start.strftime('%Y-%m-%d'),
                'blocks': count
            })

        # Percentages
        total_users = User.objects.count()
        users_who_blocked = all_blocks.values('blocker').distinct().count()
        users_blocked = all_blocks.values('blocked').distinct().count()
        blocking_percentage = round((users_who_blocked / total_users * 100) if total_users else 0, 1)
        blocked_percentage = round((users_blocked / total_users * 100) if total_users else 0, 1)

        # Recent blocks (last 20)
        recent_blocks = all_blocks.order_by('-created_at')[:20].values(
            'id',
            'blocker__username',
            'blocked__username',
            'created_at'
        )

        data = {
            "total_blocks": total_blocks,
            "blocks_today": blocks_today,
            "blocks_week": blocks_week,
            "blocks_month": blocks_month,
            "top_blockers": list(top_blockers),
            "top_blocked": list(top_blocked),
            "trends": trends,
            "blocking_percentage": blocking_percentage,
            "blocked_percentage": blocked_percentage,
            "recent_blocks": list(recent_blocks),
        }

        return Response(data)